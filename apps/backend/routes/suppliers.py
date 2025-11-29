"""
Suppliers API Routes
Manages supplier and product-supplier relationship endpoints
"""
from flask import Blueprint, request, jsonify
from sqlalchemy import or_, func, case
import sys
from pathlib import Path
import logging
import traceback

logger = logging.getLogger(__name__)

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from models.base import db
from utils.idempotency import idempotent
from utils.optimistic_locking import optimistic_lock, with_transaction

# Import models from new structure
from models.suppliers import Supplier, ProductSupplier
from models.inventory import Inventory
from models.purchase_invoice import PurchaseInvoice, PurchaseInvoiceItem, SuggestedSupplier
from models.user import User
import csv
import io
import json
import sqlite3
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None
from flask_jwt_extended import jwt_required, get_jwt_identity
from utils.idempotency import idempotent

suppliers_bp = Blueprint('suppliers', __name__)


@suppliers_bp.route('/api/suppliers/bulk_upload', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
def bulk_upload_suppliers():
    """Bulk upload suppliers via CSV/XLSX file. Returns summary {created, updated, errors}."""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part named "file" in request'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No selected file'}), 400

        filename = (file.filename or '').lower()
        raw = file.read()

        def _sanitize_cell(v):
            if v is None:
                return None
            if not isinstance(v, str):
                return v
            v = v.strip()
            if v.startswith(('=', '+', '-', '@')):
                return "'" + v
            return v

        # Parse XLSX or CSV
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                return jsonify({'success': False, 'error': 'Server missing openpyxl dependency; please install openpyxl to accept XLSX uploads'}), 500
            try:
                from io import BytesIO
                wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None:
                    return jsonify({'success': False, 'error': 'XLSX contains no sheets'}), 400
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row:
                    return jsonify({'success': False, 'error': 'XLSX first row empty'}), 400
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                rows = []
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        obj[h] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
                iterable = rows
            except Exception as e:
                return jsonify({'success': False, 'error': 'Failed to parse XLSX: ' + str(e)}), 500
        else:
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                text = raw.decode('utf-8', errors='replace')

            try:
                sample = text[:4096]
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ','

            reader_obj = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            try:
                fns = reader_obj.fieldnames or []
                if len(fns) == 1:
                    single = fns[0]
                    if ';' in single and delimiter != ';':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter=';')
                    elif '\t' in single and delimiter != '\t':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter='\t')
            except Exception:
                pass

            try:
                iterable = list(reader_obj)
            except Exception as e:
                return jsonify({'success': False, 'error': 'CSV parsing failed: ' + str(e)}), 400

        created = 0
        updated = 0
        errors = []
        row_num = 0

        for row in iterable:
            row_num += 1
            try:
                if isinstance(row, dict):
                    normalized_row = {k: _sanitize_cell(v) for k, v in row.items()}
                else:
                    normalized_row = row

                # Normalize keys to supplier model
                company_name = normalized_row.get('companyName') or normalized_row.get('company_name') or normalized_row.get('company')
                company_code = normalized_row.get('companyCode') or normalized_row.get('company_code')
                contact_person = normalized_row.get('contactPerson') or normalized_row.get('contact_person')
                phone = normalized_row.get('phone')
                email = normalized_row.get('email')
                city = normalized_row.get('city')
                country = normalized_row.get('country')
                is_active = normalized_row.get('isActive')

                if not company_name:
                    errors.append({'row': row_num, 'error': 'Missing companyName'})
                    continue

                # Find existing by company_code or company_name
                existing = None
                if company_code:
                    existing = Supplier.query.filter_by(company_code=company_code, tenant_id=user.tenant_id).one_or_none()
                if not existing:
                    existing = Supplier.query.filter_by(company_name=company_name, tenant_id=user.tenant_id).one_or_none()

                if existing:
                    # update allowed fields
                    updatable = {
                        'company_code': company_code,
                        'contact_person': contact_person,
                        'phone': phone,
                        'email': email,
                        'city': city,
                        'country': country,
                        'is_active': (is_active in (True, 'true', 'True', '1', 1)) if is_active is not None else None
                    }
                    for attr, val in updatable.items():
                        if val is not None:
                            setattr(existing, attr, val)
                    db.session.add(existing)
                    updated += 1
                else:
                    sup = Supplier(
                        company_name=company_name,
                        company_code=company_code,
                        tenant_id=user.tenant_id,
                        contact_person=contact_person,
                        phone=phone,
                        email=email,
                        city=city,
                        country=country,
                        is_active=(is_active in (True, 'true', 'True', '1', 1)) if is_active is not None else True
                    )
                    db.session.add(sup)
                    created += 1

                try:
                    db.session.flush()
                except Exception as e:
                    db.session.rollback()
                    errors.append({'row': row_num, 'error': str(e)})
                    continue

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
                continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'Commit failed: ' + str(e)}), 500

        # Log activity
        from app import log_activity
        log_activity(user.id, 'bulk_upload', 'supplier', None, {'created': created, 'updated': updated, 'errors': errors}, request, tenant_id=user.tenant_id)

        return jsonify({'success': True, 'created': created, 'updated': updated, 'errors': errors}), 200
    except sqlite3.OperationalError as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Database write failed: ' + str(e)}), 503
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================================
# SUPPLIER CRUD OPERATIONS
# ============================================================================

@suppliers_bp.route('/api/suppliers', methods=['GET'])
@jwt_required()
def get_suppliers():
    """Get all suppliers with optional filtering and pagination"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Query parameters
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 50, type=int)
        search = request.args.get('search', '')
        is_active_param = request.args.get('is_active', None)
        city = request.args.get('city', '')
        sort_by = request.args.get('sort_by', 'company_name')
        sort_order = request.args.get('sort_order', 'asc')
        
        # Base query
        query = Supplier.query.filter_by(tenant_id=user.tenant_id)
        
        # Filter by active status (only if specified)
        if is_active_param is not None:
            is_active = is_active_param.lower() == 'true'
            query = query.filter(Supplier.is_active == is_active)
        
        # City filter
        if city:
            query = query.filter(Supplier.city.ilike(f'%{city}%'))
        
        # Search filter - enhanced for autocomplete
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                or_(
                    Supplier.company_name.ilike(search_term),
                    Supplier.company_code.ilike(search_term),
                    Supplier.contact_person.ilike(search_term),
                    Supplier.email.ilike(search_term),
                    Supplier.phone.ilike(search_term)
                )
            )
            
            # For autocomplete, use simple ordering by company name
            query = query.order_by(Supplier.company_name.asc())
        else:
            # Default sorting when no search
            if hasattr(Supplier, sort_by):
                order_column = getattr(Supplier, sort_by)
                if sort_order == 'desc':
                    query = query.order_by(order_column.desc())
                else:
                    query = query.order_by(order_column.asc())
        
        # Pagination
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        return jsonify({
            'success': True,
            'data': [s.to_dict() for s in pagination.items],
            'meta': {
                'page': page,
                'perPage': per_page,
                'total': pagination.total,
                'totalPages': pagination.pages
            }
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@suppliers_bp.route('/api/suppliers/search', methods=['GET'])
@jwt_required()
def search_suppliers():
    """Fast supplier search for autocomplete"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        query_param = request.args.get('q', '').strip()
        limit = request.args.get('limit', 10, type=int)
        
        if not query_param or len(query_param) < 2:
            return jsonify({'suppliers': []}), 200
        
        # Fast search query - simplified for debugging
        search_term = f'%{query_param}%'
        suppliers = Supplier.query.filter(
            Supplier.tenant_id == user.tenant_id,
            Supplier.is_active == True,
            or_(
                Supplier.company_name.ilike(search_term),
                Supplier.company_code.ilike(search_term),
                Supplier.contact_person.ilike(search_term)
            )
        ).order_by(Supplier.company_name.asc()).limit(limit).all()
        
        return jsonify({
            'suppliers': [s.to_dict() for s in suppliers]
        }), 200
        
    except Exception as e:
        print(f"Search error: {str(e)}")  # Debug print
        return jsonify({'error': str(e)}), 500


@suppliers_bp.route('/api/suppliers/<int:supplier_id>', methods=['GET'])
@jwt_required()
def get_supplier(supplier_id):
    """Get a single supplier by ID"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        supplier = Supplier.query.get_or_404(supplier_id)
        
        if supplier.tenant_id != user.tenant_id:
            return jsonify({'error': 'Supplier not found'}), 404
        
        # Include product relationships
        supplier_dict = supplier.to_dict()
        supplier_dict['products'] = [
            ps.to_dict(include_product=True) 
            for ps in supplier.products 
            if ps.is_active
        ]
        
        return jsonify(supplier_dict), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@suppliers_bp.route('/api/suppliers', methods=['POST'])
@jwt_required()
def create_supplier():
    """Create a new supplier"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        data = request.get_json() or {}
        # Debug: log incoming payload keys for diagnosis
        try:
            print(f"DEBUG create_supplier payload keys: {list(data.keys())}")
        except Exception:
            print("DEBUG create_supplier: unable to read payload keys")

        # Accept both snake_case and camelCase keys from clients
        def _get(d, snake, camel=None, default=None):
            if snake in d and d[snake] is not None:
                return d[snake]
            if camel and camel in d and d[camel] is not None:
                return d[camel]
            return default

        company_name = _get(data, 'company_name', 'companyName')
        company_code = _get(data, 'company_code', 'companyCode')
        # Normalize empty company_code to None to avoid UNIQUE '' collisions
        if company_code is not None:
            company_code = company_code.strip() if isinstance(company_code, str) else company_code
            if company_code == '':
                company_code = None
        tax_number = data.get('taxNumber')
        tax_office = data.get('taxOffice')
        contact_person = data.get('contactPerson')
        email = data.get('email')
        phone = data.get('phone')
        mobile = data.get('mobile')
        fax = data.get('fax')
        website = data.get('website')
        address = data.get('address')
        # Normalize address to a single word (take first token)
        if address:
            try:
                address = str(address).strip().split()[0]
            except Exception:
                address = str(address).strip()
        city = data.get('city')
        country = data.get('country')
        postal_code = data.get('postalCode')
        payment_terms = data.get('paymentTerms')
        currency = data.get('currency')
        rating = data.get('rating')
        notes = data.get('notes')
        is_active = data.get('isActive', True)

        # Validate required fields
        if not company_name:
            logger.error("❌ Company name is missing!")
            return jsonify({'error': 'Company name is required'}), 400

        # Check for duplicate company name
        existing = Supplier.query.filter_by(company_name=company_name, tenant_id=user.tenant_id).first()
        if existing:
            return jsonify({'error': 'Supplier with this company name already exists'}), 409
        
        # Create supplier
        supplier = Supplier(
            tenant_id=user.tenant_id,
            company_name=company_name,
            company_code=company_code,
            tax_number=tax_number,
            tax_office=tax_office,
            contact_person=contact_person,
            email=email,
            phone=phone,
            mobile=mobile,
            fax=fax,
            website=website,
            address=address,
            city=city,
            country=country or 'Türkiye',
            postal_code=postal_code,
            payment_terms=payment_terms,
            currency=currency or 'TRY',
            rating=rating,
            notes=notes,
            is_active=is_active
        )
        
        db.session.add(supplier)
        db.session.commit()
        
        return jsonify(supplier.to_dict()), 201
        
    except db.exc.IntegrityError as e:
        db.session.rollback()
        error_msg = str(e.orig) if hasattr(e, 'orig') else str(e)
        if 'UNIQUE constraint failed: suppliers.company_name' in error_msg:
            logger.warning(f"Duplicate supplier company_name attempt: {company_name}")
            return jsonify({
                'success': False,
                'error': 'Bu şirket adı zaten kullanılıyor. Lütfen farklı bir ad deneyin.'
            }), 409
        logger.error(f"Database integrity error: {error_msg}")
        return jsonify({'success': False, 'error': 'Veritabanı kısıtlaması hatası'}), 500
    except Exception as e:
        db.session.rollback()
        logger.error(f"Create supplier error: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@suppliers_bp.route('/api/suppliers/<int:supplier_id>', methods=['PUT'])
@jwt_required()
@idempotent(methods=['PUT'])
@optimistic_lock(Supplier, id_param='supplier_id')
@with_transaction
def update_supplier(supplier_id):
    """Update an existing supplier"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        supplier = Supplier.query.get_or_404(supplier_id)
        
        if supplier.tenant_id != user.tenant_id:
            return jsonify({'error': 'Supplier not found'}), 404

        data = request.get_json() or {}

        # Normalize incoming camelCase/snake_case keys
        def _get(d, snake, camel=None, default=None):
            if snake in d and d[snake] is not None:
                return d[snake]
            if camel and camel in d and d[camel] is not None:
                return d[camel]
            return default

        new_company_name = _get(data, 'company_name', 'companyName', supplier.company_name)
        if new_company_name != supplier.company_name:
            existing = Supplier.query.filter_by(company_name=new_company_name, tenant_id=user.tenant_id).first()
            if existing:
                return jsonify({'error': 'Supplier with this company name already exists'}), 409

        # Update fields (map common camelCase keys)
        updatable_fields = [
            ('company_name', 'companyName'), ('company_code', 'companyCode'), ('tax_number', 'taxNumber'), ('tax_office', 'taxOffice'),
            ('contact_person', 'contactPerson'), ('email', None), ('phone', None), ('mobile', None), ('fax', None), ('website', None),
            ('address', None), ('city', None), ('country', None), ('postal_code', 'postalCode'), ('payment_terms', 'paymentTerms'),
            ('currency', None), ('rating', None), ('notes', None), ('is_active', 'isActive')
        ]

        for snake, camel in updatable_fields:
            val = _get(data, snake, camel)
            if val is not None:
                # If updating address, ensure single-word constraint
                if snake == 'address' and val:
                    try:
                        val = str(val).strip().split()[0]
                    except Exception:
                        val = str(val).strip()
                setattr(supplier, snake, val)
        
        db.session.commit()
        
        return jsonify(supplier.to_dict()), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@suppliers_bp.route('/api/suppliers/<int:supplier_id>', methods=['DELETE'])
@jwt_required()
def delete_supplier(supplier_id):
    """Delete a supplier (soft delete by setting is_active=False)"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        supplier = Supplier.query.get_or_404(supplier_id)
        
        if supplier.tenant_id != user.tenant_id:
            return jsonify({'error': 'Supplier not found'}), 404
        
        # Soft delete
        supplier.is_active = False
        
        # Also deactivate all product relationships
        for ps in supplier.products:
            ps.is_active = False
        
        db.session.commit()
        
        return jsonify({'message': 'Supplier deactivated successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# PRODUCT-SUPPLIER RELATIONSHIP OPERATIONS
# ============================================================================

@suppliers_bp.route('/api/products/<product_id>/suppliers', methods=['GET'])
@jwt_required()
def get_product_suppliers(product_id):
    """Get all suppliers for a specific product"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Verify product exists
        product = Inventory.query.get_or_404(product_id)
        
        if product.tenant_id != user.tenant_id:
            return jsonify({'error': 'Product not found'}), 404
        
        # Get product-supplier relationships
        product_suppliers = ProductSupplier.query.filter_by(
            product_id=product_id,
            tenant_id=user.tenant_id,
            is_active=True
        ).order_by(ProductSupplier.priority.asc()).all()
        
        return jsonify({
            'product_id': product_id,
            'product_name': product.name,
            'suppliers': [ps.to_dict(include_supplier=True) for ps in product_suppliers]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@suppliers_bp.route('/api/suppliers/<int:supplier_id>/products', methods=['GET'])
@jwt_required()
def get_supplier_products(supplier_id):
    """Get all products for a specific supplier"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Verify supplier exists
        supplier = Supplier.query.get_or_404(supplier_id)
        
        if supplier.tenant_id != user.tenant_id:
            return jsonify({'error': 'Supplier not found'}), 404
        
        # Get product-supplier relationships
        product_suppliers = ProductSupplier.query.filter_by(
            supplier_id=supplier_id,
            tenant_id=user.tenant_id,
            is_active=True
        ).all()
        
        return jsonify({
            'supplier_id': supplier_id,
            'supplier_name': supplier.company_name,
            'products': [ps.to_dict(include_product=True) for ps in product_suppliers]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 404


@suppliers_bp.route('/api/products/<product_id>/suppliers', methods=['POST'])
@jwt_required()
def add_product_supplier(product_id):
    """Add a supplier to a product"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        data = request.get_json()
        
        # Validate required fields
        if not data.get('supplier_id'):
            return jsonify({'error': 'Supplier ID is required'}), 400
        
        supplier_id = data['supplier_id']
        
        # Verify product and supplier exist
        product = Inventory.query.get_or_404(product_id)
        if product.tenant_id != user.tenant_id:
            return jsonify({'error': 'Product not found'}), 404

        supplier = Supplier.query.get_or_404(supplier_id)
        if supplier.tenant_id != user.tenant_id:
            return jsonify({'error': 'Supplier not found'}), 404
        
        # Check if relationship already exists
        existing = ProductSupplier.query.filter_by(
            product_id=product_id,
            supplier_id=supplier_id,
            tenant_id=user.tenant_id
        ).first()
        
        if existing:
            if existing.is_active:
                return jsonify({'error': 'This supplier is already linked to this product'}), 409
            else:
                # Reactivate existing relationship
                existing.is_active = True
                ps = existing
        else:
            # Create new relationship
            ps = ProductSupplier(
                product_id=product_id,
                supplier_id=supplier_id,
                tenant_id=user.tenant_id
            )
            db.session.add(ps)
        
        # Update fields
        ps.supplier_product_code = data.get('supplier_product_code')
        ps.supplier_product_name = data.get('supplier_product_name')
        ps.unit_cost = data.get('unit_cost')
        ps.currency = data.get('currency', 'TRY')
        ps.minimum_order_quantity = data.get('minimum_order_quantity', 1)
        ps.lead_time_days = data.get('lead_time_days')
        ps.is_primary = data.get('is_primary', False)
        ps.priority = data.get('priority', 1)
        ps.notes = data.get('notes')
        
        # If this is marked as primary, unmark others
        if ps.is_primary:
            ProductSupplier.query.filter_by(
                product_id=product_id,
                tenant_id=user.tenant_id,
                is_primary=True
            ).filter(ProductSupplier.id != ps.id).update({'is_primary': False})
        
        db.session.commit()
        
        return jsonify(ps.to_dict(include_supplier=True)), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@suppliers_bp.route('/api/product-suppliers/<int:ps_id>', methods=['PUT'])
@jwt_required()
@idempotent(methods=['PUT'])
@optimistic_lock(ProductSupplier, id_param='ps_id')
@with_transaction
def update_product_supplier(ps_id):
    """Update a product-supplier relationship"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        ps = ProductSupplier.query.get_or_404(ps_id)
        
        if ps.tenant_id != user.tenant_id:
            return jsonify({'error': 'Product-supplier relationship not found'}), 404

        data = request.get_json()
        
        # Update fields
        updatable_fields = [
            'supplier_product_code', 'supplier_product_name', 'unit_cost',
            'currency', 'minimum_order_quantity', 'lead_time_days',
            'is_primary', 'priority', 'notes', 'is_active'
        ]
        
        for field in updatable_fields:
            if field in data:
                setattr(ps, field, data[field])
        
        # If this is marked as primary, unmark others
        if data.get('is_primary') and ps.is_primary:
            ProductSupplier.query.filter_by(
                product_id=ps.product_id,
                tenant_id=user.tenant_id,
                is_primary=True
            ).filter(ProductSupplier.id != ps.id).update({'is_primary': False})
        
        db.session.commit()
        
        return jsonify(ps.to_dict(include_supplier=True)), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


@suppliers_bp.route('/api/product-suppliers/<int:ps_id>', methods=['DELETE'])
@jwt_required()
def delete_product_supplier(ps_id):
    """Delete a product-supplier relationship"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        ps = ProductSupplier.query.get_or_404(ps_id)
        
        if ps.tenant_id != user.tenant_id:
            return jsonify({'error': 'Product-supplier relationship not found'}), 404
        
        # Soft delete
        ps.is_active = False
        
        db.session.commit()
        
        return jsonify({'message': 'Product-supplier relationship removed successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# STATISTICS & REPORTS
# ============================================================================

@suppliers_bp.route('/api/suppliers/stats', methods=['GET'])
@jwt_required()
def get_supplier_stats():
    """Get supplier statistics"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        total_suppliers = Supplier.query.filter_by(tenant_id=user.tenant_id).count()
        active_suppliers = Supplier.query.filter_by(is_active=True, tenant_id=user.tenant_id).count()
        inactive_suppliers = total_suppliers - active_suppliers
        
        # Total product relationships
        total_relationships = ProductSupplier.query.filter_by(is_active=True, tenant_id=user.tenant_id).count()
        
        # Top suppliers by product count
        top_suppliers = db.session.query(
            Supplier.id,
            Supplier.company_name,
            func.count(ProductSupplier.id).label('product_count')
        ).join(ProductSupplier).filter(
            Supplier.tenant_id == user.tenant_id,
            Supplier.is_active == True,
            ProductSupplier.is_active == True
        ).group_by(Supplier.id).order_by(
            func.count(ProductSupplier.id).desc()
        ).limit(5).all()
        
        # Calculate average rating
        avg_rating_result = db.session.query(
            func.avg(Supplier.rating)
        ).filter(
            Supplier.tenant_id == user.tenant_id,
            Supplier.is_active == True,
            Supplier.rating.isnot(None)
        ).scalar()
        
        avg_rating = float(avg_rating_result) if avg_rating_result else 0.0
        
        return jsonify({
            'total_suppliers': total_suppliers,
            'active_suppliers': active_suppliers,
            'inactive_suppliers': inactive_suppliers,
            'total_product_relationships': total_relationships,
            'average_rating': round(avg_rating, 1),
            'top_suppliers': [
                {
                    'id': s.id,
                    'name': s.company_name,
                    'product_count': s.product_count
                } for s in top_suppliers
            ]
        }), 200
        
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ============================================================================
# SUPPLIER INVOICE OPERATIONS
# ============================================================================

@suppliers_bp.route('/api/suppliers/<int:supplier_id>/invoices', methods=['GET'])
@jwt_required()
def get_supplier_invoices(supplier_id):
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        supplier = Supplier.query.get_or_404(supplier_id)
        if supplier.tenant_id != user.tenant_id:
            return jsonify({'error': 'Supplier not found'}), 404

        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        invoice_type = request.args.get('type', 'all')
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        # Include invoices explicitly linked to supplier OR whose sender_tax_number matches supplier.tax_number
        if supplier.tax_number:
            query = PurchaseInvoice.query.filter(
                PurchaseInvoice.tenant_id == user.tenant_id,
                or_(
                    PurchaseInvoice.supplier_id == supplier_id,
                    PurchaseInvoice.sender_tax_number == supplier.tax_number
                )
            )
        else:
            query = PurchaseInvoice.query.filter_by(supplier_id=supplier_id, tenant_id=user.tenant_id)

        if invoice_type != 'all':
            query = query.filter(PurchaseInvoice.invoice_type == invoice_type.upper())

        if start_date:
            try:
                from datetime import datetime
                start = datetime.fromisoformat(start_date)
                query = query.filter(PurchaseInvoice.invoice_date >= start)
            except (ValueError, TypeError):
                pass

        if end_date:
            try:
                from datetime import datetime
                end = datetime.fromisoformat(end_date)
                query = query.filter(PurchaseInvoice.invoice_date <= end)
            except (ValueError, TypeError):
                pass

        query = query.order_by(PurchaseInvoice.invoice_date.desc())
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)

        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': True,
            'data': {
                'invoices': [inv.to_dict() for inv in pagination.items],
                'pagination': {
                    'page': page,
                    'perPage': per_page,
                    'total': pagination.total,
                    'totalPages': pagination.pages
                }
            },
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 500


# ============================================================================
# SUGGESTED SUPPLIERS OPERATIONS
# ============================================================================

@suppliers_bp.route('/api/suppliers/suggested', methods=['GET'])
@jwt_required()
def get_suggested_suppliers():
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        suggested = SuggestedSupplier.query.filter_by(status='PENDING', tenant_id=user.tenant_id).order_by(
            SuggestedSupplier.last_invoice_date.desc()
        ).all()

        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': True,
            'data': {
                'suggestedSuppliers': [s.to_dict() for s in suggested]
            },
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 500


@suppliers_bp.route('/api/suppliers/suggested/<int:suggested_id>/accept', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
@with_transaction
def accept_suggested_supplier(suggested_id):
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        suggested = SuggestedSupplier.query.get_or_404(suggested_id)
        
        if suggested.tenant_id != user.tenant_id:
            return jsonify({'success': False, 'error': 'Suggested supplier not found'}), 404

        if suggested.status == 'ACCEPTED':
            from datetime import datetime
            request_id = request.headers.get('X-Request-ID')
            return jsonify({
                'success': False,
                'error': 'Supplier already accepted',
                'data': {
                    'supplier': suggested.supplier.to_dict() if suggested.supplier else None
                },
                'requestId': request_id,
                'timestamp': datetime.utcnow().isoformat()
            }), 400

        existing = Supplier.query.filter_by(tax_number=suggested.tax_number, tenant_id=user.tenant_id).first()
        if existing:
            suggested.status = 'ACCEPTED'
            suggested.supplier_id = existing.id
            suggested.accepted_at = datetime.utcnow()

            PurchaseInvoice.query.filter_by(
                sender_tax_number=suggested.tax_number,
                tenant_id=user.tenant_id
            ).update({
                'supplier_id': existing.id,
                'is_matched': True
            })

            db.session.commit()

            from datetime import datetime
            request_id = request.headers.get('X-Request-ID')
            return jsonify({
                'success': True,
                'data': {
                    'supplier': existing.to_dict(),
                    'note': 'Linked to existing supplier'
                },
                'requestId': request_id,
                'timestamp': datetime.utcnow().isoformat()
            }), 200

        supplier = Supplier(
            company_name=suggested.company_name,
            tax_number=suggested.tax_number,
            tax_office=suggested.tax_office,
            address=suggested.address,
            city=suggested.city,
            is_active=True
        )
        db.session.add(supplier)
        db.session.flush()

        suggested.status = 'ACCEPTED'
        suggested.supplier_id = supplier.id
        suggested.accepted_at = datetime.utcnow()

        PurchaseInvoice.query.filter_by(
            sender_tax_number=suggested.tax_number
        ).update({
            'supplier_id': supplier.id,
            'is_matched': True
        })

        db.session.commit()

        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': True,
            'data': {
                'supplier': supplier.to_dict()
            },
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 201
    except Exception as e:
        db.session.rollback()
        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 500


@suppliers_bp.route('/api/suppliers/suggested/<int:suggested_id>', methods=['DELETE'])
def reject_suggested_supplier(suggested_id):
    try:
        suggested = SuggestedSupplier.query.get_or_404(suggested_id)
        suggested.status = 'REJECTED'
        db.session.commit()

        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': True,
            'data': {
                'message': 'Suggested supplier rejected'
            },
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 200
    except Exception as e:
        db.session.rollback()
        from datetime import datetime
        request_id = request.headers.get('X-Request-ID')
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': request_id,
            'timestamp': datetime.utcnow().isoformat()
        }), 500
