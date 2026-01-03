"""
Invoices Bulk Operations
------------------------
Batch generate and bulk upload endpoints.
"""

from flask import request, jsonify
from sqlalchemy import desc
from datetime import datetime
from uuid import uuid4
import csv
import io
import sqlite3

from models.base import db
from models.invoice import Invoice
from models.sales import Sale
from . import invoices_bp
from utils.decorators import unified_access
from utils.idempotency import idempotent
import logging

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


def now_utc():
    """Return current UTC timestamp"""
    return datetime.now()


@invoices_bp.route('/invoices/batch-generate', methods=['POST'])
@unified_access(resource='invoices', action='write')
def batch_generate_invoices(ctx):
    """Generate multiple invoices at once - Unified Access"""
    try:
        data = request.get_json()
        
        if not data.get('saleIds') or not isinstance(data['saleIds'], list):
            return jsonify({
                'success': False,
                'error': 'saleIds array is required'
            }), 400
        
        sale_ids = data['saleIds']
        invoice_type = data.get('invoiceType', 'standard')
        customer_info = data.get('customerInfo', {})
        
        # Determine tenant_id
        tenant_id = ctx.tenant_id
        
        generated_invoices = []
        errors = []
        
        for sale_id in sale_ids:
            try:
                # Get sale with tenant check
                sale = db.session.get(Sale, sale_id)
                if not sale:
                    errors.append(f"Sale {sale_id} not found")
                    continue
                    
                # Tenant check
                if tenant_id and sale.tenant_id != tenant_id:
                    errors.append(f"Sale {sale_id} not found")
                    continue
                
                # Check if invoice already exists
                existing_invoice = Invoice.query.filter_by(sale_id=sale_id).first()
                if existing_invoice:
                    errors.append(f"Invoice already exists for sale {sale_id}")
                    continue
                
                # Generate invoice number
                year_month = datetime.utcnow().strftime('%Y%m')
                latest = Invoice.query.filter(
                    Invoice.invoice_number.like(f'INV{year_month}%')
                ).order_by(desc(Invoice.created_at)).first()
                
                if latest:
                    last_num = int(latest.invoice_number[-4:])
                    new_num = last_num + 1
                else:
                    new_num = 1
                
                invoice_number = f"INV{year_month}{new_num:04d}"
                
                # Create invoice
                invoice = Invoice(
                    tenant_id=sale.tenant_id,
                    branch_id=sale.branch_id,
                    invoice_number=invoice_number,
                    patient_id=sale.patient_id,
                    sale_id=sale_id,
                    device_price=sale.total_amount,
                    sgk_coverage=sale.sgk_coverage or 0,
                    patient_payment=sale.patient_payment or sale.total_amount,
                    invoice_type=invoice_type,
                    status='draft',
                    created_at=now_utc(),
                    updated_at=now_utc()
                )
                
                # Add custom customer info if provided
                if customer_info:
                    invoice.customer_name = customer_info.get('name', sale.patient.name if sale.patient else '')
                    invoice.customer_address = customer_info.get('address', '')
                    invoice.customer_tax_number = customer_info.get('taxNumber', '')
                
                db.session.add(invoice)
                db.session.flush()
                
                generated_invoices.append({
                    'id': invoice.id,
                    'invoiceNumber': invoice.invoice_number,
                    'saleId': sale_id,
                    'amount': invoice.device_price,
                    'status': invoice.status
                })
                
            except Exception as e:
                errors.append(f"Error generating invoice for sale {sale_id}: {str(e)}")
        
        db.session.commit()
        
        logger.info(f"Batch invoice generation: {len(generated_invoices)} created by {ctx.principal_id}")
        
        return jsonify({
            'success': True,
            'data': {
                'generatedInvoices': generated_invoices,
                'errors': errors,
                'summary': {
                    'total': len(sale_ids),
                    'successful': len(generated_invoices),
                    'failed': len(errors)
                }
            },
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logger.error(f"Batch generate invoices error: {e}", exc_info=True)
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@invoices_bp.route('/invoices/bulk_upload', methods=['POST'])
@unified_access(resource='invoices', action='write')
@idempotent(methods=['POST'])
def bulk_upload_invoices(ctx):
    """Bulk upload invoices from CSV/XLSX - Unified Access"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part named "file" in request'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No selected file'}), 400

        # Determine tenant_id
        tenant_id = ctx.tenant_id
        if not tenant_id and ctx.is_super_admin:
            tenant_id = request.form.get('tenant_id')
            if not tenant_id:
                return jsonify({'success': False, 'error': 'tenant_id is required for super admin operations'}), 400

        filename = (file.filename or '').lower()
        raw = file.read()

        def _sanitize_cell(v):
            if v is None:
                return None
            if not isinstance(v, str):
                return v
            v = v.strip()
            if v.startswith(('=', '+', '-', '@')):
                return "'" + v
            return v

        # Parse file based on extension
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                return jsonify({'success': False, 'error': 'Server missing openpyxl dependency'}), 500
            try:
                from io import BytesIO
                wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None:
                    return jsonify({'success': False, 'error': 'XLSX contains no sheets'}), 400
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row:
                    return jsonify({'success': False, 'error': 'XLSX first row empty'}), 400
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                rows = []
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        obj[h] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
                iterable = rows
            except Exception as e:
                return jsonify({'success': False, 'error': 'Failed to parse XLSX: ' + str(e)}), 500
        else:
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                text = raw.decode('utf-8', errors='replace')

            try:
                sample = text[:4096]
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ','

            reader_obj = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            try:
                fns = reader_obj.fieldnames or []
                if len(fns) == 1:
                    single = fns[0]
                    if ';' in single and delimiter != ';':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter=';')
                    elif '\t' in single and delimiter != '\t':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter='\t')
            except Exception:
                pass

            try:
                iterable = list(reader_obj)
            except Exception as e:
                return jsonify({'success': False, 'error': 'CSV parsing failed: ' + str(e)}), 400

        created = 0
        updated = 0
        errors = []
        row_num = 0

        for row in iterable:
            row_num += 1
            try:
                if isinstance(row, dict):
                    normalized_row = {k: _sanitize_cell(v) for k, v in row.items()}
                else:
                    normalized_row = row

                invoice_number = normalized_row.get('invoiceNumber') or normalized_row.get('invoice_number')
                patient_name = normalized_row.get('patientName') or normalized_row.get('patient_name')
                patient_phone = normalized_row.get('patientPhone') or normalized_row.get('patient_phone')
                patient_tc = normalized_row.get('patientTcNumber') or normalized_row.get('patient_tc')
                issue_date = normalized_row.get('issueDate') or normalized_row.get('issue_date')
                due_date = normalized_row.get('dueDate') or normalized_row.get('due_date')
                currency = normalized_row.get('currency')
                grand_total = normalized_row.get('grandTotal') or normalized_row.get('grand_total')

                if not (invoice_number or patient_name):
                    errors.append({'row': row_num, 'error': 'Missing invoiceNumber or patientName'})
                    continue

                existing = None
                if invoice_number:
                    existing = Invoice.query.filter_by(invoice_number=invoice_number, tenant_id=tenant_id).one_or_none()

                if existing:
                    if patient_name:
                        existing.patient_name = patient_name
                    if patient_phone:
                        existing.notes = (existing.notes or '') + f"\nTelefon: {patient_phone}"
                    if patient_tc:
                        existing.patient_tc = patient_tc
                    if issue_date:
                        existing.invoice_date = issue_date
                    if due_date:
                        existing.due_date = due_date
                    if currency:
                        existing.currency = currency
                    if grand_total is not None:
                        try:
                            existing.grand_total = float(grand_total)
                        except Exception:
                            pass
                    db.session.add(existing)
                    updated += 1
                else:
                    inv = Invoice(
                        tenant_id=tenant_id,
                        invoice_number=invoice_number,
                        patient_name=patient_name,
                        patient_tc=patient_tc,
                        invoice_date=issue_date,
                        due_date=due_date,
                        currency=currency,
                        grand_total=(float(grand_total) if grand_total is not None and str(grand_total) != '' else None),
                        status='draft'
                    )
                    if patient_phone:
                        inv.notes = (inv.notes or '') + f"Telefon: {patient_phone}"
                    db.session.add(inv)
                    created += 1

                try:
                    db.session.flush()
                except Exception as e:
                    db.session.rollback()
                    errors.append({'row': row_num, 'error': str(e)})
                    continue

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
                continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'Commit failed: ' + str(e)}), 500

        logger.info(f"Bulk upload: created={created}, updated={updated} by {ctx.principal_id}")

        return jsonify({'success': True, 'created': created, 'updated': updated, 'errors': errors}), 200
        
    except sqlite3.OperationalError as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Database write failed: ' + str(e)}), 503
    except Exception as e:
        db.session.rollback()
        logger.error(f"Bulk upload invoices error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500
