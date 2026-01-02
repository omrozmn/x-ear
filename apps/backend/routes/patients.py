from flask import Blueprint, request, jsonify, make_response
from models.base import db
from models.patient import Patient
from models.sales import Sale
from datetime import datetime
import json
import logging
import sqlite3
import csv
import io
import os
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None
from flask import Response
from flask_jwt_extended import jwt_required, get_jwt_identity
from models.user import User
from utils.idempotency import idempotent
from utils.optimistic_locking import optimistic_lock, with_transaction
# Permission checks handled by middleware - see middleware/permission_middleware.py
from utils.activity_logging import log_action, create_activity_log

patients_bp = Blueprint('patients', __name__)
logger = logging.getLogger(__name__)

def _is_admin_user(user_id):
    """Simple admin determination: checks if user's username or email matches configured admin values."""
    try:
        admin_username = os.getenv('ADMIN_USERNAME', 'admin')
        admin_email = os.getenv('ADMIN_EMAIL', 'admin@x-ear.com')
        user = db.session.get(User, user_id)
        if not user:
            return False
        return (user.username == admin_username) or (user.email == admin_email)
    except Exception:
        return False


@patients_bp.route('/patients/bulk_upload', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
def bulk_upload_patients():
    """Accept a multipart/form-data CSV file containing patients and upsert them into the DB.
    This endpoint is intentionally forgiving: it returns a summary of created/updated rows
    and reports per-row errors without aborting the entire batch.
    """
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part named "file" in request'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No selected file'}), 400

        # Read CSV content (support UTF-8 BOM)
        filename = (file.filename or '').lower()
        raw = file.read()

        # Quick debug endpoint to inspect incoming file metadata
        if request.args.get('debug') == '1':
            try:
                return jsonify({'filename': file.filename, 'size': len(raw), 'content_type': file.content_type}), 200
            except Exception:
                return jsonify({'filename': file.filename, 'size': None, 'content_type': file.content_type}), 200

        def _sanitize_cell(v):
            if v is None:
                return None
            if not isinstance(v, str):
                return v
            v = v.strip()
            # Prevent CSV/Excel formula injection
            if v.startswith(('=', '+', '-', '@')):
                return "'" + v
            return v

        rows_iter = None
        # If XLSX provided, parse using openpyxl into list of dicts
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                return jsonify({'success': False, 'error': 'Server missing openpyxl dependency; please install openpyxl to accept XLSX uploads'}), 500
            try:
                # load_workbook accepts a file-like object via BytesIO
                from io import BytesIO
                wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None:
                    return jsonify({'success': False, 'error': 'XLSX contains no sheets'}), 400
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row:
                    return jsonify({'success': False, 'error': 'XLSX first row empty'}), 400
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                # Basic header sanity: ensure at least one non-empty header
                if not any(h for h in headers):
                    return jsonify({'success': False, 'error': 'XLSX appears to have no headers; please provide a file with header row or use the importer mapping UI.'}), 400

                rows = []
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        key = h
                        val = r[idx] if idx < len(r) else None
                        obj[key] = _sanitize_cell(val)
                    # skip empty rows
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
                reader = rows
            except Exception as e:
                logger.exception('Failed to parse XLSX: %s', e)
                return jsonify({'success': False, 'error': 'Failed to parse XLSX file: ' + str(e)}), 500
        else:
            # treat as CSV/text; decode with utf-8-sig fallback
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                text = raw.decode('utf-8', errors='replace')

            # Try to detect delimiter using Sniffer; fallback to comma
            try:
                sample = text[:4096]
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ','

            # Use DictReader with detected delimiter
            reader_obj = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            # Heuristic fallback: if fieldnames look like a single joined header, try common separators
            try:
                fns = reader_obj.fieldnames or []
                if len(fns) == 1:
                    single = fns[0]
                    if ';' in single and delimiter != ';':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter=';')
                    elif '\t' in single and delimiter != '\t':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter='\t')
            except Exception:
                pass

            # Materialize rows (list) for safer inspection and later iteration
            try:
                rows_from_csv = list(reader_obj)
            except Exception as e:
                logger.exception('CSV parsing failed: %s', e)
                return jsonify({'success': False, 'error': 'CSV parsing failed: ' + str(e)}), 400

            # For downstream logic, use a list of dicts
            reader = rows_from_csv
            # expose detected fieldnames for debug if requested
            detected_fieldnames = reader_obj.fieldnames
        created = 0
        updated = 0
        errors = []
        row_num = 0

        # If reader is a list (from CSV materialization or XLSX parsing) iterate directly
        iterable = reader if isinstance(reader, list) else list(reader)

        # If debugging requested, return detected fieldnames and a sample
        if request.args.get('debug') == '1':
            sample = iterable[0] if len(iterable) > 0 else None
            return jsonify({'success': True, 'detected_fieldnames': (detected_fieldnames if 'detected_fieldnames' in locals() else headers if 'headers' in locals() else None), 'sample_first_row': sample}), 200

        for row in iterable:
            row_num += 1
            try:
                # If row came from csv.DictReader keys/values are strings; sanitize
                if isinstance(row, dict):
                    normalized_row = {k: _sanitize_cell(v) for k, v in row.items()}
                else:
                    normalized_row = row

                # Normalize row to expected API-style keys
                payload = {
                    'tcNumber': (normalized_row.get('tcNumber') if isinstance(normalized_row, dict) else None) or (normalized_row.get('tc_number') if isinstance(normalized_row, dict) else None) or (normalized_row.get('tc') if isinstance(normalized_row, dict) else None),
                    'identityNumber': (normalized_row.get('identityNumber') if isinstance(normalized_row, dict) else None) or (normalized_row.get('identity_number') if isinstance(normalized_row, dict) else None),
                    'firstName': (normalized_row.get('firstName') if isinstance(normalized_row, dict) else None) or (normalized_row.get('first_name') if isinstance(normalized_row, dict) else None) or (normalized_row.get('first') if isinstance(normalized_row, dict) else None),
                    'lastName': (normalized_row.get('lastName') if isinstance(normalized_row, dict) else None) or (normalized_row.get('last_name') if isinstance(normalized_row, dict) else None) or (normalized_row.get('last') if isinstance(normalized_row, dict) else None),
                    'phone': (normalized_row.get('phone') if isinstance(normalized_row, dict) else None) or (normalized_row.get('phone_number') if isinstance(normalized_row, dict) else None) or (normalized_row.get('tel') if isinstance(normalized_row, dict) else None),
                    'email': (normalized_row.get('email') if isinstance(normalized_row, dict) else None),
                    'birthDate': (normalized_row.get('birthDate') if isinstance(normalized_row, dict) else None) or (normalized_row.get('dob') if isinstance(normalized_row, dict) else None),
                    'gender': (normalized_row.get('gender') if isinstance(normalized_row, dict) else None),
                    'status': (normalized_row.get('status') if isinstance(normalized_row, dict) else None),
                    'segment': (normalized_row.get('segment') if isinstance(normalized_row, dict) else None)
                }

                # Address fields support flat CSV columns
                address = {}
                if row.get('address_city'):
                    address['city'] = row.get('address_city')
                if row.get('address_district'):
                    address['district'] = row.get('address_district')
                if row.get('address_full'):
                    address['fullAddress'] = row.get('address_full')
                if address:
                    payload['address'] = address

                # Tags may be provided as a comma/semicolon separated string
                tags_value = row.get('tags')
                if tags_value:
                    # split on comma or semicolon and strip
                    split_tags = [t.strip() for t in csv.reader([tags_value]).__next__() if t.strip()]
                    payload['tags'] = split_tags

                # Attempt to find existing patient by tc_number (strong identifier) within tenant
                existing = None
                if payload.get('tcNumber'):
                    existing = Patient.query.filter_by(tc_number=payload['tcNumber'], tenant_id=user.tenant_id).one_or_none()

                # Basic required-field guard: avoid DB integrity errors by validating
                if not payload.get('firstName') and not payload.get('phone') and not payload.get('tcNumber'):
                    errors.append({'row': row_num, 'error': 'Missing required identifying fields (tcNumber, firstName or phone)'} )
                    continue

                if existing:
                    # apply updatable fields similar to update endpoint
                    updatable = ['firstName','lastName','phone','email','birthDate','gender','status','segment','tags']
                    for k in updatable:
                        if payload.get(k) is not None:
                            if k == 'tags':
                                existing.tags = json.dumps(payload['tags'])
                            elif k == 'birthDate':
                                try:
                                    existing.birth_date = datetime.fromisoformat(payload['birthDate'])
                                except Exception:
                                    pass
                            else:
                                setattr(existing, {'firstName':'first_name','lastName':'last_name','phone':'phone','email':'email','gender':'gender','status':'status','segment':'segment'}[k], payload[k])
                    db.session.add(existing)
                    updated += 1
                else:
                    # create new patient
                    patient = Patient.from_dict(payload)
                    patient.tenant_id = user.tenant_id
                    db.session.add(patient)
                    created += 1

                # Flush per-row to capture integrity errors early without losing context
                try:
                    db.session.flush()
                except Exception as e:
                    # Capture rollback cause and continue
                    db.session.rollback()
                    errors.append({'row': row_num, 'error': str(e)})
                    continue

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
                continue

        # Commit everything that successfully flushed
        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'Commit failed: ' + str(e)}), 500

        # Log activity
        from app import log_activity
        actor = current_user_id
        log_activity(actor, 'bulk_upload', 'patient', None, {'created': created, 'updated': updated, 'errors': errors}, request)

        return jsonify({'success': True, 'created': created, 'updated': updated, 'errors': errors}), 200
    except sqlite3.OperationalError as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Database write failed: ' + str(e)}), 503
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@patients_bp.route('/patients/export', methods=['GET'])
@jwt_required()
def export_patients_csv():
    """Export patients as CSV.
    Supports optional query params: status, segment, q (search term).
    """
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        status = request.args.get('status')
        segment = request.args.get('segment')
        q = request.args.get('q')

        query = Patient.query.filter_by(tenant_id=user.tenant_id)
        if status:
            query = query.filter_by(status=status)
        if segment:
            query = query.filter_by(segment=segment)
        if q:
            search_filter = f"%{q}%"
            query = query.filter(db.or_(
                Patient.first_name.ilike(search_filter),
                Patient.last_name.ilike(search_filter),
                Patient.tc_number.ilike(search_filter),
                Patient.phone.ilike(search_filter),
                Patient.email.ilike(search_filter),
            ))

        patients = query.order_by(Patient.created_at.desc()).all()

        # Build CSV in memory
        output = io.StringIO()
        writer = csv.writer(output)
        # Header
        writer.writerow(['id','tcNumber','firstName','lastName','phone','email','birthDate','gender','status','segment','tags','createdAt'])
        for p in patients:
            writer.writerow([
                p.id,
                p.tc_number,
                p.first_name,
                p.last_name,
                p.phone,
                p.email,
                p.created_at.isoformat() if hasattr(p, 'created_at') and p.created_at else '',
                p.gender,
                p.status,
                p.segment,
                json.dumps(json.loads(p.tags) if p.tags else []),
                p.created_at.isoformat() if hasattr(p, 'created_at') and p.created_at else ''
            ])

        csv_bytes = output.getvalue().encode('utf-8')
        output.close()

        # Log export
        from app import log_activity        
        # Log activity
        log_activity(current_user_id or 'unknown', 'export', 'patient', None, {'count': len(patients)}, request)

        response = make_response(csv_bytes)
        response.headers.set('Content-Type', 'text/csv; charset=utf-8')
        response.headers.set('Content-Disposition', 'attachment', filename=f'patients_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv')
        return response
    except Exception as e:
        logger.exception('Export failed: %s', e)
        return jsonify({'success': False, 'error': str(e)}), 500


@patients_bp.route('/patients', methods=['GET'])
@jwt_required()
def list_patients():
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Support both offset-based and cursor-based pagination
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        cursor = request.args.get('cursor')  # For cursor-based pagination
        search = request.args.get('search', '')
        status = request.args.get('status', '')
        city = request.args.get('city', '')
        district = request.args.get('district', '')
        
        query = Patient.query.filter_by(tenant_id=user.tenant_id)

        # If user is admin (branch admin), filter by assigned branches
        if user.role == 'admin':
            user_branch_ids = [b.id for b in user.branches]
            if user_branch_ids:
                query = query.filter(Patient.branch_id.in_(user_branch_ids))
            else:
                # If admin has no branches assigned, they see nothing
                return jsonify({
                    'success': True,
                    'data': [],
                    'pagination': {
                        'page': page,
                        'perPage': per_page,
                        'total': 0,
                        'totalPages': 0,
                        'hasNext': False,
                        'nextCursor': None
                    }
                })
        
        # Apply search filter if provided
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                db.or_(
                    Patient.first_name.ilike(search_term),
                    Patient.last_name.ilike(search_term),
                    Patient.phone.ilike(search_term),
                    Patient.email.ilike(search_term),
                    Patient.id.ilike(search_term)
                )
            )
        
        # Apply status filter
        if status:
            query = query.filter(Patient.status == status)
        
        # Apply city filter
        if city:
            query = query.filter(Patient.address_city == city)
        
        # Apply district filter
        if district:
            query = query.filter(Patient.address_district == district)
        
        # Cursor-based pagination for better performance with large datasets
        if cursor:
            try:
                # Decode cursor (base64 encoded patient ID)
                import base64
                cursor_id = base64.b64decode(cursor.encode()).decode()
                query = query.filter(Patient.id > cursor_id)
            except Exception:
                # Invalid cursor, fall back to first page
                pass
        
        # Order by ID for consistent pagination
        query = query.order_by(Patient.id)
        
        # Limit results (add 1 to check if there are more pages)
        per_page = min(per_page, 200)  # Cap at 200 for performance
        patients_list = query.limit(per_page + 1).all()
        
        # Check if there are more results
        has_next = len(patients_list) > per_page
        if has_next:
            patients_list = patients_list[:-1]  # Remove the extra item
        
        # Generate next cursor if there are more results
        next_cursor = None
        if has_next and patients_list:
            import base64
            last_id = patients_list[-1].id
            next_cursor = base64.b64encode(str(last_id).encode()).decode()
        
        results = [p.to_dict() for p in patients_list]
        
        # For backward compatibility, also support offset-based pagination
        if not cursor:
            # Traditional pagination for UI compatibility
            total_query = Patient.query.filter_by(tenant_id=user.tenant_id)
            if search:
                search_term = f"%{search}%"
                total_query = total_query.filter(
                    db.or_(
                        Patient.first_name.ilike(search_term),
                        Patient.last_name.ilike(search_term),
                        Patient.phone.ilike(search_term),
                        Patient.email.ilike(search_term)
                    )
                )
            if status:
                total_query = total_query.filter(Patient.status == status)
            if city:
                total_query = total_query.filter(Patient.address_city == city)
            if district:
                total_query = total_query.filter(Patient.address_district == district)
            
            total_count = total_query.count()
            total_pages = (total_count + per_page - 1) // per_page
            
            return jsonify({
                'success': True,
                'data': results,
                'pagination': {
                    'page': page,
                    'perPage': per_page,
                    'total': total_count,
                    'totalPages': total_pages,
                    'hasNext': has_next,
                    'nextCursor': next_cursor
                }
            })
        else:
            # Cursor-based response
            return jsonify({
                'success': True,
                'data': results,
                'pagination': {
                    'perPage': per_page,
                    'hasNext': has_next,
                    'nextCursor': next_cursor,
                    'cursor': cursor
                }
            })
            
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e), 'timestamp': datetime.now().isoformat()}), 500


@patients_bp.route('/patients/<patient_id>', methods=['GET'])
@jwt_required()
def get_patient(patient_id):
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        patient = db.session.get(Patient, patient_id)
        if not patient or patient.tenant_id != user.tenant_id:
            return jsonify({'success': False, 'error': 'Patient not found'}), 404
        return jsonify({'success': True, 'data': patient.to_dict()})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@patients_bp.route('/patients', methods=['POST'])
@idempotent(methods=['POST'])
@log_action(action="patient.created", entity_id_from_response="data.id")
def create_patient():
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        logger.info(f'🔍 CREATE PATIENT - Received data: {data}')
        logger.info(f'🔍 CREATE PATIENT - city: {data.get("city")}, addressCity: {data.get("addressCity")}')
        logger.info(f'🔍 CREATE PATIENT - district: {data.get("district")}, addressDistrict: {data.get("addressDistrict")}')
        logger.info(f'🔍 CREATE PATIENT - address: {data.get("address")}')

        # Basic validation: ensure required fields exist to avoid DB integrity errors
        missing = []
        for field in ('firstName', 'lastName', 'phone'):
            if not data.get(field):
                missing.append(field)
        if missing:
            return jsonify({'success': False, 'error': f'Missing required fields: {",".join(missing)}'}), 400
        
        # Validate phone format
        phone = data.get('phone', '').strip()
        if phone:
            # Remove common phone separators
            phone_digits = ''.join(c for c in phone if c.isdigit())
            # Turkish phone numbers should be 10 or 11 digits (with or without country code)
            if len(phone_digits) < 10 or len(phone_digits) > 11:
                return jsonify({
                    'success': False, 
                    'error': 'Geçerli bir telefon numarası giriniz',
                    'details': {'phone': 'Geçerli bir telefon numarası giriniz'}
                }), 400

        # Validate TC number if provided
        tc_number = data.get('tcNumber')
        if tc_number and tc_number.strip():
            from utils.tc_validator import validate_tc_number
            is_valid, error_msg = validate_tc_number(tc_number)
            if not is_valid:
                return jsonify({'success': False, 'error': error_msg}), 400
            
            # Check if TC number already exists within tenant
            existing_tc = Patient.query.filter_by(tc_number=tc_number, tenant_id=user.tenant_id).first()
            if existing_tc:
                return jsonify({
                    'success': False, 
                    'error': 'Bu TC Kimlik No ile kayıtlı bir hasta zaten mevcut',
                    'existingPatient': {
                        'id': existing_tc.id,
                        'firstName': existing_tc.first_name,
                        'lastName': existing_tc.last_name,
                        'phone': existing_tc.phone
                    }
                }), 409

        # Check if patient with same phone number already exists
        phone = data.get('phone')
        if phone:
            existing_patient = Patient.query.filter_by(phone=phone, tenant_id=user.tenant_id).first()
            if existing_patient:
                return jsonify({
                    'success': False, 
                    'error': 'Bu telefon numarası ile kayıtlı bir hasta zaten mevcut',
                    'existingPatient': {
                        'id': existing_patient.id,
                        'firstName': existing_patient.first_name,
                        'lastName': existing_patient.last_name,
                        'phone': existing_patient.phone
                    }
                }), 409

        patient = Patient.from_dict(data)
        patient.tenant_id = user.tenant_id
        
        logger.info('🔍 CREATE PATIENT - After from_dict:')
        logger.info('🔍   address_full: %s', patient.address_full)
        logger.info('🔍   address_city: %s', patient.address_city)
        logger.info('🔍   address_district: %s', patient.address_district)
        
        db.session.add(patient)
        db.session.commit()

        from app import log_activity
        log_activity('system', 'create', 'patient', patient.id, {'payload': data}, request)

        resp = make_response(jsonify({'success': True, 'data': patient.to_dict()}), 201)
        resp.headers['Location'] = f"/api/patients/{patient.id}"
        return resp
    except Exception as e:
        # Rollback and provide more informative error messages in development
        db.session.rollback()
        logger.exception('Create patient failed')
        # If SQLAlchemy IntegrityError, surface as a 409 conflict
        from sqlalchemy.exc import IntegrityError
        if isinstance(e, IntegrityError):
            # Check if it's a TC number duplicate
            if 'tc_number' in str(e).lower():
                return jsonify({'success': False, 'error': 'Bu TC Kimlik No ile kayıtlı bir hasta zaten mevcut'}), 409
            return jsonify({'success': False, 'error': 'Veritabanı bütünlük hatası: Bu kayıt zaten mevcut'}), 409
        # Surface read-only DB errors with a distinct status so frontend can
        # display a clear message to operators. SQLite raises sqlite3.OperationalError
        # when the DB file is not writable (e.g. permissions or mounted read-only).
        if isinstance(e, sqlite3.OperationalError) and 'readonly' in str(e).lower():
            logger.exception('Database is read-only')
            return jsonify({'success': False, 'error': 'Database write failed: read-only database'}), 503
        return jsonify({'success': False, 'error': str(e), 'timestamp': datetime.now().isoformat()}), 500


@patients_bp.route('/patients/<patient_id>', methods=['PUT','PATCH'])
@idempotent(methods=['PUT', 'PATCH'])
@optimistic_lock(Patient, id_param='patient_id')
@with_transaction
@log_action(action="patient.updated", entity_id_param="patient_id")
def update_patient(patient_id):
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        data = request.get_json()
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        logger.info(f'🔍 UPDATE PATIENT - Patient ID: {patient_id}')
        logger.info(f'🔍 UPDATE PATIENT - Received data: {data}')
        logger.info(f'🔍 UPDATE PATIENT - branchId in data: {data.get("branchId")}')

        patient = db.session.get(Patient, patient_id)
        if not patient or patient.tenant_id != user.tenant_id:
            return jsonify({'success': False, 'error': 'Patient not found'}), 404

        logger.info(f'🔍 UPDATE PATIENT - Current branch_id: {patient.branch_id}')

        # Normalize keys: accept camelCase incoming fields and map to model attributes
        key_map = {
            'firstName': 'first_name',
            'lastName': 'last_name',
            'tcNumber': 'tc_number',
            'identityNumber': 'identity_number',
            'birthDate': 'birth_date',
            'gender': 'gender',
            'address': 'address',
            'status': 'status',
            'segment': 'segment',
            'acquisitionType': 'acquisition_type',
            'conversionStep': 'conversion_step',
            'referredBy': 'referred_by',
            'priorityScore': 'priority_score',
            'branchId': 'branch_id',
            'tags': 'tags',
            'sgkInfo': 'sgk_info',
            'phone': 'phone',
            'email': 'email',
            'city': 'address_city',
            'district': 'address_district',
            'addressCity': 'address_city',
            'addressDistrict': 'address_district'
        }

        # Whitelist of allowed direct-mapped model attributes (snake_case)
        allowed_attrs = {'first_name','last_name','tc_number','identity_number','phone','email','gender','status','segment','acquisition_type','conversion_step','referred_by','priority_score','branch_id','address_city','address_district','address_full'}

        for k, v in data.items():
            # Special fields handled explicitly
            if k == 'tags':
                patient.tags = json.dumps(v) if v is not None else json.dumps([])
                continue
            if k == 'sgkInfo':
                patient.sgk_info = json.dumps(v) if v is not None else json.dumps({
                    'rightEarDevice': 'available',
                    'leftEarDevice': 'available',
                    'rightEarBattery': 'available',
                    'leftEarBattery': 'available'
                })
                continue
            if k == 'birthDate' and v:
                try:
                    patient.birth_date = datetime.fromisoformat(v)
                except (ValueError, TypeError):
                    pass
                continue
            if k == 'address' or k == 'addressFull':
                logger.info('🔍 UPDATE - address field (%s): type=%s, value=%s', k, type(v).__name__, v)
                if isinstance(v, dict):
                    # Legacy dict format
                    patient.address_city = v.get('city')
                    patient.address_district = v.get('district')
                    patient.address_full = v.get('fullAddress') or v.get('address')
                    logger.info('🔍 UPDATE - address dict processed: city=%s, district=%s, address_full=%s', 
                              patient.address_city, patient.address_district, patient.address_full)
                elif isinstance(v, str):
                    # New string format - just store the address text
                    patient.address_full = v
                    logger.info('🔍 UPDATE - address string set: %s', patient.address_full)
                continue

            # Normalized key to model attribute
            normalized = key_map.get(k, k)

            # If normalized is a direct attribute, set it
            if normalized in allowed_attrs and hasattr(patient, normalized):
                try:
                    logger.info(f'🔍 Setting {normalized} = {v}')
                    setattr(patient, normalized, v)
                except Exception as e:
                    logger.debug('Failed to set attribute %s on patient: %s', normalized, e)
                continue

            # Last resort: if the original key matches a model attr, set it
            if hasattr(patient, k):
                try:
                    setattr(patient, k, v)
                except Exception as e:
                    logger.debug('Failed to set attribute %s on patient: %s', k, e)
                continue

            # Unknown/unhandled fields
            logger.debug('Ignored unknown patient update field: %s', k)

        logger.info('🔍 UPDATE PATIENT - After processing:')
        logger.info('🔍   address_full: %s', patient.address_full)
        logger.info('🔍   address_city: %s', patient.address_city)
        logger.info('🔍   address_district: %s', patient.address_district)
        logger.info('🔍   branch_id: %s', patient.branch_id)
        
        db.session.commit()
        from app import log_activity
        
        # detailed change message
        changes_list = []
        for k, v in data.items():
             # skip complex objects or long fields
             if k in ('photo', 'documents', 'notes'): continue
             changes_list.append(f"{k}: {v}")
        
        change_msg = ", ".join(changes_list)
        if len(change_msg) > 100: change_msg = change_msg[:97] + "..."
        
        log_activity('system', 'update', 'patient', patient.id, 
                    {'changes': data}, 
                    request,
                    message=f"Hasta güncellendi: {change_msg}")
        
        return jsonify({'success': True, 'data': patient.to_dict()}), 200
    except Exception as e:
        db.session.rollback()
        # Surface read-only DB errors with a distinct status so frontend can
        # display a clear message to operators. SQLite raises sqlite3.OperationalError
        # when the DB file is not writable (e.g. permissions or mounted read-only).
        if isinstance(e, sqlite3.OperationalError) and 'readonly' in str(e).lower():
            logger.exception('Database is read-only')
            return jsonify({'success': False, 'error': 'Database write failed: read-only database'}), 503
        return jsonify({'success': False, 'error': str(e)}), 500


@patients_bp.route('/patients/<patient_id>', methods=['DELETE'])
@log_action(action="patient.deleted", entity_id_param="patient_id", is_critical=True)
def delete_patient(patient_id):
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        patient = db.session.get(Patient, patient_id)
        if not patient or patient.tenant_id != user.tenant_id:
            return jsonify({'success': False, 'error': 'Patient not found'}), 404
        db.session.delete(patient)
        db.session.commit()
        from app import log_activity
        log_activity('system', 'delete', 'patient', patient_id, {'tcNumber': patient.tc_number}, request)
        return jsonify({'success': True, 'data': {'message': 'Patient deleted successfully'}}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@patients_bp.route('/patients/search', methods=['GET'])
@jwt_required()
def search_patients():
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        search_term = request.args.get('q', '')
        status = request.args.get('status')
        segment = request.args.get('segment')
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))

        query = Patient.query.filter_by(tenant_id=user.tenant_id)

        # If user is admin (branch admin), filter by assigned branches
        if user.role == 'admin':
            user_branch_ids = [b.id for b in user.branches]
            if user_branch_ids:
                query = query.filter(Patient.branch_id.in_(user_branch_ids))
            else:
                return jsonify({
                    'success': True,
                    'data': [],
                    'pagination': {
                        'page': page,
                        'perPage': per_page,
                        'total': 0,
                        'totalPages': 0
                    }
                })
        if search_term:
            search_filter = f"%{search_term}%"
            query = query.filter(
                db.or_(
                    Patient.first_name.ilike(search_filter),
                    Patient.last_name.ilike(search_filter),
                    Patient.tc_number.ilike(search_filter),
                    Patient.phone.ilike(search_filter),
                    Patient.email.ilike(search_filter)
                )
            )
        if status:
            query = query.filter_by(status=status)
        if segment:
            query = query.filter_by(segment=segment)

        patients_paginated = query.order_by(Patient.created_at.desc()).paginate(page=page, per_page=per_page, error_out=False)
        results = [p.to_dict() for p in patients_paginated.items]

        return jsonify({
            'success': True,
            'data': results,
            'results': results,  # backward compat
            'meta': {
                'total': patients_paginated.total,
                'page': page,
                'perPage': per_page,
                'totalPages': patients_paginated.pages
            }
        }), 200
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@patients_bp.route('/patients/<patient_id>/devices', methods=['GET'])
@jwt_required()
def get_patient_devices(patient_id):
    """Get all devices assigned to a specific patient"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        from models.sales import DeviceAssignment
        from models.inventory import InventoryItem as Inventory
        
        # Get patient to verify existence
        patient = db.session.get(Patient, patient_id)
        if not patient or patient.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Patient not found',
                'timestamp': datetime.now().isoformat()
            }), 404
        
        # Get all device assignments for this patient
        assignments = DeviceAssignment.query.filter_by(patient_id=patient_id).all()
        
        devices_data = []
        for assignment in assignments:
            # Get inventory item details if available
            inventory_item = None
            if assignment.inventory_id:
                from models.inventory import InventoryItem
                inventory_item = db.session.get(InventoryItem, assignment.inventory_id)
            
            # Start with the model's to_dict output which contains most fields
            device_dict = assignment.to_dict()
            
            # Enhance with inventory details if linked
            if inventory_item:
                device_dict['brand'] = inventory_item.brand
                device_dict['model'] = inventory_item.model
                device_dict['deviceName'] = f"{inventory_item.brand} {inventory_item.model}"
                device_dict['category'] = inventory_item.category
                device_dict['barcode'] = inventory_item.barcode
            else:
                # Fallback if no inventory link (e.g., manual entry)
                device_dict['deviceName'] = f"{assignment.loaner_brand or 'Unknown'} {assignment.loaner_model or 'Device'}" if assignment.is_loaner else f"Device {assignment.device_id or ''}"
            
            # Explicitly ensure frontend-required fields are present
            # Note: assignment.to_dict() already includes keys like 'deliveryStatus', 'isLoaner'
            # but we allow overrides here if logic differs.
            
            device_dict['earSide'] = assignment.ear
            device_dict['status'] = 'assigned' # Base status
            
            # Map delivery status explicitly if not in to_dict (safety net)
            if 'deliveryStatus' not in device_dict:
                 device_dict['deliveryStatus'] = getattr(assignment, 'delivery_status', 'pending')

            # Ensure loaner fields are present
            if 'isLoaner' not in device_dict:
                device_dict['isLoaner'] = getattr(assignment, 'is_loaner', False)
            if 'loanerInventoryId' not in device_dict:
                device_dict['loanerInventoryId'] = getattr(assignment, 'loaner_inventory_id', None)
            
            # Ensure Sale ID is passed and fetch sale details
            device_dict['saleId'] = assignment.sale_id
            
            # Fetch related sale to get down payment info (source from PaymentRecord for accuracy)
            device_dict['downPayment'] = 0.0  # Default initialization
            
            if assignment.sale_id:
                sale = db.session.get(Sale, assignment.sale_id)
                if sale:
                    # Try to get explicit down payment record first
                    from models.sales import PaymentRecord
                    down_payment_record = PaymentRecord.query.filter_by(
                        sale_id=sale.id,
                        payment_type='down_payment'
                    ).first()
                    
                    if down_payment_record:
                        device_dict['downPayment'] = float(down_payment_record.amount)
                    else:
                        # Fallback to sale.paid_amount if no specific record (legacy data)
                        device_dict['downPayment'] = float(sale.paid_amount) if sale.paid_amount else 0.0

            # assignedDate
            device_dict['assignedDate'] = assignment.created_at.isoformat() if assignment.created_at else None

            # Pricing fields - Ensure they are float for JSON serialization
            try:
                device_dict['sgkReduction'] = float(assignment.sgk_support) if getattr(assignment, 'sgk_support', None) is not None else 0.0
                device_dict['patientPayment'] = float(assignment.net_payable) if getattr(assignment, 'net_payable', None) is not None else 0.0
                device_dict['salePrice'] = float(assignment.sale_price) if getattr(assignment, 'sale_price', None) is not None else 0.0
                device_dict['listPrice'] = float(assignment.list_price) if getattr(assignment, 'list_price', None) is not None else 0.0
                
                # Map sgkScheme to sgkSupportType for frontend compatibility
                device_dict['sgkSupportType'] = assignment.sgk_scheme
                
            except (ValueError, TypeError):
                logger.warning(f"Error converting prices for assignment {assignment.id}")

            devices_data.append(device_dict)
        
        return jsonify({
            'success': True,
            'data': devices_data,
            'meta': {
                'patientId': patient_id,
                'patientName': f"{patient.first_name} {patient.last_name}",
                'deviceCount': len(devices_data)
            },
            'timestamp': datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        logger.error(f"Get patient devices error: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500


@patients_bp.route('/patients/<patient_id>/sales', methods=['GET'])
def get_patient_sales(patient_id):
    """Get all sales for a specific patient"""
    try:
        from models.sales import Sale, DeviceAssignment, PaymentRecord
        
        # Get patient to verify existence
        patient = db.session.get(Patient, patient_id)
        if not patient:
            return jsonify({
                'success': False,
                'error': 'Patient not found',
                'timestamp': datetime.now().isoformat()
            }), 404
        
        # Get all sales for this patient
        sales = Sale.query.filter_by(patient_id=patient_id).order_by(Sale.sale_date.desc()).all()
        
        sales_data = []
        for sale in sales:
            sale_dict = sale.to_dict()
            
            # Add devices from assignments
            devices = []
            assignments = DeviceAssignment.query.filter_by(sale_id=sale.id).all()
            for assignment in assignments:
                # Get inventory item details if available
                inventory_item = None
                if assignment.inventory_id:
                    from models.inventory import InventoryItem as Inventory
                    inventory_item = db.session.get(Inventory, assignment.inventory_id)
                
                # Build device info with inventory details
                if inventory_item:
                    device_name = f"{inventory_item.brand} {inventory_item.model}"
                    brand = inventory_item.brand
                    model = inventory_item.model
                    barcode = inventory_item.barcode
                else:
                    # Fallback for manual/loaner devices
                    device_name = f"{assignment.loaner_brand or 'Unknown'} {assignment.loaner_model or 'Device'}".strip()
                    brand = assignment.loaner_brand or ''
                    model = assignment.loaner_model or ''
                    barcode = None
                
                device_info = {
                    'id': assignment.inventory_id or assignment.device_id,
                    'name': device_name,
                    'brand': brand,
                    'model': model,
                    'serialNumber': assignment.serial_number or assignment.serial_number_left or assignment.serial_number_right,
                    'barcode': barcode,
                    'ear': assignment.ear,
                    'listPrice': float(assignment.list_price) if assignment.list_price else None,
                    'salePrice': float(assignment.sale_price) if assignment.sale_price else None,
                    'sgkCoverageAmount': float(assignment.sgk_support) if assignment.sgk_support else 0.0,
                    'patientResponsibleAmount': float(assignment.net_payable) if assignment.net_payable else None
                }
                devices.append(device_info)
            
            sale_dict['devices'] = devices
            
            # Add payment records
            payment_records = []
            payments = PaymentRecord.query.filter_by(sale_id=sale.id).all()
            for payment in payments:
                payment_info = {
                    'id': payment.id,
                    'amount': float(payment.amount) if payment.amount else 0.0,
                    'paymentDate': payment.payment_date.isoformat() if payment.payment_date else None,
                    'paymentMethod': payment.payment_method,
                    'paymentType': payment.payment_type,
                    'status': 'paid',  # PaymentRecord existence implies paid
                    'referenceNumber': None,
                    'notes': None
                }
                payment_records.append(payment_info)
            
            sale_dict['paymentRecords'] = payment_records
            
            sales_data.append(sale_dict)
        
        return jsonify({
            'success': True,
            'data': sales_data,
            'meta': {
                'patientId': patient_id,
                'patientName': f"{patient.first_name} {patient.last_name}",
                'salesCount': len(sales_data)
            },
            'timestamp': datetime.now().isoformat()
        }), 200
        
    except Exception as e:
        logger.error(f"Get patient sales error: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500

@patients_bp.route('/patients/count', methods=['GET'])
@jwt_required()
def count_patients():
    """Count patients with optional filters for SMS campaigns.
    
    Query params:
    - status: Patient status filter (active, passive)
    - segment: Patient segment filter (lead, potential, etc.)
    - acquisitionType: How the patient was acquired (advertisement, referral, etc.)
    - branchId: Filter by specific branch
    - dateStart: Filter by acquisition date start (ISO format)
    - dateEnd: Filter by acquisition date end (ISO format)
    
    Returns count of matching patients with valid phone numbers.
    """
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        query = Patient.query.filter_by(tenant_id=user.tenant_id)
        
        # Only count patients with valid phone numbers (for SMS campaigns)
        query = query.filter(Patient.phone.isnot(None))
        query = query.filter(Patient.phone != '')

        # If user is admin (branch admin), filter by assigned branches
        if user.role == 'admin':
            user_branch_ids = [b.id for b in user.branches]
            if user_branch_ids:
                query = query.filter(Patient.branch_id.in_(user_branch_ids))
            else:
                return jsonify({
                    'success': True,
                    'data': {'count': 0}
                })

        # Apply filters
        status = request.args.get('status')
        if status:
            query = query.filter(Patient.status == status)

        segment = request.args.get('segment')
        if segment:
            query = query.filter(Patient.segment == segment)

        acquisition_type = request.args.get('acquisitionType')
        if acquisition_type:
            query = query.filter(Patient.acquisition_type == acquisition_type)

        branch_id = request.args.get('branchId')
        if branch_id:
            query = query.filter(Patient.branch_id == branch_id)

        date_start = request.args.get('dateStart')
        if date_start:
            try:
                start_date = datetime.fromisoformat(date_start.replace('Z', '+00:00'))
                query = query.filter(Patient.created_at >= start_date)
            except ValueError:
                pass

        date_end = request.args.get('dateEnd')
        if date_end:
            try:
                end_date = datetime.fromisoformat(date_end.replace('Z', '+00:00'))
                query = query.filter(Patient.created_at <= end_date)
            except ValueError:
                pass

        count = query.count()

        return jsonify({
            'success': True,
            'data': {
                'count': count
            }
        }), 200

    except Exception as e:
        logger.error(f"Count patients error: {str(e)}")
        return jsonify({
            'success': False,
            'error': str(e),
            'timestamp': datetime.now().isoformat()
        }), 500