# Inventory API Routes for X-Ear CRM
# RESTful API endpoints for inventory management

from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from models.base import db
from models.user import User
import sys
from pathlib import Path
from datetime import datetime, timezone
from utils.idempotency import idempotent
import csv
import io
try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None

def now_utc():
    """Return current UTC timestamp"""
    return datetime.now(timezone.utc)

# Add models directory to path to import inventory
from models.inventory import Inventory, UNIT_TYPES
from models.brand import Brand
from models.category import Category
from uuid import uuid4

inventory_bp = Blueprint('inventory', __name__, url_prefix='/api/inventory')


@inventory_bp.route('', methods=['GET'])
@jwt_required()
def get_all_inventory():
    """Get all inventory items with optional filtering and pagination"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Query parameters for filtering
        category = request.args.get('category')
        brand = request.args.get('brand')
        supplier = request.args.get('supplier')
        low_stock = request.args.get('low_stock', 'false').lower() == 'true'
        out_of_stock = request.args.get('out_of_stock', 'false').lower() == 'true'
        search = request.args.get('search', '').lower()
        
        # Pagination parameters
        page = int(request.args.get('page', 1))
        per_page = int(request.args.get('per_page', 20))
        
        # Validate pagination parameters
        if page < 1:
            page = 1
        if per_page < 1 or per_page > 100:
            per_page = 20
        
        # Build query
        query = Inventory.query.filter_by(tenant_id=user.tenant_id)

        # If user is admin (branch admin), filter by assigned branches
        if user.role == 'admin':
            user_branch_ids = [b.id for b in user.branches]
            if user_branch_ids:
                query = query.filter(Inventory.branch_id.in_(user_branch_ids))
            else:
                return jsonify({
                    'success': True,
                    'data': [],
                    'meta': {
                        'page': page,
                        'perPage': per_page,
                        'total': 0,
                        'totalPages': 0
                    }
                }), 200
        
        if category:
            query = query.filter_by(category=category)
        
        if brand:
            query = query.filter(Inventory.brand.ilike(f'%{brand}%'))
        
        if supplier:
            query = query.filter(Inventory.supplier.ilike(f'%{supplier}%'))
        
        if low_stock:
            query = query.filter(
                Inventory.available_inventory > 0,
                Inventory.available_inventory <= Inventory.reorder_level
            )
        
        if out_of_stock:
            query = query.filter(Inventory.available_inventory == 0)
        
        if search:
            query = query.filter(
                db.or_(
                    Inventory.name.ilike(f'%{search}%'),
                    Inventory.brand.ilike(f'%{search}%'),
                    Inventory.model.ilike(f'%{search}%')
                )
            )
        
        # Get total count before pagination
        total_count = query.count()
        
        # Apply pagination
        items = query.order_by(Inventory.name).offset((page - 1) * per_page).limit(per_page).all()
        
        # Calculate pagination metadata
        total_pages = (total_count + per_page - 1) // per_page
        has_next = page < total_pages
        has_prev = page > 1
        
        # Safely serialize items so a single bad row doesn't cause a 500
        data_list = []
        for item in items:
            try:
                data_list.append(item.to_dict())
            except Exception as _e:
                print(f"Warning: failed to serialize inventory item {getattr(item, 'id', None)}: {_e}")
                data_list.append({'id': getattr(item, 'id', None), 'name': getattr(item, 'name', None)})
        return jsonify({
            'success': True,
            'data': data_list,
            'meta': {
                'page': page,
                'perPage': per_page,
                'total': total_count,
                'totalPages': total_pages
            }
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/bulk_upload', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
def bulk_upload_inventory():
    """Bulk upload inventory items from CSV/XLSX. Returns {created, updated, errors}."""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part named "file" in request'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No selected file'}), 400

        filename = (file.filename or '').lower()
        raw = file.read()

        def _sanitize_cell(v):
            if v is None:
                return None
            if not isinstance(v, str):
                return v
            v = v.strip()
            if v.startswith(('=', '+', '-', '@')):
                return "'" + v
            return v

        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                return jsonify({'success': False, 'error': 'Server missing openpyxl dependency; please install openpyxl to accept XLSX uploads'}), 500
            try:
                from io import BytesIO
                wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None:
                    return jsonify({'success': False, 'error': 'XLSX contains no sheets'}), 400
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row:
                    return jsonify({'success': False, 'error': 'XLSX first row empty'}), 400
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                rows = []
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        obj[h] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
                iterable = rows
            except Exception as e:
                return jsonify({'success': False, 'error': 'Failed to parse XLSX: ' + str(e)}), 500
        else:
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                text = raw.decode('utf-8', errors='replace')

            try:
                sample = text[:4096]
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ','

            reader_obj = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            try:
                fns = reader_obj.fieldnames or []
                if len(fns) == 1:
                    single = fns[0]
                    if ';' in single and delimiter != ';':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter=';')
                    elif '\t' in single and delimiter != '\t':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter='\t')
            except Exception:
                pass

            try:
                iterable = list(reader_obj)
            except Exception as e:
                return jsonify({'success': False, 'error': 'CSV parsing failed: ' + str(e)}), 400

        created = 0
        updated = 0
        errors = []
        row_num = 0

        for row in iterable:
            row_num += 1
            try:
                if isinstance(row, dict):
                    normalized_row = {k: _sanitize_cell(v) for k, v in row.items()}
                else:
                    normalized_row = row

                # Map common headers to Inventory fields
                payload = {}
                # Prefer camelCase and snake_case mappings
                payload['name'] = normalized_row.get('name') or normalized_row.get('productName') or normalized_row.get('Ürün Adı') or normalized_row.get('product_name')
                payload['brand'] = normalized_row.get('brand') or normalized_row.get('Marka')
                payload['model'] = normalized_row.get('model') or normalized_row.get('Model')
                payload['category'] = normalized_row.get('category') or normalized_row.get('Kategori')
                payload['availableInventory'] = normalized_row.get('availableInventory') or normalized_row.get('available_inventory') or normalized_row.get('stock') or normalized_row.get('Stok')
                payload['price'] = normalized_row.get('price') or normalized_row.get('Fiyat')
                payload['barcode'] = normalized_row.get('barcode') or normalized_row.get('Barkod')
                payload['supplier'] = normalized_row.get('supplier') or normalized_row.get('Tedarikçi')
                payload['stockCode'] = normalized_row.get('stockCode') or normalized_row.get('stock_code')

                if not payload.get('name') and not payload.get('barcode'):
                    errors.append({'row': row_num, 'error': 'Missing name and barcode'})
                    continue

                # Try to find existing item by barcode first, then by name within tenant
                existing = None
                if payload.get('barcode'):
                    existing = Inventory.query.filter_by(barcode=payload.get('barcode'), tenant_id=user.tenant_id).one_or_none()
                if not existing and payload.get('name'):
                    existing = Inventory.query.filter(Inventory.name.ilike(payload.get('name')), Inventory.tenant_id == user.tenant_id).one_or_none()

                if existing:
                    # map fields onto existing
                    for k, v in payload.items():
                        if v is None:
                            continue
                        # convert inventory model expected fields
                        if k == 'availableInventory':
                            try:
                                existing.available_inventory = int(v)
                            except Exception:
                                pass
                        elif k == 'price':
                            try:
                                existing.price = float(v)
                            except Exception:
                                pass
                        elif k == 'stockCode':
                            existing.stock_code = v
                        else:
                            setattr(existing, k, v)
                    db.session.add(existing)
                    updated += 1
                else:
                    try:
                        # Convert keys expected by Inventory.from_dict
                        create_payload = {
                            'name': payload.get('name'),
                            'brand': payload.get('brand'),
                            'model': payload.get('model'),
                            'category': payload.get('category'),
                            'availableInventory': int(payload.get('availableInventory')) if payload.get('availableInventory') not in (None, '') else 0,
                            'price': float(payload.get('price')) if payload.get('price') not in (None, '') else 0,
                            'barcode': payload.get('barcode'),
                            'supplier': payload.get('supplier'),
                            'stockCode': payload.get('stockCode')
                        }
                        item = Inventory.from_dict(create_payload)
                        item.tenant_id = user.tenant_id
                        db.session.add(item)
                        created += 1
                    except Exception as e:
                        errors.append({'row': row_num, 'error': str(e)})
                        db.session.rollback()
                        continue

                try:
                    db.session.flush()
                except Exception as e:
                    db.session.rollback()
                    errors.append({'row': row_num, 'error': str(e)})
                    continue

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
                continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'Commit failed: ' + str(e)}), 500

        return jsonify({'success': True, 'created': created, 'updated': updated, 'errors': errors}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
    # Advanced search endpoint moved here: preserve original advanced_search behavior
@inventory_bp.route('/search', methods=['GET'])
@jwt_required()
def advanced_search():
    """Advanced product search with comprehensive filtering"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Search parameters
        q = request.args.get('q', '').strip()
        category = request.args.get('category')
        brand = request.args.get('brand')
        min_price = request.args.get('minPrice', type=float)
        max_price = request.args.get('maxPrice', type=float)
        in_stock = request.args.get('inStock', type=bool)
        low_stock = request.args.get('lowStock', 'false').lower() == 'true'
        features = request.args.get('features', '').split(',') if request.args.get('features') else []
        supplier = request.args.get('supplier')
        warranty_period = request.args.get('warrantyPeriod', type=int)
        
        # Sorting parameters
        sort_by = request.args.get('sortBy', 'name')
        sort_order = request.args.get('sortOrder', 'asc')
        
        # Pagination parameters
        page = int(request.args.get('page', 1))
        limit = int(request.args.get('limit', 20))
        
        # Validate parameters
        if page < 1:
            page = 1
        if limit < 1 or limit > 100:
            limit = 20
        
        valid_sort_fields = ['name', 'price', 'stock', 'brand', 'category', 'createdAt']
        if sort_by not in valid_sort_fields:
            sort_by = 'name'
        
        if sort_order not in ['asc', 'desc']:
            sort_order = 'asc'
        
        # Build query
        query = Inventory.query.filter_by(tenant_id=user.tenant_id)

        # If user is admin (branch admin), filter by assigned branches
        if user.role == 'admin':
            user_branch_ids = [b.id for b in user.branches]
            if user_branch_ids:
                query = query.filter(Inventory.branch_id.in_(user_branch_ids))
            else:
                return jsonify({
                    'success': True,
                    'data': {
                        'items': [],
                        'pagination': {
                            'page': page,
                            'limit': limit,
                            'total': 0,
                            'totalPages': 0
                        },
                        'filters': {}
                    },
                    'requestId': str(uuid4()),
                    'timestamp': now_utc().isoformat()
                }), 200
        
        # Text search across multiple fields
        if q:
            search_filter = db.or_(
                Inventory.name.ilike(f'%{q}%'),
                Inventory.brand.ilike(f'%{q}%'),
                Inventory.model.ilike(f'%{q}%'),
                Inventory.barcode.ilike(f'%{q}%'),
                Inventory.description.ilike(f'%{q}%')
            )
            query = query.filter(search_filter)
        
        # Category filter
        if category:
            query = query.filter(Inventory.category == category)
        
        # Brand filter
        if brand:
            query = query.filter(Inventory.brand == brand)
        
        # Price range filter
        if min_price is not None:
            query = query.filter(Inventory.price >= min_price)
        if max_price is not None:
            query = query.filter(Inventory.price <= max_price)
        
        # Stock filters
        if in_stock is not None:
            if in_stock:
                query = query.filter(Inventory.available_inventory > 0)
            else:
                query = query.filter(Inventory.available_inventory <= 0)
        
        if low_stock:
            query = query.filter(Inventory.available_inventory <= Inventory.reorder_level)
        
        # Features filter (assuming features are stored as JSON or comma-separated)
        if features and features[0]:  # Check if features list is not empty
            for feature in features:
                if feature.strip():
                    query = query.filter(Inventory.features.ilike(f'%{feature.strip()}%'))
        
        # Supplier filter
        if supplier:
            query = query.filter(Inventory.supplier == supplier)
        
        # Warranty period filter
        if warranty_period is not None:
            query = query.filter(Inventory.warranty_period >= warranty_period)
        
        # Apply sorting
        sort_column = getattr(Inventory, sort_by, Inventory.name)
        if sort_order == 'desc':
            query = query.order_by(sort_column.desc())
        else:
            query = query.order_by(sort_column.asc())
        
        # Get total count before pagination
        total_count = query.count()
        
        # Apply pagination
        offset = (page - 1) * limit
        items = query.offset(offset).limit(limit).all()
        
        # Calculate pagination metadata
        total_pages = (total_count + limit - 1) // limit
        
        # Get filter options for frontend
        categories = db.session.query(Inventory.category).filter_by(tenant_id=user.tenant_id).distinct().filter(
            Inventory.category.isnot(None)
        ).all()
        brands = db.session.query(Inventory.brand).filter_by(tenant_id=user.tenant_id).distinct().filter(
            Inventory.brand.isnot(None)
        ).all()
        suppliers = db.session.query(Inventory.supplier).filter_by(tenant_id=user.tenant_id).distinct().filter(
            Inventory.supplier.isnot(None)
        ).all()
        
        # Get price range
        price_stats = db.session.query(
            db.func.min(Inventory.price),
            db.func.max(Inventory.price)
        ).filter_by(tenant_id=user.tenant_id).first()
        
        # Safely serialize items for the response
        safe_items = []
        for item in items:
            try:
                safe_items.append(item.to_dict())
            except Exception as _e:
                print(f"Warning: failed to serialize inventory item {getattr(item, 'id', None)}: {_e}")
                safe_items.append({'id': getattr(item, 'id', None), 'name': getattr(item, 'name', None)})

        return jsonify({
            'success': True,
            'data': {
                'items': safe_items,
                'pagination': {
                    'page': page,
                    'limit': limit,
                    'total': total_count,
                    'totalPages': total_pages
                },
                'filters': {
                    'categories': [cat[0] for cat in categories if cat[0]],
                    'brands': [brand[0] for brand in brands if brand[0]],
                    'suppliers': [sup[0] for sup in suppliers if sup[0]],
                    'priceRange': {
                        'min': float(price_stats[0]) if price_stats[0] else 0,
                        'max': float(price_stats[1]) if price_stats[1] else 0
                    }
                }
            },
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/categories_old', methods=['GET'])
@jwt_required()
def get_categories_old():
    """Get all available product categories with counts"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Get categories with item counts
        categories = db.session.query(
            Inventory.category,
            db.func.count(Inventory.id).label('count')
        ).filter(
            Inventory.category.isnot(None),
            Inventory.tenant_id == user.tenant_id
        ).group_by(Inventory.category).all()
        
        category_list = []
        for category, count in categories:
            if category and category.strip():
                category_list.append({
                    'id': category.lower().replace(' ', '_'),
                    'name': category,
                    'count': count,
                    'description': f'{count} items in {category}'
                })
        
        return jsonify({
            'success': True,
            'data': category_list,
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/brands_old', methods=['GET'])
@jwt_required()
def get_brands_old():
    """Get all available product brands with counts and categories"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Get brands with item counts and associated categories
        brands_data = db.session.query(
            Inventory.brand,
            Inventory.category,
            db.func.count(Inventory.id).label('count')
        ).filter(
            Inventory.brand.isnot(None),
            Inventory.tenant_id == user.tenant_id
        ).group_by(Inventory.brand, Inventory.category).all()
        
        # Organize brands with their categories
        brands_dict = {}
        for brand, category, count in brands_data:
            if brand and brand.strip():
                if brand not in brands_dict:
                    brands_dict[brand] = {
                        'name': brand,
                        'count': 0,
                        'categories': []
                    }
                brands_dict[brand]['count'] += count
                if category and category not in brands_dict[brand]['categories']:
                    brands_dict[brand]['categories'].append(category)
        
        brand_list = list(brands_dict.values())
        
        return jsonify({
            'success': True,
            'data': brand_list,
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/features', methods=['GET'])
@jwt_required()
def get_features():
    """Get all available product features with counts"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Get all items with features
        items_with_features = db.session.query(Inventory.features, Inventory.category).filter(
            Inventory.features.isnot(None),
            Inventory.tenant_id == user.tenant_id
        ).all()
        
        features_dict = {}
        for features_str, category in items_with_features:
            if features_str and features_str.strip():
                try:
                    # Accept either JSON array or comma-separated string
                    if features_str.strip().startswith('['):
                        parsed = json.loads(features_str)
                        feature_list = [f.strip() for f in parsed if isinstance(f, str) and f.strip()]
                    else:
                        feature_list = [f.strip() for f in features_str.split(',') if f.strip()]
                except Exception:
                    # Skip malformed feature values
                    continue

                for feature in feature_list:
                    if feature not in features_dict:
                        features_dict[feature] = {
                            'name': feature,
                            'count': 0,
                            'category': category
                        }
                    features_dict[feature]['count'] += 1
        
        feature_list = list(features_dict.values())
        
        return jsonify({
            'success': True,
            'data': feature_list,
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/brands', methods=['GET'])
@jwt_required()
def get_brands():
    """Get available brands from both brands table and inventory"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Get brands from brands table
        brand_objects = Brand.query.all()
        brands_from_table = [b.name for b in brand_objects]
        
        # Get distinct brands from inventory (for backwards compatibility)
        inventory_brands = db.session.query(Inventory.brand).filter_by(tenant_id=user.tenant_id).distinct().filter(Inventory.brand.isnot(None)).all()
        brands_from_inventory = [str(brand[0]) for brand in inventory_brands if brand[0] and str(brand[0]).strip()]
        
        # Combine and deduplicate
        all_brands = list(set(brands_from_table + brands_from_inventory))
        all_brands.sort()
        
        return jsonify({
            'success': True,
            'data': {
                'brands': all_brands
            },
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/brands', methods=['POST'])
@idempotent(methods=['POST'])
def create_brand():
    """Create a new brand"""
    try:
        data = request.get_json()
        if not data or 'name' not in data:
            return jsonify({
                'success': False,
                'error': 'Brand name is required',
                'requestId': str(uuid4()),
                'timestamp': now_utc().isoformat()
            }), 400

        brand_name = data['name'].strip()
        if not brand_name:
            return jsonify({
                'success': False,
                'error': 'Brand name cannot be empty',
                'requestId': str(uuid4()),
                'timestamp': now_utc().isoformat()
            }), 400

        # Check if brand already exists in brands table
        existing_brand = Brand.query.filter_by(name=brand_name).first()
        
        if existing_brand:
            return jsonify({
                'success': False,
                'error': f'Brand "{brand_name}" already exists',
                'requestId': str(uuid4()),
                'timestamp': now_utc().isoformat()
            }), 409

        # Create new brand
        new_brand = Brand(name=brand_name)
        db.session.add(new_brand)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': {
                'brand': brand_name,
                'id': new_brand.id
            },
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/<item_id>', methods=['GET'])
@jwt_required()
def get_inventory_item(item_id):
    """Get a single inventory item by ID"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        item = db.session.get(Inventory, item_id)
        
        if not item or item.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Inventory item not found'
            }), 404
        
        return jsonify({
            'success': True,
            'data': item.to_dict()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
def create_inventory_item():
    """Create a new inventory item"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 400
        
        # Validate required fields
        required_fields = ['name', 'brand', 'category', 'price']
        for field in required_fields:
            if not data.get(field):
                return jsonify({
                    'success': False,
                    'error': f'Missing required field: {field}'
                }), 400
        
        # Check for duplicate barcode
        if data.get('barcode'):
            existing = Inventory.query.filter_by(barcode=data['barcode']).first()
            if existing:
                return jsonify({
                    'success': False,
                    'error': 'Barcode already exists'
                }), 400
        
        # Create inventory item
        item = Inventory.from_dict(data)
        item.tenant_id = user.tenant_id
        
        # Assign branch_id if provided
        if data.get('branchId'):
            item.branch_id = data.get('branchId')

        db.session.add(item)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': item.to_dict(),
            'message': 'Inventory item created successfully'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/<item_id>', methods=['PUT', 'PATCH'])
@jwt_required()
def update_inventory_item(item_id):
    """Update an existing inventory item"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        item = db.session.get(Inventory, item_id)
        
        if not item or item.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Inventory item not found'
            }), 404
        
        data = request.get_json()
        
        if not data:
            return jsonify({
                'success': False,
                'error': 'No data provided'
            }), 400
        
        # Check for duplicate barcode (if changed)
        if data.get('barcode') and data['barcode'] != item.barcode:
            existing = Inventory.query.filter_by(barcode=data['barcode']).first()
            if existing:
                return jsonify({
                    'success': False,
                    'error': 'Barcode already exists'
                }), 400
        
        # Update fields
        item.name = data.get('name', item.name)
        item.brand = data.get('brand', item.brand)
        item.model = data.get('model', item.model)
        item.category = data.get('category', item.category)
        item.barcode = data.get('barcode', item.barcode)
        item.stock_code = data.get('stockCode', item.stock_code)
        item.supplier = data.get('supplier', item.supplier)
        item.unit = data.get('unit', item.unit)
        item.description = data.get('description', item.description)
        item.price = float(data.get('price', item.price))
        item.cost = float(data.get('cost', item.cost)) if 'cost' in data else item.cost
        # VAT/KDV handling (support both vatRate and kdv payload fields)
        if 'vatRate' in data:
            # Treat empty strings/null as "no change" unless explicit kdv provided
            v = data.get('vatRate')
            if v is None or (isinstance(v, str) and str(v).strip() == ''):
                if 'kdv' in data:
                    try:
                        item.vat_rate = float(data.get('kdv'))
                    except Exception:
                        pass
                else:
                    # leave existing vat_rate untouched
                    pass
            else:
                try:
                    item.vat_rate = float(v)
                except Exception:
                    # fallback to previous value
                    pass
        elif 'kdv' in data:
            try:
                item.vat_rate = float(data.get('kdv'))
            except Exception:
                pass
        # Price/cost include flags (frontend toggles) — support camelCase and snake_case
        if 'priceIncludesKdv' in data:
            item.price_includes_kdv = bool(data.get('priceIncludesKdv'))
        elif 'price_includes_kdv' in data:
            item.price_includes_kdv = bool(data.get('price_includes_kdv'))
        if 'costIncludesKdv' in data:
            item.cost_includes_kdv = bool(data.get('costIncludesKdv'))
        elif 'cost_includes_kdv' in data:
            item.cost_includes_kdv = bool(data.get('cost_includes_kdv'))
        item.direction = data.get('direction') or data.get('ear', item.direction)
        item.ear = data.get('ear') or data.get('direction', item.ear)
        item.warranty = data.get('warranty', item.warranty)
        item.reorder_level = data.get('reorderLevel') or data.get('minInventory', item.reorder_level)
        
        # Update features if provided
        if 'features' in data:
            import json
            features = data.get('features', [])
            item.features = json.dumps(features) if features else None
        
        # Update inventory levels if provided
        if 'availableInventory' in data or 'inventory' in data:
            item.available_inventory = data.get('availableInventory') or data.get('inventory', item.available_inventory)
        
        if 'onTrial' in data:
            item.on_trial = data.get('onTrial', item.on_trial)
        
        # Update serial numbers if provided
        if 'availableSerials' in data:
            import json
            serials = data.get('availableSerials', [])
            item.available_serials = json.dumps(serials) if serials else None
        
        item.updated_at = now_utc()
        
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': item.to_dict(),
            'message': 'Inventory item updated successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/<item_id>', methods=['DELETE'])
@jwt_required()
def delete_inventory_item(item_id):
    """Delete an inventory item"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        item = db.session.get(Inventory, item_id)
        
        if not item or item.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Inventory item not found'
            }), 404
        
        # Import ProductSupplier model to handle related records
        from models.suppliers import ProductSupplier
        
        # Delete related product_suppliers records first
        related_suppliers = ProductSupplier.query.filter_by(product_id=item_id).all()
        for supplier_link in related_suppliers:
            db.session.delete(supplier_link)
        
        # Now delete the inventory item
        db.session.delete(item)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Inventory item deleted successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/<item_id>/activity', methods=['POST'])
@jwt_required()
def log_inventory_activity(item_id):
    """Log an activity for an inventory item"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        item = db.session.get(Inventory, item_id)
        
        if not item or item.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Inventory item not found'
            }), 404
        
        data = request.get_json() or {}
        
        # Store activity in a simple JSON format
        # For now, we'll use a simple list stored in a new column
        activity = {
            'id': data.get('id', f"activity_{uuid4().hex[:16]}"),
            'productId': item_id,
            'action': data.get('action'),
            'description': data.get('description'),
            'metadata': data.get('metadata', {}),
            'timestamp': data.get('timestamp', now_utc().isoformat()),
            'user': data.get('user', 'Admin User')
        }
        
        # Get existing activities
        import json
        activities = json.loads(item.activity_log) if hasattr(item, 'activity_log') and item.activity_log else []
        activities.insert(0, activity)  # Add to beginning
        activities = activities[:100]  # Keep only last 100
        
        # Store back
        if not hasattr(item, 'activity_log'):
            # Column doesn't exist yet, skip storage but return success
            return jsonify({
                'success': True,
                'data': activity,
                'message': 'Activity logged (in-memory only, column not yet migrated)'
            }), 201
        
        item.activity_log = json.dumps(activities)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': activity,
            'message': 'Activity logged successfully'
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/<item_id>/activity', methods=['GET'])
@jwt_required()
def get_inventory_activities(item_id):
    """Get activity log for an inventory item"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        item = db.session.get(Inventory, item_id)
        
        if not item or item.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Inventory item not found'
            }), 404
        
        import json
        activities = json.loads(item.activity_log) if hasattr(item, 'activity_log') and item.activity_log else []
        
        return jsonify({
            'success': True,
            'activities': activities
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/<item_id>/serials', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
def add_serial_numbers(item_id):
    """Add serial numbers to an inventory item"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        item = db.session.get(Inventory, item_id)
        
        if not item or item.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Inventory item not found'
            }), 404
        
        data = request.get_json()
        serials = data.get('serials', [])
        
        if not serials:
            return jsonify({
                'success': False,
                'error': 'No serial numbers provided'
            }), 400
        
        added_count = 0
        for serial in serials:
            if item.add_serial_number(serial):
                added_count += 1
        
        item.updated_at = now_utc()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': item.to_dict(),
            'message': f'{added_count} serial number(s) added successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/<item_id>/assign', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
def assign_to_patient(item_id):
    """Assign inventory item to a patient (reduces stock)"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        item = db.session.get(Inventory, item_id)
        
        if not item or item.tenant_id != user.tenant_id:
            return jsonify({
                'success': False,
                'error': 'Inventory item not found'
            }), 404
        
        data = request.get_json()
        patient_id = data.get('patientId')
        serial_number = data.get('serialNumber')
        quantity = data.get('quantity', 1)
        
        if not patient_id:
            return jsonify({
                'success': False,
                'error': 'Patient ID is required'
            }), 400
        
        # If serial number provided, remove it from available serials
        if serial_number:
            if not item.remove_serial_number(serial_number):
                return jsonify({
                    'success': False,
                    'error': 'Serial number not found in inventory'
                }), 400
        else:
            # Otherwise, just decrease inventory count
            if not item.update_inventory(-quantity):
                return jsonify({
                    'success': False,
                    'error': 'Insufficient inventory'
                }), 400
        
        item.updated_at = now_utc()
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': item.to_dict(),
            'message': 'Item assigned to patient successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/low-stock', methods=['GET'])
@jwt_required()
def get_low_stock_items():
    """Get all items with low stock levels"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        items = Inventory.query.filter(
            Inventory.available_inventory <= Inventory.reorder_level,
            Inventory.tenant_id == user.tenant_id
        ).order_by(Inventory.available_inventory).all()
        
        safe_items = []
        for item in items:
            try:
                safe_items.append(item.to_dict())
            except Exception as _e:
                print(f"Warning: failed to serialize inventory item {getattr(item, 'id', None)}: {_e}")
                safe_items.append({'id': getattr(item, 'id', None), 'name': getattr(item, 'name', None)})

        return jsonify({
            'success': True,
            'data': safe_items,
            'count': len(safe_items)
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@inventory_bp.route('/stats', methods=['GET'])
@jwt_required()
def get_inventory_stats():
    """Get inventory statistics"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        total_items = Inventory.query.filter_by(tenant_id=user.tenant_id).count()
        low_stock_count = Inventory.query.filter(
            Inventory.available_inventory <= Inventory.reorder_level,
            Inventory.tenant_id == user.tenant_id
        ).count()
        out_of_stock_count = Inventory.query.filter(
            Inventory.available_inventory == 0,
            Inventory.tenant_id == user.tenant_id
        ).count()
        # Total stock across all inventory items
        try:
            total_stock = int(db.session.query(db.func.coalesce(db.func.sum(Inventory.available_inventory), 0)).filter_by(tenant_id=user.tenant_id).scalar() or 0)
        except Exception:
            total_stock = 0
        
        # Calculate total value taking into account whether stored price includes KDV (VAT)
        # and the item's kdv_rate. If price includes KDV, use price as-is; otherwise
        # multiply by (1 + kdv_rate/100).
        try:
            # SQL expression: available_inventory * (CASE WHEN price_includes_kdv THEN price ELSE price * (1 + kdv_rate/100) END)
            total_value_expr = Inventory.available_inventory * db.case(
                (Inventory.price_includes_kdv == True, Inventory.price),
                else_=Inventory.price * (1 + (Inventory.kdv_rate / 100.0))
            )
            total_value = db.session.query(db.func.sum(total_value_expr)).filter_by(tenant_id=user.tenant_id).scalar() or 0
        except Exception:
            # Fallback to simple multiplication if DB backend can't handle the case expression
            total_value = db.session.query(
                db.func.sum(Inventory.available_inventory * Inventory.price)
            ).filter_by(tenant_id=user.tenant_id).scalar() or 0
        
        # Category breakdown
        try:
            # Use same VAT-aware expression for category breakdown values
            category_value_expr = Inventory.available_inventory * db.case(
                (Inventory.price_includes_kdv == True, Inventory.price),
                else_=Inventory.price * (1 + (Inventory.kdv_rate / 100.0))
            )
            category_stats = db.session.query(
                Inventory.category,
                db.func.count(Inventory.id).label('count'),
                db.func.sum(category_value_expr).label('value')
            ).filter_by(tenant_id=user.tenant_id).group_by(Inventory.category).all()
            
            category_breakdown = {}
            for category, count, value in category_stats:
                category_breakdown[category or 'Diğer'] = {
                    'count': count,
                    'value': float(value or 0)
                }
        except Exception as e:
            category_breakdown = {}
        
        # Brand breakdown
        try:
            # Use same VAT-aware expression for brand breakdown values
            brand_value_expr = Inventory.available_inventory * db.case(
                (Inventory.price_includes_kdv == True, Inventory.price),
                else_=Inventory.price * (1 + (Inventory.kdv_rate / 100.0))
            )
            brand_stats = db.session.query(
                Inventory.brand,
                db.func.count(Inventory.id).label('count'),
                db.func.sum(brand_value_expr).label('value')
            ).filter_by(tenant_id=user.tenant_id).group_by(Inventory.brand).order_by(db.func.count(Inventory.id).desc()).limit(10).all()
            
            brand_breakdown = {}
            for brand, count, value in brand_stats:
                brand_breakdown[brand or 'Bilinmeyen'] = {
                    'count': count,
                    'value': float(value or 0)
                }
        except Exception as e:
            brand_breakdown = {}
        
        return jsonify({
            'success': True,
            'data': {
                'totalItems': total_items,
                'lowStockCount': low_stock_count,
                'outOfStockCount': out_of_stock_count,
                'totalValue': float(total_value),
                # DB-level total count (alias for clarity); frontend should use this
                'dbTotalItems': total_items,
                # Total stock across all items
                'totalStock': total_stock,
                # Hearing aid specific KPIs
                'hearingAid': {
                    'count': Inventory.query.filter(Inventory.category == 'hearing_aid', Inventory.tenant_id == user.tenant_id).count(),
                    'stock': int(db.session.query(db.func.coalesce(db.func.sum(Inventory.available_inventory), 0)).filter(Inventory.category == 'hearing_aid', Inventory.tenant_id == user.tenant_id).scalar() or 0)
                },
                'categoryBreakdown': category_breakdown,
                'brandBreakdown': brand_breakdown
            },
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/categories', methods=['GET'])
@jwt_required()
def get_categories():
    """Get all unique inventory categories from both categories table and inventory"""
    try:
        current_user_id = get_jwt_identity()
        user = db.session.get(User, current_user_id)
        if not user:
            return jsonify({'success': False, 'error': 'User not found'}), 401

        # Get categories from categories table
        category_objects = Category.query.all()
        categories_from_table = [c.name for c in category_objects]
        
        # Get distinct categories from inventory items (for backwards compatibility)
        inventory_categories = db.session.query(Inventory.category).filter_by(tenant_id=user.tenant_id).distinct().filter(
            Inventory.category.isnot(None),
            Inventory.category != ''
        ).all()
        categories_from_inventory = [category[0] for category in inventory_categories if category[0]]
        
        # Combine and deduplicate
        all_categories = list(set(categories_from_table + categories_from_inventory))
        all_categories.sort()
        
        return jsonify({
            'success': True,
            'data': {
                'categories': all_categories
            },
            'meta': {
                'total': len(all_categories)
            },
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 200
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/categories', methods=['POST'])
@jwt_required()
@idempotent(methods=['POST'])
def create_category():
    """Create a new inventory category"""
    try:
        data = request.get_json()
        
        if not data or 'category' not in data:
            return jsonify({
                'success': False,
                'error': 'Category name is required',
                'requestId': str(uuid4()),
                'timestamp': now_utc().isoformat()
            }), 400
        
        category_name = data['category'].strip()
        
        if not category_name:
            return jsonify({
                'success': False,
                'error': 'Category name cannot be empty',
                'requestId': str(uuid4()),
                'timestamp': now_utc().isoformat()
            }), 400
        
        # Check if category already exists in categories table
        existing_category = Category.query.filter_by(name=category_name).first()
        
        if existing_category:
            return jsonify({
                'success': False,
                'error': f'Category "{category_name}" already exists',
                'requestId': str(uuid4()),
                'timestamp': now_utc().isoformat()
            }), 409
        
        # Create new category
        new_category = Category(name=category_name)
        db.session.add(new_category)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'data': {
                'category': category_name,
                'id': new_category.id
            },
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({
            'success': False,
            'error': str(e),
            'requestId': str(uuid4()),
            'timestamp': now_utc().isoformat()
        }), 500


@inventory_bp.route('/units', methods=['GET'])
@jwt_required()
def get_units():
    """Get all available unit types for inventory items"""
    try:
        return jsonify({
            'success': True,
            'data': UNIT_TYPES,
            'message': 'Unit types retrieved successfully'
        }), 200
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500
