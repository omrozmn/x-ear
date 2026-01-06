"""
Patients Bulk Operations
------------------------
Bulk upload patients from CSV/XLSX files.
Uses @unified_access for tenant scoping.
"""

from flask import request, jsonify
from models.base import db
from models.patient import Patient
import csv
import io
import logging
import sqlite3
import json
from datetime import datetime
from . import patients_bp
from utils.decorators import unified_access
from utils.idempotency import idempotent

logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except Exception:
    load_workbook = None


@patients_bp.route('/patients/bulk_upload', methods=['POST'])
@unified_access(resource='patients', action='write')
@idempotent(methods=['POST'])
def bulk_upload_patients(ctx):
    """Accept a multipart/form-data CSV file containing patients and upsert them into the DB - Unified Access."""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No file part named "file" in request'}), 400

        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No selected file'}), 400

        # Determine tenant_id
        tenant_id = ctx.tenant_id
        if not tenant_id:
            # Super admin must provide tenant_id
            tenant_id = request.form.get('tenant_id')
            if not tenant_id:
                return jsonify({'success': False, 'error': 'tenant_id is required for super admin operations'}), 400

        # Read CSV content (support UTF-8 BOM)
        filename = (file.filename or '').lower()
        raw = file.read()

        # Quick debug endpoint to inspect incoming file metadata
        if request.args.get('debug') == '1':
            try:
                return jsonify({'filename': file.filename, 'size': len(raw), 'content_type': file.content_type}), 200
            except Exception:
                return jsonify({'filename': file.filename, 'size': None, 'content_type': file.content_type}), 200

        def _sanitize_cell(v):
            if v is None:
                return None
            if not isinstance(v, str):
                return v
            v = v.strip()
            # Prevent CSV/Excel formula injection
            if v.startswith(('=', '+', '-', '@')):
                return "'" + v
            return v

        rows_iter = None
        # If XLSX provided, parse using openpyxl into list of dicts
        reader = None
        detected_fieldnames = None
        
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                return jsonify({'success': False, 'error': 'Server missing openpyxl dependency; please install openpyxl to accept XLSX uploads'}), 500
            try:
                from io import BytesIO
                wb = load_workbook(filename=BytesIO(raw), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None:
                    return jsonify({'success': False, 'error': 'XLSX contains no sheets'}), 400
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row:
                    return jsonify({'success': False, 'error': 'XLSX first row empty'}), 400
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                if not any(h for h in headers):
                    return jsonify({'success': False, 'error': 'XLSX appears to have no headers; please provide a file with header row or use the importer mapping UI.'}), 400

                rows = []
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        key = h
                        val = r[idx] if idx < len(r) else None
                        obj[key] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
                reader = rows
            except Exception as e:
                logger.exception('Failed to parse XLSX: %s', e)
                return jsonify({'success': False, 'error': 'Failed to parse XLSX file: ' + str(e)}), 500
        else:
            try:
                text = raw.decode('utf-8-sig')
            except Exception:
                text = raw.decode('utf-8', errors='replace')

            try:
                sample = text[:4096]
                dialect = csv.Sniffer().sniff(sample)
                delimiter = dialect.delimiter
            except Exception:
                delimiter = ','

            reader_obj = csv.DictReader(io.StringIO(text), delimiter=delimiter)
            # Heuristic fallback for header detection
            try:
                fns = reader_obj.fieldnames or []
                if len(fns) == 1:
                    single = fns[0]
                    if ';' in single and delimiter != ';':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter=';')
                    elif '\t' in single and delimiter != '\t':
                        reader_obj = csv.DictReader(io.StringIO(text), delimiter='\t')
            except Exception:
                pass

            try:
                rows_from_csv = list(reader_obj)
            except Exception as e:
                logger.exception('CSV parsing failed: %s', e)
                return jsonify({'success': False, 'error': 'CSV parsing failed: ' + str(e)}), 400

            reader = rows_from_csv
            detected_fieldnames = reader_obj.fieldnames
        
        created = 0
        updated = 0
        errors = []
        row_num = 0

        iterable = reader if isinstance(reader, list) else list(reader)

        if request.args.get('debug') == '1':
            sample = iterable[0] if len(iterable) > 0 else None
            return jsonify({'success': True, 'detected_fieldnames': (detected_fieldnames if detected_fieldnames else headers if 'headers' in locals() else None), 'sample_first_row': sample}), 200

        for row in iterable:
            row_num += 1
            try:
                if isinstance(row, dict):
                    normalized_row = {k: _sanitize_cell(v) for k, v in row.items()}
                else:
                    normalized_row = row

                payload = {
                    'tcNumber': (normalized_row.get('tcNumber') if isinstance(normalized_row, dict) else None) or (normalized_row.get('tc_number') if isinstance(normalized_row, dict) else None) or (normalized_row.get('tc') if isinstance(normalized_row, dict) else None),
                    'identityNumber': (normalized_row.get('identityNumber') if isinstance(normalized_row, dict) else None) or (normalized_row.get('identity_number') if isinstance(normalized_row, dict) else None),
                    'firstName': (normalized_row.get('firstName') if isinstance(normalized_row, dict) else None) or (normalized_row.get('first_name') if isinstance(normalized_row, dict) else None) or (normalized_row.get('first') if isinstance(normalized_row, dict) else None),
                    'lastName': (normalized_row.get('lastName') if isinstance(normalized_row, dict) else None) or (normalized_row.get('last_name') if isinstance(normalized_row, dict) else None) or (normalized_row.get('last') if isinstance(normalized_row, dict) else None),
                    'phone': (normalized_row.get('phone') if isinstance(normalized_row, dict) else None) or (normalized_row.get('phone_number') if isinstance(normalized_row, dict) else None) or (normalized_row.get('tel') if isinstance(normalized_row, dict) else None),
                    'email': (normalized_row.get('email') if isinstance(normalized_row, dict) else None),
                    'birthDate': (normalized_row.get('birthDate') if isinstance(normalized_row, dict) else None) or (normalized_row.get('dob') if isinstance(normalized_row, dict) else None),
                    'gender': (normalized_row.get('gender') if isinstance(normalized_row, dict) else None),
                    'status': (normalized_row.get('status') if isinstance(normalized_row, dict) else None),
                    'segment': (normalized_row.get('segment') if isinstance(normalized_row, dict) else None)
                }

                address = {}
                if row.get('address_city'): address['city'] = row.get('address_city')
                if row.get('address_district'): address['district'] = row.get('address_district')
                if row.get('address_full'): address['fullAddress'] = row.get('address_full')
                if address: payload['address'] = address

                tags_value = row.get('tags')
                if tags_value:
                    split_tags = [t.strip() for t in csv.reader([tags_value]).__next__() if t.strip()]
                    payload['tags'] = split_tags

                existing = None
                if payload.get('tcNumber'):
                    existing = Patient.query.filter_by(tc_number=payload['tcNumber'], tenant_id=tenant_id).one_or_none()

                if not payload.get('firstName') and not payload.get('phone') and not payload.get('tcNumber'):
                    errors.append({'row': row_num, 'error': 'Missing required identifying fields'})
                    continue

                if existing:
                    updatable = ['firstName','lastName','phone','email','birthDate','gender','status','segment','tags']
                    for k in updatable:
                        if payload.get(k) is not None:
                            if k == 'tags':
                                existing.tags = json.dumps(payload['tags'])
                            elif k == 'birthDate':
                                try:
                                    existing.birth_date = datetime.fromisoformat(payload['birthDate'])
                                except Exception:
                                    pass
                            else:
                                key_map = {'firstName':'first_name','lastName':'last_name','phone':'phone','email':'email','gender':'gender','status':'status','segment':'segment'}
                                setattr(existing, key_map[k], payload[k])
                    db.session.add(existing)
                    updated += 1
                else:
                    patient = Patient.from_dict(payload)
                    patient.tenant_id = tenant_id
                    db.session.add(patient)
                    created += 1

                try:
                    db.session.flush()
                except Exception as e:
                    db.session.rollback()
                    errors.append({'row': row_num, 'error': str(e)})
                    continue

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
                continue

        try:
            db.session.commit()
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': 'Commit failed: ' + str(e)}), 500

        from app import log_activity
        log_activity(ctx.principal_id, 'bulk_upload', 'patient', None, {'created': created, 'updated': updated, 'errors': errors}, request, tenant_id=tenant_id)

        logger.info(f"Bulk upload: created={created}, updated={updated} by {ctx.principal_id}")

        return jsonify({'success': True, 'created': created, 'updated': updated, 'errors': errors}), 200
    except sqlite3.OperationalError as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': 'Database write failed: ' + str(e)}), 503
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500
