from fastapi import APIRouter, Depends, HTTPException, status, Query, Request, UploadFile, File
from typing import List, Optional, Any, Dict
from datetime import datetime
from sqlalchemy import or_
from sqlalchemy.orm import Session, joinedload
import json
import base64
import csv
import io
from enum import Enum

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from schemas.patients import (
    PatientRead, PatientCreate, PatientUpdate, PatientSearchFilters, BulkUploadResponse
)
from schemas.base import ResponseEnvelope, ResponseMeta
from schemas.base import ApiError
from models.patient import Patient
from models.sales import Sale, DeviceAssignment, PaymentRecord
from models.inventory import InventoryItem

from database import get_db
from middleware.unified_access import UnifiedAccess, require_access

import logging

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Patients"])

# --- HELPERS ---

def get_patient_or_404(db_session: Session, patient_id: str, access: UnifiedAccess) -> Patient:
    patient = db_session.get(Patient, patient_id)
    if not patient:
            raise HTTPException(
                status_code=404,
                detail=ApiError(message="Patient not found", code="PATIENT_NOT_FOUND").model_dump(mode="json"),
            )
    if access.tenant_id and patient.tenant_id != access.tenant_id:
        raise HTTPException(
            status_code=404,
            detail=ApiError(message="Patient not found", code="PATIENT_NOT_FOUND").model_dump(mode="json"),
        ) # Hide cross-tenant
    return patient

# --- ROUTES ---

@router.get("/patients", operation_id="listPatients", response_model=ResponseEnvelope[List[PatientRead]])
def list_patients(
    page: int = 1,
    per_page: int = 20,
    search: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    city: Optional[str] = None,
    district: Optional[str] = None,
    cursor: Optional[str] = None,
    access: UnifiedAccess = Depends(require_access("patients.view")),
    db: Session = Depends(get_db)
):
    """List patients with filtering and pagination"""
    try:
        query = db.query(Patient).options(joinedload(Patient.branch))
        
        # Tenant Scope
        if access.tenant_id:
            query = query.filter_by(tenant_id=access.tenant_id)
            
        # Branch Logic (Legacy Admin restriction)
        if access.is_tenant_admin and access.user and access.user.role == 'admin':
             user_branch_ids = [b.id for b in getattr(access.user, 'branches', [])]
             if user_branch_ids:
                 query = query.filter(Patient.branch_id.in_(user_branch_ids))
             elif getattr(access.user, 'branches', []):
                  return ResponseEnvelope(data=[], meta=ResponseMeta(total=0, page=page, perPage=per_page, totalPages=0))
        
        # Search
        if search:
            search_term = f"%{search}%"
            query = query.filter(
                or_(
                    Patient.first_name.ilike(search_term),
                    Patient.last_name.ilike(search_term),
                    Patient.phone.ilike(search_term),
                    Patient.email.ilike(search_term),
                    Patient.tc_number.ilike(search_term)
                )
            )
            
        # Filters
        if status_filter:
            from models.enums import PatientStatus
            # Convert status filter to enum value if possible
            try:
                status_enum = PatientStatus.from_legacy(status_filter)
                # Use .value for safe comparison with SQLAlchemy
                query = query.filter(Patient.status == status_enum.value)
            except (ValueError, AttributeError):
                query = query.filter(Patient.status == status_filter)
        if city:
            query = query.filter(Patient.address_city == city)
        if district:
            query = query.filter(Patient.address_district == district)
            
        # Cursor Pagination Logic
        if cursor:
             try:
                 cursor_id = base64.b64decode(cursor.encode()).decode()
                 query = query.filter(Patient.id > cursor_id)
             except Exception:
                 pass
        
        # Count total (separate query)
        total = query.count() if not cursor else 0 # optimized, don't count on cursor pages usually
        
        # Paginate
        query = query.order_by(Patient.id)
        limit = min(per_page, 200)
        items = query.limit(limit + 1).all()
        
        has_next = len(items) > limit
        if has_next:
            items = items[:-1]
            
        # Next Cursor
        next_cursor = None
        if has_next and items:
            next_cursor = base64.b64encode(str(items[-1].id).encode()).decode()
            
        # Manual pagination meta calculation if not cursor
        total_pages = 0
        if not cursor and total > 0:
            total_pages = (total + per_page - 1) // per_page
            
        return ResponseEnvelope(
            data=[item.to_dict() for item in items],
            meta={
                "total": total,
                "page": page,
                "perPage": per_page,
                "totalPages": total_pages,
                "hasNext": has_next,
                "nextCursor": next_cursor
            }
        )
    except Exception as e:
        logger.error(f"List patients error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
# ... imports ...

@router.post("/patients", operation_id="createPatients", response_model=ResponseEnvelope[PatientRead], status_code=201)
def create_patient(
    patient_in: PatientCreate,
    access: UnifiedAccess = Depends(require_access("patients.create")),
    db: Session = Depends(get_db)
):
    """Create a new patient"""
    try:
        if not access.tenant_id:
              raise HTTPException(
                  status_code=400,
                  detail=ApiError(message="Tenant context required", code="TENANT_REQUIRED").model_dump(mode="json"),
              )

        # Create instance
        # Pydantic model dump -> kwargs
        # Note: We need to handle aliasing. patient_in has fields like 'firstName', 
        # but model expects 'first_name'. Pydantic v2 model_dump(by_alias=False) gives snake_case default names.
        data = patient_in.model_dump(exclude_unset=True)

        # Force conversion just in case
        if 'birth_date' in data and data['birth_date']:
             d = data['birth_date']
             from datetime import date, datetime, time
             if type(d) == date:  # Strict type check to catch pure date
                 data['birth_date'] = datetime.combine(d, time.min)
        
        # Manual overrides/handling
        
        # Manual overrides/handling
        tags = data.pop('tags', [])
        sgk_info = data.pop('sgk_info', {})
        
        # Handle address flat vs nested
        # Schema might have 'address' dict. Model expects address_city, etc.
        # But our Schema PatientCreate defines flat fields: address_city, address_district...
        # Wait, looked at schemas/patients.py, it uses AddressSchema for 'address' field?
        # Let's check schema again. Step 110 shows PatientCreate inherits PatientBase.
        # PatientBase has `first_name`, `last_name`, etc.
        # It has `address: Optional[AddressSchema]`.
        # Database has `address_city`, `address_district`, `address_full`.
        address_data = data.pop('address', None)
        
        if 'status' in data:
             from models.enums import PatientStatus as ModelPatientStatus
             # Convert schema status (lowercase) to model status (UPPERCASE)
             status_val = data['status']
             if isinstance(status_val, str) or isinstance(status_val, Enum):
                 # Handle Pydantic Enum or string
                 val_str = status_val.value if hasattr(status_val, 'value') else str(status_val)
                 data['status'] = ModelPatientStatus.from_legacy(val_str)

        # Handle Date -> DateTime conversion for SQLite
        if 'birth_date' in data and data['birth_date']:
            logger.info(f"Birthdate type before: {type(data['birth_date'])} value: {data['birth_date']}")
            from datetime import date as date_type, datetime as datetime_type
            if isinstance(data['birth_date'], date_type) and not isinstance(data['birth_date'], datetime_type):
                data['birth_date'] = datetime_type.combine(data['birth_date'], datetime_type.min.time())
            logger.info(f"Birthdate type after: {type(data['birth_date'])}")
        
        patient = Patient(tenant_id=access.tenant_id, **data)
        
        # Set JSON fields
        patient.tags_json = tags if tags else []
        patient.sgk_info_json = sgk_info if sgk_info else {'rightEarDevice': 'available', 'leftEarDevice': 'available'}
        
        # Set Address flattened
        if address_data:
            # address_data is a dict or object depending on model_dump
            patient.address_city = address_data.get('city')
            patient.address_district = address_data.get('district')
            patient.address_full = address_data.get('fullAddress') or address_data.get('address')
            
        # Defaults
        if not patient.status: 
             patient.status = 'ACTIVE'
        
        db.add(patient)
        db.commit()
        
        return ResponseEnvelope(data=patient.to_dict())
    except Exception as e:
        db.rollback()
        logger.error(f"Create patient error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/patients/export", operation_id="exportPatients")
def export_patients(
    q: Optional[str] = None,
    status: Optional[str] = None,
    segment: Optional[str] = None,
    access: UnifiedAccess = Depends(require_access("patients.export")),
    db: Session = Depends(get_db)
):
    """Export patients as CSV"""
    from fastapi.responses import StreamingResponse
    import io
    import csv
    from sqlalchemy import text # Import text for branch logic

    if not access.tenant_id:
         raise HTTPException(status_code=400, detail="Tenant context required")

    query = db.query(Patient).options(joinedload(Patient.branch))
    query = query.filter_by(tenant_id=access.tenant_id)
    
    # Branch Logic
    if access.is_tenant_admin and access.user and access.user.role == 'admin':
         user_branch_ids = [b.id for b in getattr(access.user, 'branches', [])]
         if user_branch_ids:
             query = query.filter(Patient.branch_id.in_(user_branch_ids))
         elif getattr(access.user, 'branches', []):
              # Empty list but restricted
              query = query.filter(text("1=0")) 

    if status:
        query = query.filter(Patient.status == status)
    if segment:
        query = query.filter(Patient.segment == segment)
    if q:
        search_term = f"%{q}%"
        query = query.filter(
            or_(
                Patient.first_name.ilike(search_term),
                Patient.last_name.ilike(search_term),
                Patient.phone.ilike(search_term),
                Patient.email.ilike(search_term),
                Patient.tc_number.ilike(search_term)
            )
        )
        
    patients = query.order_by(Patient.created_at.desc()).all()

    # Generator for streaming
    def iter_csv():
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['id','tcNumber','firstName','lastName','phone','email','birthDate','gender','status','segment','tags','createdAt'])
        output.seek(0)
        yield output.read()
        output.seek(0)
        output.truncate(0)

        for p in patients:
            tags_str = json.dumps(p.tags_json) if p.tags_json else "[]"
            # Handle potential None dates
            created = p.created_at.isoformat() if p.created_at else ''
            
            writer.writerow([
                str(p.id),
                str(p.tc_number or ''),
                str(p.first_name or ''),
                str(p.last_name or ''),
                str(p.phone or ''),
                str(p.email or ''),
                str(p.birth_date or ''),
                str(p.gender or ''),
                str(p.status or ''),
                str(p.segment or ''),
                tags_str,
                created
            ])
            output.seek(0)
            yield output.read()
            output.seek(0)
            output.truncate(0)
            
    filename = f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/patients/count", operation_id="listPatientCount")
def count_patients(
    access: UnifiedAccess = Depends(require_access("patients.view")),
    db: Session = Depends(get_db),
    status: Optional[str] = None,
    segment: Optional[str] = None
):
    """Count patients"""
    if not access.tenant_id:
        return {"data": {"count": 0}}
        
    query = db.query(Patient).filter_by(tenant_id=access.tenant_id)
    # Filter valid phone
    query = query.filter(Patient.phone.isnot(None))
    
    if status:
        query = query.filter(Patient.status == status)
    if segment:
        query = query.filter(Patient.segment == segment)
        
    return ResponseEnvelope(data={'count': query.count()})

@router.get("/patients/{patient_id}", operation_id="getPatient", response_model=ResponseEnvelope[PatientRead])
def get_patient(
    patient_id: str,
    access: UnifiedAccess = Depends(require_access("patients.view")),
    db: Session = Depends(get_db)
):
    """Get single patient"""
    patient = get_patient_or_404(db, patient_id, access)
    return ResponseEnvelope(data=patient.to_dict())

@router.put("/patients/{patient_id}", operation_id="updatePatient", response_model=ResponseEnvelope[PatientRead])
def update_patient(
    patient_id: str,
    patient_in: PatientUpdate,
    access: UnifiedAccess = Depends(require_access("patients.edit")),
    db: Session = Depends(get_db)
):
    """Update patient"""
    patient = get_patient_or_404(db, patient_id, access)
    
    data = patient_in.model_dump(exclude_unset=True)
    
    # Handle complex fields same as create
    if 'tags' in data:
        patient.tags_json = data.pop('tags')
    
    if 'sgk_info' in data:
        patient.sgk_info_json = data.pop('sgk_info')
        
    if 'address' in data:
        addr = data.pop('address')
        if addr:
            patient.address_city = addr.get('city')
            patient.address_district = addr.get('district')
            patient.address_full = addr.get('fullAddress') or addr.get('address')

    for k, v in data.items():
        if hasattr(patient, k):
            setattr(patient, k, v)
            
    try:
        db.commit()
        return ResponseEnvelope(data=patient.to_dict())
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/patients/{patient_id}", operation_id="deletePatient")
def delete_patient(
    patient_id: str,
    access: UnifiedAccess = Depends(require_access("patients.delete")),
    db: Session = Depends(get_db)
):
    """Delete patient"""
    patient = get_patient_or_404(db, patient_id, access)
    try:
        db.delete(patient)
        db.commit()
        return ResponseEnvelope(message="Patient deleted")
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/patients/bulk-upload", operation_id="bulkUploadPatients", response_model=ResponseEnvelope[BulkUploadResponse])
async def bulk_upload_patients(
    file: UploadFile = File(...),
    access: UnifiedAccess = Depends(require_access("patients.create")),
    db: Session = Depends(get_db)
):
    """Bulk upload patients from CSV or XLSX"""
    try:
        if not access.tenant_id:
            raise HTTPException(
                status_code=400,
                detail=ApiError(message="Tenant context required", code="TENANT_REQUIRED").model_dump(mode="json")
            )

        filename = (file.filename or '').lower()
        content = await file.read()
        
        rows = []
        headers = []

        def _sanitize_cell(v):
            if v is None:
                return None
            if not isinstance(v, str):
                return v
            v = v.strip()
            # Prevent CSV/Excel formula injection
            if v.startswith(('=', '+', '-', '@')):
                return "'" + v
            return v

        # Parse File
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                raise HTTPException(status_code=500, detail="Server missing openpyxl dependency")
            
            try:
                wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None:
                     raise HTTPException(status_code=400, detail="XLSX contains no sheets")
                
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row:
                     raise HTTPException(status_code=400, detail="XLSX first row empty")
                
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                if not any(h for h in headers):
                     raise HTTPException(status_code=400, detail="XLSX has no headers")
                     
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        obj[h] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
            except Exception as e:
                logger.error(f"XLSX parse error: {e}")
                raise HTTPException(status_code=400, detail=f"Failed to parse XLSX: {str(e)}")
        else:
            # CSV assumption
            try:
                # Try UTF-8-SIG then UTF-8
                try:
                    text_content = content.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text_content = content.decode('utf-8', errors='replace')
                
                # Sniffer
                try:
                    dialect = csv.Sniffer().sniff(text_content[:4096])
                    delimiter = dialect.delimiter
                except Exception:
                    delimiter = ','
                
                # Fallback check for delimiter
                first_line = text_content.split('\n')[0]
                if ';' in first_line and delimiter != ';':
                    delimiter = ';'
                
                reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
                headers = reader.fieldnames or []
                rows = [row for row in reader]
                
            except Exception as e:
                logger.error(f"CSV parse error: {e}")
                raise HTTPException(status_code=400, detail=f"Failed to parse CSV: {str(e)}")

        # Process Rows
        created = 0
        updated = 0
        errors = []
        row_num = 0

        for row in rows:
            row_num += 1
            try:
                if isinstance(row, dict):
                    normalized_row = {k: _sanitize_cell(v) for k, v in row.items()}
                else:
                    normalized_row = row
                
                # Normalize keys (handle variations)
                def get_val(keys):
                    for k in keys:
                        if k in normalized_row and normalized_row[k]:
                            return normalized_row[k]
                    return None
                    
                tc_number = get_val(['tcNumber', 'tc_number', 'tc', 'TC', 'tc_no'])
                first_name = get_val(['firstName', 'first_name', 'first', 'isim', 'ad'])
                last_name = get_val(['lastName', 'last_name', 'last', 'soyisim', 'soyad'])
                phone = get_val(['phone', 'phone_number', 'tel', 'telefon', 'cep_telefon', 'cep'])
                
                if not (first_name or phone or tc_number):
                    # Skip empty rows or not enough info
                    errors.append({'row': row_num, 'error': 'Missing required identifying fields'})
                    continue

                # Prepare Payload
                payload = {
                    'tc_number': tc_number,
                    'identity_number': get_val(['identityNumber', 'identity_number']),
                    'first_name': first_name,
                    'last_name': last_name,
                    'phone': phone,
                    'email': get_val(['email', 'e-mail', 'eposta']),
                    'birth_date': get_val(['birthDate', 'birth_date', 'dob', 'dogum_tarihi']),
                    'gender': get_val(['gender', 'cinsiyet']),
                    'status': get_val(['status', 'durum']),
                    'segment': get_val(['segment']),
                }
                
                # Clean up None values
                payload = {k: v for k, v in payload.items() if v is not None}
                
                # Address
                address_city = get_val(['address_city', 'city', 'il', 'sehir'])
                address_district = get_val(['address_district', 'district', 'ilce'])
                address_full = get_val(['address_full', 'fullAddress', 'address', 'adres'])
                
                if address_city: payload['address_city'] = address_city
                if address_district: payload['address_district'] = address_district
                if address_full: payload['address_full'] = address_full
                
                # Tags
                tags_val = get_val(['tags', 'etiketler'])
                if tags_val:
                    if isinstance(tags_val, str):
                        payload['tags_json'] = [t.strip() for t in tags_val.split(',') if t.strip()]
                    else:
                        payload['tags_json'] = tags_val

                # Check Existing
                existing = None
                if tc_number:
                    existing = db.query(Patient).filter(Patient.tc_number == tc_number, Patient.tenant_id == access.tenant_id).first()
                
                if existing:
                    # Update Logic
                    for k, v in payload.items():
                        if hasattr(existing, k) and v is not None:
                            # Handle Date parsing for birth_date if strictly string
                            if k == 'birth_date' and isinstance(v, str):
                                try:
                                    dt = None
                                    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y/%m/%d'):
                                        try:
                                            dt = datetime.strptime(v, fmt)
                                            break
                                        except ValueError:
                                            pass
                                    if dt:
                                         setattr(existing, k, dt)
                                    else:
                                         # Try isoformat
                                         setattr(existing, k, datetime.fromisoformat(v))
                                except:
                                    pass 
                            else:
                                setattr(existing, k, v)
                    updated += 1
                    db.add(existing)
                else:
                    # Create Logic
                    if not payload.get('first_name') or not payload.get('last_name'):
                         errors.append({'row': row_num, 'error': 'First Name and Last Name required for new patients'})
                         continue
                         
                    # Handle Date parsing
                    if 'birth_date' in payload and isinstance(payload['birth_date'], str):
                         v = payload['birth_date']
                         dt = None
                         for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y/%m/%d'):
                            try:
                                dt = datetime.strptime(v, fmt)
                                break
                            except ValueError:
                                pass
                         if dt:
                             payload['birth_date'] = dt
                         else:
                             try:
                                 payload['birth_date'] = datetime.fromisoformat(v)
                             except:
                                 del payload['birth_date']
                    
                    patient = Patient(tenant_id=access.tenant_id, **payload)
                    if not patient.status: patient.status = 'ACTIVE'
                    
                    db.add(patient)
                    created += 1
                
                # Use savepoint to prevent full rollback on single row failure
                db.begin_nested() 
                try:
                    db.flush()
                    db.commit() # Commit the savepoint
                except Exception as e:
                    db.rollback() # Rollback to savepoint
                    errors.append({'row': row_num, 'error': str(e)})

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        
        db.commit()
        return ResponseEnvelope(data={
            'success': True,
            'created': created,
            'updated': updated,
            'errors': errors
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Bulk upload patient error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

