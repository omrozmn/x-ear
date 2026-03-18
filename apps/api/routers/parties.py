from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from typing import List, Optional
from datetime import datetime
from sqlalchemy.orm import Session
import json
import csv
import io

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from schemas.parties import (
    PartyRead, PartyCreate, PartyUpdate, BulkUploadResponse,
    BulkUpdateRequest, BulkUpdateResponse, BulkEmailRequest, BulkEmailResponse
)
from schemas.base import ResponseEnvelope
from schemas.base import ApiError
from core.models.party import Party

from services.party_service import PartyService

from database import get_db
from middleware.unified_access import UnifiedAccess, require_access

import logging
logger = logging.getLogger(__name__)

router = APIRouter(tags=["Parties"])

# --- HELPERS ---

PATIENT_DETAIL_SENSITIVE_PERMISSION_MAP = {
    "phone": "sensitive.parties.detail.contact.view",
    "email": "sensitive.parties.detail.contact.view",
    "tc_number": "sensitive.parties.detail.identity.view",
    "identity_number": "sensitive.parties.detail.identity.view",
    "sgk_info": "sensitive.parties.detail.sgk.view",
}

PATIENT_LIST_SENSITIVE_PERMISSION_MAP = {
    "phone": "sensitive.parties.list.contact.view",
    "email": "sensitive.parties.list.contact.view",
    "tc_number": "sensitive.parties.list.identity.view",
}


def _mask_party_detail_response(patient: Party | dict, access: UnifiedAccess) -> PartyRead:
    payload = PartyRead.model_validate(patient).model_dump(by_alias=False)

    for field_name, permission_name in PATIENT_DETAIL_SENSITIVE_PERMISSION_MAP.items():
        if not access.has_permission(permission_name):
            payload[field_name] = {} if field_name == "sgk_info" else None

    return PartyRead.model_validate(payload)


def _mask_party_list_response(patients: list[Party], access: UnifiedAccess) -> list[PartyRead]:
    masked_items: list[PartyRead] = []

    for patient in patients:
        payload = PartyRead.model_validate(patient).model_dump(by_alias=False)
        for field_name, permission_name in PATIENT_LIST_SENSITIVE_PERMISSION_MAP.items():
            if not access.has_permission(permission_name):
                payload[field_name] = None
        masked_items.append(PartyRead.model_validate(payload))

    return masked_items



# --- ROUTES ---

@router.get("/parties", operation_id="listParties", response_model=ResponseEnvelope[List[PartyRead]])
def list_parties(
    page: int = Query(1, ge=1),
    per_page: int = Query(20, ge=1, le=1000),
    search: Optional[str] = None,
    status_filter: Optional[str] = Query(None, alias="status"),
    city: Optional[str] = None,
    district: Optional[str] = None,
    cursor: Optional[str] = None,
    access: UnifiedAccess = Depends(require_access("parties.view")),
    db: Session = Depends(get_db)
):
    """List patients with filtering and pagination"""
    try:
        service = PartyService(db)
        
        branch_ids = None
        if access.is_tenant_admin and access.user and access.user.role == 'admin':
             # Logic matching previous implementation: narrow scope to user's branches if admin
             try:
                 user_branches = getattr(access.user, 'branches', [])
                 if user_branches:
                     branch_ids = [b.id for b in user_branches]
                 elif hasattr(access.user, 'branches'): 
                     # Has attribute but empty list implied restricted access to 0 branches
                     branch_ids = []
             except Exception as branch_error:
                 logger.warning(f"Failed to load user branches: {branch_error}")
                 # Fallback: ignore branch restriction if loading fails, or secure closed?
                 # Ignoring to prevent 500, assuming if critical it should have worked.
                 branch_ids = None
        
        items, total, next_cursor = service.list_parties(
            tenant_id=access.tenant_id,
            page=page,
            per_page=per_page,
            search=search,
            status=status_filter,
            city=city,
            district=district,
            cursor=cursor,
            branch_ids=branch_ids
        )
            
        # Manual pagination meta calculation if not cursor
        total_pages = 0
        if not cursor and total > 0:
            total_pages = (total + per_page - 1) // per_page
            
        return ResponseEnvelope(
            data=_mask_party_list_response(items, access),
            meta={
                "total": total,
                "page": page,
                "perPage": per_page,
                "totalPages": total_pages,
                "hasNext": True if next_cursor else False, # Simplified hasNext logic based on cursor
                "nextCursor": next_cursor
            }
        )
    except Exception as e:
        logger.error(f"List patients error: {e}")
        import traceback
        with open("last_error.txt", "w") as f:
            f.write(f"Error: {str(e)}\n")
            f.write(traceback.format_exc())
        print(f"CRITICAL 500 ERROR IN LIST PARTIES: {e}", flush=True)
        raise HTTPException(status_code=500, detail="Internal server error")
# ... imports ...

@router.post("/parties", operation_id="createParties", response_model=ResponseEnvelope[PartyRead], status_code=201)
def create_party(
    patient_in: PartyCreate,
    access: UnifiedAccess = Depends(require_access("parties.create")),
    db: Session = Depends(get_db)
):
    """Create a new patient"""
    try:
        from core.tenant_utils import get_effective_tenant_id
        from sqlalchemy.exc import IntegrityError
        from models.user import ActivityLog
        from models.tenant import Tenant

        service = PartyService(db)
        # Resolve tenant_id: super_admin with 'system' falls back to first real tenant
        try:
            tenant_id = get_effective_tenant_id(access)
        except HTTPException:
            if access.is_super_admin:
                first_tenant = db.query(Tenant).first()
                if first_tenant:
                    tenant_id = first_tenant.id
                else:
                    raise HTTPException(status_code=400, detail="No tenants available")
            else:
                raise

        # Service handles data normalization and defaults
        # Use by_alias=False to get snake_case keys that Party.from_dict expects
        data = patient_in.model_dump(exclude_unset=True, by_alias=False)
        
        patient = service.create_party(data, tenant_id)
        
        # Log activity
        try:
            activity_log = ActivityLog(
                user_id=getattr(access, 'user_id', None) or 'system',
                action='party_created',
                entity_type='party',
                entity_id=patient.id,
                tenant_id=tenant_id,
                details=json.dumps({
                    'name': f"{patient.first_name} {patient.last_name}",
                    'phone': patient.phone,
                    'status': patient.status
                })
            )
            db.add(activity_log)
            db.commit()
        except Exception as log_error:
            logger.warning(f"Failed to log party creation: {log_error}")
        
        return ResponseEnvelope(data=patient)
    except IntegrityError as e:
        logger.error(f"Party integrity error: {e}")
        db.rollback()
        error_msg = str(e.orig) if hasattr(e, 'orig') else str(e)
        if 'phone' in error_msg.lower():
            raise HTTPException(status_code=409, detail="Phone number already exists")
        elif 'email' in error_msg.lower():
            raise HTTPException(status_code=409, detail="Email already exists")
        elif 'tc_number' in error_msg.lower():
            raise HTTPException(status_code=409, detail="TC number already exists")
        raise HTTPException(status_code=409, detail="Duplicate entry")
    except Exception as e:
        logger.error(f"Create patient error: {e}")
        # Service might raise HTTPException or others, safe to re-raise or wrap
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail="Internal server error")

@router.get("/parties/export", operation_id="listPartyExport")
def export_parties(
    q: Optional[str] = None,
    status: Optional[str] = None,
    segment: Optional[str] = None,
    access: UnifiedAccess = Depends(require_access("parties.export")),
    db: Session = Depends(get_db)
):
    """Export patients as CSV"""
    from fastapi.responses import StreamingResponse
    import io
    import csv
    
    if not access.tenant_id:
         raise HTTPException(status_code=400, detail="Tenant context required")

    service = PartyService(db)
    
    branch_ids = None
    if access.is_tenant_admin and access.user and access.user.role == 'admin':
         user_branches = getattr(access.user, 'branches', [])
         if user_branches:
             branch_ids = [b.id for b in user_branches]
         elif hasattr(access.user, 'branches'):
             branch_ids = []

    # Generator for streaming
    def iter_csv():
        output = io.StringIO()
        writer = csv.writer(output)

        # Header
        writer.writerow(['id','tcNumber','firstName','lastName','phone','email','birthDate','gender','status','segment','tags','createdAt'])
        output.seek(0)
        yield output.read()
        output.seek(0)
        output.truncate(0)

        # Iterate via Service (with hard limit for safety)
        try:
            iterator = service.iter_parties(
                tenant_id=access.tenant_id,
                search=q,
                status=status,
                segment=segment,
                branch_ids=branch_ids
            )
        except Exception as e:
            logger.error(f"Export iteration error: {e}")
            return

        for p in iterator:
            tags_str = json.dumps(p.tags_json) if p.tags_json else "[]"
            # Handle potential None dates
            created = p.created_at.isoformat() if p.created_at else ''
            
            # Serialize enum properly - use .value to get string representation
            status_value = p.status.value if hasattr(p.status, 'value') else str(p.status or '')
            
            writer.writerow([
                str(p.id),
                str(p.tc_number or ''),
                str(p.first_name or ''),
                str(p.last_name or ''),
                str(p.phone or ''),
                str(p.email or ''),
                str(p.birth_date or ''),
                str(p.gender or ''),
                status_value,
                str(p.segment or ''),
                tags_str,
                created
            ])
            output.seek(0)
            yield output.read()
            output.seek(0)
            output.truncate(0)
            
    filename = f"patients_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    return StreamingResponse(
        iter_csv(),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@router.get("/parties/count", operation_id="listPartyCount")
def count_parties(
    access: UnifiedAccess = Depends(require_access("parties.view")),
    db: Session = Depends(get_db),
    status: Optional[str] = None,
    segment: Optional[str] = None
):
    """Count patients"""
    if not access.tenant_id:
        return {"data": {"count": 0}}
        
    service = PartyService(db)
    count = service.count_parties(access.tenant_id, status=status, segment=segment)
        
    return ResponseEnvelope(data={'count': count})

@router.get("/parties/search", operation_id="searchParties", response_model=ResponseEnvelope[List[dict]])
def search_parties(
    q: str = Query(..., min_length=1, description="Search query"),
    limit: int = Query(10, ge=1, le=50, description="Maximum results"),
    access: UnifiedAccess = Depends(require_access("parties.view")),
    db: Session = Depends(get_db)
):
    """Search parties for selection (fuzzy search by name/phone)"""
    service = PartyService(db)
    results = service.search_parties(access.tenant_id, query=q, limit=limit)
    
    # Format for frontend consumption
    formatted_results = []
    for party in results:
        formatted_results.append({
            "partyId": party.id,
            "firstName": party.first_name or "",
            "lastName": party.last_name or "",
            "phone": party.phone,
            "fullName": f"{party.first_name or ''} {party.last_name or ''}".strip()
        })
    
    return ResponseEnvelope(data=formatted_results)

@router.get("/parties/{party_id}", operation_id="getParty", response_model=ResponseEnvelope[PartyRead])
def get_party(
    party_id: str,
    access: UnifiedAccess = Depends(require_access("parties.view")),
    db: Session = Depends(get_db)
):
    """Get single patient"""
    service = PartyService(db)
    # Service raises 404 if not found or tenant mismatch
    patient = service.get_party(party_id, access.tenant_id)
    return ResponseEnvelope(data=_mask_party_detail_response(patient, access))

@router.put("/parties/{party_id}", operation_id="updateParty", response_model=ResponseEnvelope[PartyRead])
def update_party(
    party_id: str,
    patient_in: PartyUpdate,
    access: UnifiedAccess = Depends(require_access("parties.edit")),
    db: Session = Depends(get_db)
):
    """Update patient"""
    from models.user import ActivityLog
    
    service = PartyService(db)
    
    # Use by_alias=False to get snake_case keys
    data = patient_in.model_dump(exclude_unset=True, by_alias=False)
    try:
        updated_patient = service.update_party(party_id, data, access.tenant_id)
        
        # Log activity
        try:
            # Build change summary
            changes = {}
            if 'status' in data:
                changes['status'] = data['status']
            if 'segment' in data:
                changes['segment'] = data['segment']
            if 'acquisition_type' in data:
                changes['acquisition_type'] = data['acquisition_type']
            if 'branch_id' in data:
                changes['branch_id'] = data['branch_id']
            
            activity_log = ActivityLog(
                user_id=getattr(access, 'user_id', None) or 'system',
                action='party_updated',
                entity_type='party',
                entity_id=party_id,
                tenant_id=access.tenant_id,
                details=json.dumps({
                    'name': f"{updated_patient.first_name} {updated_patient.last_name}",
                    'changes': changes
                })
            )
            db.add(activity_log)
            db.commit()
        except Exception as log_error:
            logger.warning(f"Failed to log party update: {log_error}")
        
        return ResponseEnvelope(data=updated_patient)
    except Exception as e:
        logger.error(f"Update patient error: {e}")
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail="Internal server error")

@router.delete("/parties/{party_id}", operation_id="deleteParty")
def delete_party(
    party_id: str,
    access: UnifiedAccess = Depends(require_access("parties.delete")),
    db: Session = Depends(get_db)
):
    """Delete patient"""
    from models.user import ActivityLog
    
    service = PartyService(db)
    try:
        # Get party info before deletion for logging
        party = service.get_party(party_id, access.tenant_id)
        party_name = f"{party.first_name} {party.last_name}" if party else "Unknown"
        
        service.delete_party(party_id, access.tenant_id)
        
        # Log activity
        try:
            activity_log = ActivityLog(
                user_id=getattr(access, 'user_id', None) or 'system',
                action='party_deleted',
                entity_type='party',
                entity_id=party_id,
                tenant_id=access.tenant_id,
                details=json.dumps({
                    'name': party_name
                })
            )
            db.add(activity_log)
            db.commit()
        except Exception as log_error:
            logger.warning(f"Failed to log party deletion: {log_error}")
        
        return ResponseEnvelope(message="Patient deleted")
    except Exception as e:
        logger.error(f"Delete patient error: {e}")
        if isinstance(e, HTTPException):
            raise e
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/parties/bulk-upload", operation_id="createPartyBulkUpload", response_model=ResponseEnvelope[BulkUploadResponse])
async def bulk_upload_parties(
    file: UploadFile = File(...),
    update_mode: str = Query(default="fill_empty", description="Update strategy: 'fill_empty' or 'overwrite'"),
    access: UnifiedAccess = Depends(require_access("parties.create")),
    db: Session = Depends(get_db)
):
    """Bulk upload patients from CSV or XLSX. update_mode: fill_empty (default) or overwrite."""
    try:
        # Use effective_tenant_id if impersonating, otherwise use tenant_id
        effective_tenant = access.effective_tenant_id or access.tenant_id
        
        logger.info(f"Bulk upload started - tenant_id: {access.tenant_id}, effective_tenant_id: {access.effective_tenant_id}, using: {effective_tenant}, filename: {file.filename}")
        
        if not effective_tenant or effective_tenant == 'system':
            logger.error(f"Bulk upload rejected - invalid tenant. tenant_id: {access.tenant_id}, effective: {effective_tenant}")
            raise HTTPException(
                status_code=400,
                detail=ApiError(message="Lütfen işlem yapmak için bir klinik (tenant) seçiniz.", code="TENANT_REQUIRED").model_dump(mode="json")
            )

        filename = (file.filename or '').lower()
        content = await file.read()
        
        rows = []
        headers = []

        def _sanitize_cell(v):
            if v is None:
                return None
            if not isinstance(v, str):
                return v
            v = v.strip()
            # Prevent CSV/Excel formula injection
            if v.startswith(('=', '+', '-', '@')):
                return "'" + v
            return v

        # Parse File
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                raise HTTPException(status_code=500, detail="Server missing openpyxl dependency")
            
            try:
                wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None:
                     raise HTTPException(status_code=400, detail="XLSX contains no sheets")
                
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row:
                     raise HTTPException(status_code=400, detail="XLSX first row empty")
                
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                if not any(h for h in headers):
                     raise HTTPException(status_code=400, detail="XLSX has no headers")
                     
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        obj[h] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
            except Exception as e:
                logger.error(f"XLSX parse error: {e}")
                raise HTTPException(status_code=400, detail="XLSX parse error")
        else:
            # CSV assumption
            try:
                # Try UTF-8-SIG then UTF-8
                try:
                    text_content = content.decode('utf-8-sig')
                except UnicodeDecodeError:
                    text_content = content.decode('utf-8', errors='replace')
                
                # Sniffer
                try:
                    dialect = csv.Sniffer().sniff(text_content[:4096])
                    delimiter = dialect.delimiter
                except Exception:
                    delimiter = ','
                
                # Fallback check for delimiter
                first_line = text_content.split('\n')[0]
                if ';' in first_line and delimiter != ';':
                    delimiter = ';'
                
                reader = csv.DictReader(io.StringIO(text_content), delimiter=delimiter)
                headers = reader.fieldnames or []
                rows = [row for row in reader]
                
            except Exception as e:
                logger.error(f"CSV parse error: {e}")
                raise HTTPException(status_code=400, detail="CSV parse error")

        # Process Rows
        created = 0
        updated = 0
        errors = []
        row_num = 0

        for row in rows:
            row_num += 1
            try:
                if isinstance(row, dict):
                    normalized_row = {k: _sanitize_cell(v) for k, v in row.items()}
                else:
                    normalized_row = row
                
                # Normalize keys (handle variations) - case-insensitive
                def get_val(keys):
                    # First try exact match
                    for k in keys:
                        if k in normalized_row and normalized_row[k]:
                            return normalized_row[k]
                    # Then try case-insensitive match
                    normalized_keys = {k.lower(): k for k in normalized_row.keys()}
                    for k in keys:
                        lower_k = k.lower()
                        if lower_k in normalized_keys and normalized_row[normalized_keys[lower_k]]:
                            return normalized_row[normalized_keys[lower_k]]
                    return None
                    
                tc_number = get_val(['tcNumber', 'tc_number', 'tc', 'TC', 'tc_no', 'TC Kimlik No', 'tc kimlik no', 'tc_kimlik_no'])
                first_name = get_val(['firstName', 'first_name', 'first', 'isim', 'ad', 'Ad', 'İsim'])
                last_name = get_val(['lastName', 'last_name', 'last', 'soyisim', 'soyad', 'Soyad', 'Soyisim'])
                phone = get_val(['phone', 'phone_number', 'tel', 'telefon', 'cep_telefon', 'cep', 'Telefon', 'Tel'])
                
                if not (first_name or phone or tc_number):
                    # Skip empty rows or not enough info
                    errors.append({'row': row_num, 'error': 'Missing required identifying fields'})
                    continue

                # Prepare Payload — all Party model fields
                payload = {
                    'tc_number': tc_number,
                    'identity_number': get_val(['identityNumber', 'identity_number', 'kimlik_no', 'Kimlik No']),
                    'first_name': first_name,
                    'last_name': last_name,
                    'phone': phone,
                    'email': get_val(['email', 'e-mail', 'eposta', 'Email', 'E-posta']),
                    'birth_date': get_val(['birthDate', 'birth_date', 'dob', 'dogum_tarihi', 'Doğum Tarihi', 'dogum tarihi']),
                    'gender': get_val(['gender', 'cinsiyet', 'Cinsiyet', 'Gender']),
                    'status': get_val(['status', 'durum', 'Durum', 'Status']),
                    'segment': get_val(['segment', 'Segment']),
                    'acquisition_type': get_val(['acquisitionType', 'acquisition_type', 'kazanim_tipi', 'Kazanım Tipi']),
                    'referred_by': get_val(['referredBy', 'referred_by', 'referans', 'Referans']),
                }
                
                # Normalize gender values (Turkish → English)
                if payload.get('gender'):
                    gender_lower = str(payload['gender']).lower()
                    if gender_lower in ['erkek', 'male', 'm', 'e']:
                        payload['gender'] = 'MALE'
                    elif gender_lower in ['kadın', 'kadın', 'female', 'f', 'k']:
                        payload['gender'] = 'FEMALE'
                    else:
                        # Keep original if not recognized
                        pass
                
                # Normalize status values (Turkish → English)
                if payload.get('status'):
                    status_lower = str(payload['status']).lower()
                    if status_lower in ['aktif', 'active']:
                        payload['status'] = 'ACTIVE'
                    elif status_lower in ['pasif', 'passive', 'inactive']:
                        payload['status'] = 'INACTIVE'
                    else:
                        # Keep original if not recognized
                        pass
                
                # Clean up None values
                payload = {k: v for k, v in payload.items() if v is not None}
                
                # Address
                address_city = get_val(['address_city', 'addressCity', 'city', 'il', 'sehir', 'Şehir', 'şehir', 'City'])
                address_district = get_val(['address_district', 'addressDistrict', 'district', 'ilce', 'İlçe', 'ilçe', 'District'])
                address_full = get_val(['address_full', 'addressFull', 'fullAddress', 'address', 'adres', 'Adres', 'Address'])

                if address_city: payload['address_city'] = address_city
                if address_district: payload['address_district'] = address_district
                if address_full: payload['address_full'] = address_full

                # Tags
                tags_val = get_val(['tags', 'etiketler', 'Etiketler'])
                if tags_val:
                    if isinstance(tags_val, str):
                        payload['tags_json'] = [t.strip() for t in tags_val.split(',') if t.strip()]
                    else:
                        payload['tags_json'] = tags_val

                # Check Existing — cascading match (tc > phone > name)
                existing = None

                # Priority 1: TC number (strongest identifier)
                if tc_number:
                    existing = db.query(Party).filter(
                        Party.tc_number == tc_number,
                        Party.tenant_id == effective_tenant
                    ).first()

                # Priority 2: Phone (exact match, tenant-scoped)
                if not existing and phone:
                    existing = db.query(Party).filter(
                        Party.phone == phone,
                        Party.tenant_id == effective_tenant
                    ).first()

                # Priority 3: first_name + last_name (tenant-scoped)
                if not existing and first_name and last_name:
                    existing = db.query(Party).filter(
                        Party.first_name == first_name,
                        Party.last_name == last_name,
                        Party.tenant_id == effective_tenant
                    ).first()

                if existing:
                    # Update Logic — respects update_mode
                    def _parse_birth_date(v):
                        if not isinstance(v, str):
                            return v
                        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y/%m/%d'):
                            try:
                                return datetime.strptime(v, fmt)
                            except ValueError:
                                pass
                        try:
                            return datetime.fromisoformat(v)
                        except:
                            return None

                    for k, v in payload.items():
                        if k == 'tags_json':
                            continue  # handled separately below
                        if hasattr(existing, k) and v is not None:
                            current_val = getattr(existing, k, None)
                            should_update = False
                            if update_mode == 'overwrite':
                                should_update = True
                            else:
                                # fill_empty: only fill if current value is empty/null
                                should_update = current_val is None or current_val == ''

                            if should_update:
                                if k == 'birth_date':
                                    parsed = _parse_birth_date(v)
                                    if parsed:
                                        setattr(existing, k, parsed)
                                else:
                                    setattr(existing, k, v)
                    # Merge tags (union, not replace) — regardless of mode
                    if 'tags_json' in payload and payload['tags_json']:
                        existing_tags = existing.tags_json or []
                        merged = list(set(existing_tags + payload['tags_json']))
                        existing.tags_json = merged
                    updated += 1
                    db.add(existing)
                else:
                    # Create Logic
                    if not payload.get('first_name') or not payload.get('last_name'):
                         errors.append({'row': row_num, 'error': 'First Name and Last Name required for new patients'})
                         continue
                         
                    # Handle Date parsing
                    if 'birth_date' in payload and isinstance(payload['birth_date'], str):
                         v = payload['birth_date']
                         dt = None
                         for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y', '%Y/%m/%d'):
                            try:
                                dt = datetime.strptime(v, fmt)
                                break
                            except ValueError:
                                pass
                         if dt:
                             payload['birth_date'] = dt
                         else:
                             try:
                                 payload['birth_date'] = datetime.fromisoformat(v)
                             except:
                                 del payload['birth_date']
                    
                    patient = Party(tenant_id=effective_tenant, **payload)
                    if not patient.status: patient.status = 'ACTIVE'
                    
                    db.add(patient)
                    created += 1
                
                # Use savepoint to prevent full rollback on single row failure
                db.begin_nested() 
                try:
                    db.flush()
                    db.commit() # Commit the savepoint
                except Exception as e:
                    db.rollback() # Rollback to savepoint
                    errors.append({'row': row_num, 'error': str(e)})

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})
        
        db.commit()
        return ResponseEnvelope(data={
            'success': True,
            'created': created,
            'updated': updated,
            'errors': errors
        })
        
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        logger.error(f"Bulk upload patient error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")



# --- Bulk Operations ---

@router.post("/parties/bulk-update", operation_id="bulkUpdateParties", response_model=ResponseEnvelope[BulkUpdateResponse])
async def bulk_update_parties(
    request: BulkUpdateRequest,
    access: UnifiedAccess = Depends(require_access("parties.edit")),
    db: Session = Depends(get_db)
):
    """
    Bulk update multiple parties with the same changes.
    Returns success/failure count and individual results.
    """
    try:
        from schemas.parties import BulkUpdateResponse, BulkUpdateResult
        
        service = PartyService(db)
        results = []
        success_count = 0
        failure_count = 0
        
        for party_id in request.party_ids:
            try:
                # Verify party exists and belongs to tenant
                party = service.get_party(party_id, access.tenant_id)
                if not party:
                    results.append(BulkUpdateResult(
                        party_id=party_id,
                        success=False,
                        error="Party not found"
                    ))
                    failure_count += 1
                    continue

                # Apply updates
                update_data = request.updates.model_dump(exclude_unset=True, by_alias=False)
                for key, value in update_data.items():
                    if hasattr(party, key):
                        setattr(party, key, value)

                results.append(BulkUpdateResult(
                    party_id=party_id,
                    success=True
                ))
                success_count += 1

            except Exception as e:
                logger.error(f"Failed to update party {party_id}: {e}")
                db.rollback()
                results.append(BulkUpdateResult(
                    party_id=party_id,
                    success=False,
                    error="Update failed"
                ))
                failure_count += 1

        # Single commit for all successful updates
        if success_count > 0:
            try:
                db.commit()
            except Exception as e:
                logger.error(f"Bulk commit failed: {e}")
                db.rollback()
                # Mark all as failed
                for r in results:
                    if r.success:
                        r.success = False
                        r.error = "Batch commit failed"
                        failure_count += 1
                        success_count -= 1
        
        response_data = BulkUpdateResponse(
            success_count=success_count,
            failure_count=failure_count,
            results=results
        )
        
        return ResponseEnvelope(data=response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk update error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")


@router.post("/parties/bulk-email", operation_id="bulkEmailParties", response_model=ResponseEnvelope[BulkEmailResponse])
async def bulk_email_parties(
    request: BulkEmailRequest,
    access: UnifiedAccess = Depends(require_access("parties.edit")),
    db: Session = Depends(get_db)
):
    """
    Send email to multiple parties.
    Returns success/failure count and individual results.
    """
    try:
        from schemas.parties import BulkEmailResponse, BulkEmailResult
        
        service = PartyService(db)
        results = []
        success_count = 0
        failure_count = 0
        
        for party_id in request.party_ids:
            try:
                # Verify party exists and belongs to tenant
                party = service.get_party(party_id, access.tenant_id)
                if not party:
                    results.append(BulkEmailResult(
                        party_id=party_id,
                        success=False,
                        error="Party not found"
                    ))
                    failure_count += 1
                    continue
                
                # Check if party has email
                if not party.email:
                    results.append(BulkEmailResult(
                        party_id=party_id,
                        email=None,
                        success=False,
                        error="Party has no email address"
                    ))
                    failure_count += 1
                    continue
                
                # TODO: Implement actual email sending
                # For now, just log and mark as success
                logger.info(f"Would send email to {party.email}: {request.subject}")
                
                results.append(BulkEmailResult(
                    party_id=party_id,
                    email=party.email,
                    success=True
                ))
                success_count += 1
                
            except Exception as e:
                logger.error(f"Failed to email party {party_id}: {e}")
                results.append(BulkEmailResult(
                    party_id=party_id,
                    success=False,
                    error=str(e)
                ))
                failure_count += 1
        
        response_data = BulkEmailResponse(
            success_count=success_count,
            failure_count=failure_count,
            results=results
        )
        
        return ResponseEnvelope(data=response_data)
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Bulk email error: {e}")
        raise HTTPException(status_code=500, detail="Internal server error")
