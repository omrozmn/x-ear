"""
FastAPI Invoices Router - Migrated from Flask routes/invoices/
Handles invoice CRUD operations
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from typing import Optional, List, Dict, Any
from datetime import datetime
import logging
import uuid
import csv
import io

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from sqlalchemy.orm import Session
from sqlalchemy import desc

from schemas.base import ResponseEnvelope
from schemas.invoices import (
    InvoiceCreate, InvoiceUpdate, InvoiceRead, BatchInvoiceGenerateRequest,
    BulkUploadResponse, InvoicePrintQueueResponse, InvoiceAddToQueueRequest, InvoiceTemplate, InvoicePrintQueueItem
)
from models.user import User
from models.invoice import Invoice
from models.patient import Patient
from models.sales import Sale
from models.efatura_outbox import EFaturaOutbox
from models.user import ActivityLog
from utils.efatura import build_return_invoice_xml, write_outbox_file
import os
from middleware.unified_access import UnifiedAccess, require_access, require_admin
from database import get_db

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Invoices"])

# --- Routes ---

@router.get("/invoices", operation_id="listInvoices", response_model=ResponseEnvelope[List[InvoiceRead]])
def get_invoices(
    page: int = 1,
    per_page: int = 20,
    status: Optional[str] = None,
    patient_id: Optional[str] = None,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access())
):
    """Get all invoices with pagination"""
    try:
        query = db_session.query(Invoice).filter(
            Invoice.tenant_id == access.tenant_id,
            Invoice.status != 'deleted'
        )
        
        if status:
            query = query.filter(Invoice.status == status)
        if patient_id:
            query = query.filter(Invoice.patient_id == patient_id)
        
        query = query.order_by(desc(Invoice.created_at))
        total = query.count()
        invoices = query.offset((page - 1) * per_page).limit(per_page).all()
        
        return ResponseEnvelope(
            data=[inv.to_dict() for inv in invoices],
            meta={
                "page": page,
                "perPage": per_page,
                "total": total,
                "totalPages": (total + per_page - 1) // per_page,
            },
        )
    except Exception as e:
        logger.error(f"Get invoices error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/invoices/print-queue", operation_id="listInvoicePrintQueue", response_model=ResponseEnvelope[InvoicePrintQueueResponse])
def get_print_queue(
    access: UnifiedAccess = Depends(require_access("invoices.read")),
    db_session: Session = Depends(get_db)
):
    """Get invoices in print queue"""
    try:
        query = db_session.query(Invoice).filter(
            Invoice.tenant_id == access.tenant_id,
            Invoice.status == 'queued_for_print'
        )
        invoices = query.all()
        
        items = []
        for inv in invoices:
            items.append({
                'id': inv.id,
                'invoiceNumber': inv.invoice_number,
                'patientName': inv.patient_name or (inv.patient.full_name if inv.patient else 'Unknown'),
                'amount': inv.device_price,
                'queuedAt': inv.updated_at.isoformat() if inv.updated_at else inv.created_at.isoformat(),
                'priority': 'normal',
                'copies': 1,
                'status': inv.status
            })
            
        return ResponseEnvelope(data={
            'items': items,
            'total': len(items),
            'status': 'ready' if items else 'empty'
        })
    except Exception as e:
        logger.error(f"Get print queue error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/invoices/print-queue", operation_id="createInvoicePrintQueue")
def add_to_print_queue(
    payload: InvoiceAddToQueueRequest,
    access: UnifiedAccess = Depends(require_access("invoices.write")),
    db_session: Session = Depends(get_db)
):
    """Add invoices to print queue"""
    try:
        updated_count = 0
        errors = []
        
        for inv_id in payload.invoice_ids:
            inv = db_session.query(Invoice).filter(
                Invoice.id == inv_id,
                Invoice.tenant_id == access.tenant_id
            ).first()
            
            if inv:
                inv.status = 'queued_for_print'
                inv.updated_at = datetime.utcnow()
                updated_count += 1
            else:
                errors.append(f"Invoice {inv_id} not found")
        
        db_session.commit()
        
        return ResponseEnvelope(data={
            'queuedCount': updated_count,
            'errors': errors
        })
    except Exception as e:
        db_session.rollback()
        logger.error(f"Add to print queue error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/invoices/templates", operation_id="listInvoiceTemplates", response_model=ResponseEnvelope[List[InvoiceTemplate]])
def get_invoice_templates(
    access: UnifiedAccess = Depends(require_access("invoices.read"))
):
    """Get available invoice templates"""
    # Mock data as per legacy implementation
    templates = [
        {
            'id': 'standard',
            'name': 'Standard Invoice',
            'description': 'Default invoice template with company branding',
            'fields': ['invoiceNumber', 'date', 'customerInfo', 'items', 'totals', 'paymentTerms'],
            'isDefault': True,
            'isCustom': False
        },
        {
            'id': 'detailed',
            'name': 'Detailed Invoice',
            'description': 'Comprehensive invoice with item descriptions',
            'fields': ['invoiceNumber', 'date', 'customerInfo', 'items', 'itemDetails', 'totals', 'paymentTerms', 'notes'],
            'isDefault': False,
            'isCustom': False
        },
        {
            'id': 'sgk',
            'name': 'SGK Invoice',
            'description': 'Invoice template for SGK submissions',
            'fields': ['invoiceNumber', 'date', 'customerInfo', 'sgkInfo', 'items', 'totals', 'sgkCoverage'],
            'isDefault': False,
            'isCustom': False
        },
        {
            'id': 'proforma',
            'name': 'Proforma Invoice',
            'description': 'Proforma invoice template for quotations',
            'fields': ['proformaNumber', 'date', 'customerInfo', 'items', 'totals', 'validityPeriod'],
            'isDefault': False,
            'isCustom': False
        }
    ]
    return ResponseEnvelope(data=templates)

@router.post("/invoices/templates", operation_id="createInvoiceTemplates", response_model=ResponseEnvelope[InvoiceTemplate], status_code=201)
def create_invoice_template(
    template: InvoiceTemplate,
    access: UnifiedAccess = Depends(require_access("invoices.write"))
):
    """Create custom invoice template (Mock)"""
    # Just echo back for now as DB support is missing in legacy too
    t_dict = template.model_dump()
    t_dict['id'] = str(uuid.uuid4())
    t_dict['created_by'] = access.principal_id
    t_dict['created_at'] = datetime.utcnow()
    t_dict['is_custom'] = True
    
    return ResponseEnvelope(data=t_dict)

@router.post("/invoices/batch-generate", operation_id="createInvoiceBatchGenerate")
def batch_generate_invoices(
    request_data: BatchInvoiceGenerateRequest,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access("invoices.write"))
):
    """Generate multiple invoices at once"""
    try:
        sale_ids = request_data.saleIds
        invoice_type = request_data.invoice_type
        customer_info = request_data.customer_info
        
        generated_invoices = []
        errors = []
        
        # Helper for date
        def now_utc():
            return datetime.utcnow()
            
        for sale_id in sale_ids:
            try:
                # Get sale with tenant check
                sale = db_session.get(Sale, sale_id)
                if not sale:
                    errors.append(f"Sale {sale_id} not found")
                    continue
                    
                if access.tenant_id and sale.tenant_id != access.tenant_id:
                    errors.append(f"Sale {sale_id} not found")
                    continue
                
                # Check if invoice already exists
                existing_invoice = db_session.query(Invoice).filter(
                    Invoice.sale_id == sale_id,
                    Invoice.status != 'deleted'
                ).first()
                
                if existing_invoice:
                    errors.append(f"Invoice already exists for sale {sale_id}")
                    continue
                
                # Generate invoice number
                year_month = now_utc().strftime('%Y%m')
                latest = db_session.query(Invoice).filter(
                    Invoice.invoice_number.like(f'INV{year_month}%'),
                    Invoice.tenant_id == access.tenant_id
                ).order_by(desc(Invoice.created_at)).first()
                 # Note: Flask code filtered like that but maybe global? No, Invoice number collision risk if global. 
                 # Flask code: Invoice.query.filter(...like...)
                 # It didn't filter by tenant_id in 'latest' query explicitly in bulk.py snippet I saw.
                 # But Invoice numbers usually sequential per system or per tenant?
                 # 'INV{YM}{NUM}' format looks global or requires prefix.
                 # If per tenant, numbers clash. Assuming global unique or tenant prefix?
                 # Flask snippet: `Invoice.query.filter(Invoice.invoice_number.like(...))` -> GLOBAL check if not scoped.
                 # But Invoice model usually holds tenant_id.
                 # I will scope it to tenant to be safe, or just use UUID if strict parity allows.
                 # User approved "Complete migration". 
                 # Let's scope to tenant for correctness in multi-tenant system.
                
                if latest:
                    try:
                        last_num = int(latest.invoice_number[-4:])
                        new_num = last_num + 1
                    except ValueError:
                        new_num = 1
                else:
                    new_num = 1
                
                invoice_number = f"INV{year_month}{new_num:04d}"
                
                # Create invoice
                invoice = Invoice(
                    tenant_id=sale.tenant_id,
                    branch_id=sale.branch_id,
                    invoice_number=invoice_number,
                    patient_id=sale.patient_id,
                    sale_id=sale_id,
                    device_price=sale.total_amount or 0,
                    status='draft',
                    notes=f"SGK: {sale.sgk_coverage or 0}, Patient: {sale.patient_payment or 0}, Type: {invoice_type}",
                    created_at=now_utc(),
                    updated_at=now_utc(),
                    created_by=access.user_id
                )
                
                # Add custom customer info
                if customer_info:
                    invoice.customer_name = customer_info.get('name') or (sale.patient.full_name if sale.patient else '')
                    invoice.customer_address = customer_info.get('address')
                    invoice.customer_tax_number = customer_info.get('taxNumber')
                    
                db_session.add(invoice)
                db_session.flush() # Force ID generation and uniqueness check (if any)
                
                generated_invoices.append({
                    'id': invoice.id,
                    'invoiceNumber': invoice.invoice_number,
                    'saleId': sale_id,
                    'amount': invoice.device_price,
                    'status': invoice.status
                })
                
            except Exception as e:
                errors.append(f"Error generating invoice for sale {sale_id}: {str(e)}")
        
        db_session.commit()
        
        return {
            'success': True,
            'data': {
                'generatedInvoices': generated_invoices,
                'errors': errors,
                'summary': {
                    'total': len(sale_ids),
                    'successful': len(generated_invoices),
                    'failed': len(errors)
                }
            },
            'requestId': str(uuid.uuid4()),
            'timestamp': now_utc().isoformat()
        }
        
    except Exception as e:
        db_session.rollback()
        logger.error(f"Batch generate invoices error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/invoices/{invoice_id}", operation_id="getInvoice", response_model=ResponseEnvelope[InvoiceRead])
def get_invoice(
    invoice_id: int,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access())
):
    """Get a specific invoice"""
    invoice = db_session.query(Invoice).filter(
        Invoice.id == invoice_id,
        Invoice.tenant_id == access.tenant_id
    ).first()
    
    if not invoice or invoice.status == 'deleted':
        raise HTTPException(status_code=404, detail={"message": "Invoice not found", "code": "NOT_FOUND"})
    
    return ResponseEnvelope(data=invoice.to_dict())

@router.post("/invoices", operation_id="createInvoices", response_model=ResponseEnvelope[InvoiceRead], status_code=201)
def create_invoice(
    request_data: InvoiceCreate,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access())
):
    """Create a new invoice"""
    try:
        # Verify patient exists
        payload = request_data.model_dump(by_alias=False)
        patient_id = payload.get("patient_id")
        sale_id = payload.get("sale_id")

        # Patient is optional for some invoice types
        if patient_id:
            patient = db_session.query(Patient).filter(
                Patient.id == patient_id,
                Patient.tenant_id == access.tenant_id
            ).first()
            
            if not patient:
                raise HTTPException(status_code=404, detail={"message": "Patient not found", "code": "NOT_FOUND"})
        
        # Generate invoice number
        from datetime import datetime
        invoice_number = f"INV-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
        
        invoice = Invoice(
            # id is auto-generated (Integer primary key)
            tenant_id=access.tenant_id,
            invoice_number=invoice_number,
            patient_id=patient_id,
            sale_id=sale_id,
            status=(payload.get("status") or "draft"),
            notes=payload.get("notes"),
            device_price=payload.get("total_amount", 0),
            created_by=access.user_id
        )
        
        db_session.add(invoice)
        db_session.commit()
        db_session.refresh(invoice)
        
        return ResponseEnvelope(data=invoice.to_dict())
    except HTTPException:
        raise
    except Exception as e:
        db_session.rollback()
        logger.error(f"Create invoice error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@router.put("/invoices/{invoice_id}", operation_id="updateInvoice", response_model=ResponseEnvelope[InvoiceRead])
def update_invoice(
    invoice_id: int,
    request_data: InvoiceUpdate,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access())
):
    """Update an invoice"""
    try:
        invoice = db_session.query(Invoice).filter(
            Invoice.id == invoice_id,
            Invoice.tenant_id == access.tenant_id
        ).first()
        
        if not invoice or invoice.status == 'deleted':
            raise HTTPException(status_code=404, detail={"message": "Invoice not found", "code": "NOT_FOUND"})
        
        update_data = request_data.model_dump(exclude_unset=True, by_alias=False)
        for key, value in update_data.items():
            if value is None:
                continue
            if hasattr(invoice, key):
                setattr(invoice, key, value)
        
        invoice.updated_at = datetime.utcnow()
        db_session.commit()
        
        return ResponseEnvelope(data=invoice.to_dict())
    except HTTPException:
        raise
    except Exception as e:
        db_session.rollback()
        logger.error(f"Update invoice error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@router.delete("/invoices/{invoice_id}", operation_id="deleteInvoice", response_model=ResponseEnvelope[None])
def delete_invoice(
    invoice_id: int,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access())
):
    """Delete an invoice (soft delete)"""
    try:
        invoice = db_session.query(Invoice).filter(
            Invoice.id == invoice_id,
            Invoice.tenant_id == access.tenant_id
        ).first()
        
        if not invoice:
            raise HTTPException(status_code=404, detail={"message": "Invoice not found", "code": "NOT_FOUND"})
        
        invoice.status = 'deleted'
        invoice.updated_at = datetime.utcnow()
        db_session.commit()
        
        return ResponseEnvelope(message="Invoice deleted successfully")
    except HTTPException:
        raise
    except Exception as e:
        db_session.rollback()
        logger.error(f"Delete invoice error: {e}")
        raise HTTPException(status_code=400, detail=str(e))

@router.get("/patients/{patient_id}/invoices", operation_id="listPatientInvoices", response_model=ResponseEnvelope[List[InvoiceRead]])
def get_patient_invoices(
    patient_id: str,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access())
):
    """Get all invoices for a patient"""
    try:
        patient = db_session.query(Patient).filter(
            Patient.id == patient_id,
            Patient.tenant_id == access.tenant_id
        ).first()
        
        if not patient:
            raise HTTPException(status_code=404, detail={"message": "Patient not found", "code": "NOT_FOUND"})
        
        invoices = db_session.query(Invoice).filter(
            Invoice.patient_id == patient_id,
            Invoice.tenant_id == access.tenant_id,
            Invoice.status != 'deleted'
        ).order_by(desc(Invoice.created_at)).all()
        
        return ResponseEnvelope(data=[inv.to_dict() for inv in invoices])
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get patient invoices error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/invoices/{invoice_id}/send-to-gib", operation_id="createInvoiceSendToGib", response_model=ResponseEnvelope[dict])
def send_to_gib(
    invoice_id: int,
    db_session: Session = Depends(get_db),
    access: UnifiedAccess = Depends(require_access())
):
    """Send invoice to GIB (Generate e-Fatura/e-Arşiv)"""
    try:
        invoice = db_session.query(Invoice).filter(
            Invoice.id == invoice_id,
            Invoice.tenant_id == access.tenant_id
        ).first()
        
        if not invoice:
            raise HTTPException(status_code=404, detail={"message": "Invoice not found", "code": "NOT_FOUND"})

        if getattr(invoice, 'sent_to_gib', False):
             raise HTTPException(status_code=400, detail={"message": "Invoice already sent to GİB", "code": "ALREADY_SENT"})

        # Build minimal xml using invoice metadata
        invoice_meta = {
            'invoiceNumber': getattr(invoice, 'invoice_number', str(invoice.id)),
            'items': getattr(invoice, 'items', [])
        }

        # Mock object for type('X', ...)
        class InvoiceMock:
            id = invoice.id
            
        file_name, ettn, uid, xml_content = build_return_invoice_xml(InvoiceMock(), invoice_meta=invoice_meta)

        outbox = EFaturaOutbox(
            invoice_id=str(invoice.id),
            file_name=file_name,
            ettn=ettn,
            uuid=uid,
            xml_content=xml_content,
            status='pending',
            tenant_id=access.tenant_id # Ensure tenant ownership
        )
        db_session.add(outbox)

        # Attempt to write outbox file to instance folder (non-fatal)
        try:
            # Assuming instance/efatura_outbox path structure
            outbox_dir = os.path.abspath(os.path.join(os.getcwd(), 'instance', 'efatura_outbox'))
            if not os.path.exists(outbox_dir):
                os.makedirs(outbox_dir, exist_ok=True)
                
            write_outbox_file(outbox_dir, file_name, xml_content)
            outbox.status = 'sent'
        except Exception as e:
            logger.warning(f"Failed to write outbox file: {e}")
            outbox.status = 'error'

        # Mark invoice metadata
        invoice.sent_to_gib = True
        # invoice.sent_to_gib_at = datetime.utcnow() # SQLAlchemy DateTime defaults usually handles it or model
        invoice.status = 'sent'
        # invoice.updated_at = datetime.utcnow()

        # Activity log
        try:
            log = ActivityLog(
                user_id=access.user_id or 'system', 
                action='invoice_issued', 
                entity_type='invoice', 
                entity_id=str(invoice.id),
                tenant_id=access.tenant_id
            )
            db_session.add(log)
        except Exception:
            pass

        db_session.commit()

        return ResponseEnvelope(data={'invoice': invoice.to_dict(), 'outbox': outbox.to_dict()})
    except HTTPException:
        raise
    except Exception as e:
        db_session.rollback()
        logger.error(f"Send to GIB error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/invoices/bulk-upload", operation_id="createInvoiceBulkUpload", response_model=ResponseEnvelope[BulkUploadResponse])
async def bulk_upload_invoices(
    file: UploadFile = File(...),
    access: UnifiedAccess = Depends(require_access("invoices.write")),
    db_session: Session = Depends(get_db)
):
    """Bulk upload invoices from CSV/XLSX"""
    try:
        if not access.tenant_id:
             # Allow if super admin sends tenant_id in form? FastAPI UploadFile doesn't mix easily with Form params in same obj usually?
             # Actually I can add `tenant_id: str = Form(...)`.
             # But for now, let's assume tenant context is usually present or required.
             raise HTTPException(status_code=400, detail="Tenant context required")

        filename = (file.filename or '').lower()
        content = await file.read()
        
        rows = []
        headers = []

        def _sanitize_cell(v):
             if v is None: return None
             if not isinstance(v, str): return v
             v = v.strip()
             if v.startswith(('=', '+', '-', '@')): return "'" + v
             return v

        # Parse Logic (Shared mostly)
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                raise HTTPException(status_code=500, detail="Server missing openpyxl dependency")
            try:
                wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None: raise HTTPException(status_code=400, detail="XLSX contains no sheets")
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row: raise HTTPException(status_code=400, detail="Empty sheet")
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        obj[h] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"XLSX Parsing error: {e}")
        else:
             try:
                try: text = content.decode('utf-8-sig')
                except: text = content.decode('utf-8', errors='replace')
                
                try: dialect = csv.Sniffer().sniff(text[:4096]); delimiter = dialect.delimiter
                except: delimiter = ','
                
                try:
                    rows_iter = csv.DictReader(io.StringIO(text), delimiter=delimiter)
                    rows = [r for r in rows_iter]
                except:
                     # Fallback for semi-colon
                     rows = [r for r in csv.DictReader(io.StringIO(text), delimiter=';')]

                # Sanitize
                rows = [{k: _sanitize_cell(v) for k, v in r.items()} for r in rows]
             except Exception as e:
                raise HTTPException(status_code=400, detail=f"CSV Parsing error: {e}")

        created = 0
        updated = 0
        errors = []
        row_num = 0
        
        for row in rows:
            row_num += 1
            try:
                # Normalization
                def get_val(keys):
                    for k in keys:
                         if k in row and row[k]: return row[k]
                    return None
                
                invoice_number = get_val(['invoiceNumber', 'invoice_number', 'no', 'fatura_no'])
                patient_name = get_val(['patientName', 'patient_name', 'hasta', 'isim'])
                
                if not (invoice_number or patient_name):
                    errors.append({'row': row_num, 'error': 'Missing invoice number or patient name'})
                    continue
                
                # Check existing
                existing = None
                if invoice_number:
                    existing = db_session.query(Invoice).filter(
                        Invoice.invoice_number == invoice_number, 
                        Invoice.tenant_id == access.tenant_id
                    ).first()
                
                # Payload prep
                patient_phone = get_val(['patientPhone', 'patient_phone', 'telefon'])
                patient_tc = get_val(['patientTcNumber', 'patient_tc', 'tc'])
                issue_date = get_val(['issueDate', 'issue_date', 'tarih'])
                due_date = get_val(['dueDate', 'due_date', 'vade'])
                grand_total = get_val(['grandTotal', 'grand_total', 'tutar', 'toplam'])
                currency = get_val(['currency', 'pb', 'doviz'])
                
                # Date parsing helper
                def parse_date(d):
                    if not d: return None
                    if isinstance(d, datetime): return d
                    for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%d/%m/%Y'):
                         try: return datetime.strptime(d, fmt)
                         except: pass
                    try: return datetime.fromisoformat(d)
                    except: return None

                i_date = parse_date(issue_date) or datetime.utcnow()
                d_date = parse_date(due_date)

                if existing:
                    existing.patient_name = patient_name or existing.patient_name
                    existing.patient_tc = patient_tc or existing.patient_tc
                    existing.currency = currency or existing.currency
                    if i_date: existing.invoice_date = i_date
                    if d_date: existing.due_date = d_date
                    
                    if grand_total:
                        try: existing.grand_total = float(grand_total); existing.device_price = float(grand_total)
                        except: pass
                    
                    if patient_phone:
                         existing.notes = (existing.notes or '') + "\\nPhone: " + str(patient_phone)
                    
                    updated += 1
                else:
                    new_inv = Invoice(
                        tenant_id=access.tenant_id,
                        invoice_number=invoice_number,
                        patient_name=patient_name,
                        patient_tc=patient_tc,
                        invoice_date=i_date,
                        due_date=d_date,
                        currency=currency,
                        created_by=access.user_id,
                        status='draft'
                    )
                    if grand_total:
                        try: val = float(grand_total); new_inv.grand_total = val; new_inv.device_price = val
                        except: pass
                    
                    if patient_phone:
                        new_inv.notes = f"Phone: {patient_phone}"
                        
                    db_session.add(new_inv)
                    created += 1
                
                # Flush logic
                db_session.begin_nested()
                try:
                    db_session.flush()
                    db_session.commit()
                except:
                    db_session.rollback()
                    raise

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})

        db_session.commit()
        return ResponseEnvelope(data={
            'success': True, 
            'created': created, 
            'updated': updated, 
            'errors': errors,
            'timestamp': datetime.utcnow()
        })
        
    except HTTPException: raise
    except Exception as e:
        db_session.rollback()
        logger.error(f"Bulk invoice upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


