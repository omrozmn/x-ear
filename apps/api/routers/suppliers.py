"""
FastAPI Suppliers Router - Migrated from Flask routes/suppliers/
Supplier CRUD, search, stats, and product relationships
"""
from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from typing import Optional, List, Dict, Any
from pydantic import BaseModel, Field
import logging
import csv
import io

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

from sqlalchemy.orm import Session
from sqlalchemy import or_, func, text
from sqlalchemy.exc import IntegrityError

from schemas.base import ResponseEnvelope, ApiError
from schemas.suppliers import SupplierCreate as SupplierCreateSchema
from schemas.suppliers import SupplierRead, SupplierUpdate as SupplierUpdateSchema, ProductSupplierRead
from schemas.suppliers import SupplierStats, SupplierSearchResponse

from models.suppliers import Supplier, ProductSupplier
from models.tenant import Tenant
from middleware.unified_access import UnifiedAccess, require_access
from database import get_db
from core.database import engine

logger = logging.getLogger(__name__)

router = APIRouter(tags=["Suppliers"])
_SUPPLIER_SCHEMA_ENSURED = False


def _ensure_supplier_schema() -> None:
    global _SUPPLIER_SCHEMA_ENSURED
    if _SUPPLIER_SCHEMA_ENSURED:
        return
    with engine.begin() as connection:
        if connection.dialect.name == "sqlite":
            rows = connection.exec_driver_sql("PRAGMA table_info(suppliers)").fetchall()
            columns = {row[1] for row in rows}
        else:
            rows = connection.execute(text(
                "SELECT column_name FROM information_schema.columns WHERE table_name = 'suppliers'"
            )).fetchall()
            columns = {row[0] for row in rows}
        if "institution_number" not in columns:
            connection.exec_driver_sql("ALTER TABLE suppliers ADD COLUMN institution_number VARCHAR(50)")
    _SUPPLIER_SCHEMA_ENSURED = True

# --- Helper Functions ---

def tenant_scoped_query(access: UnifiedAccess, model, session: Session):
    """Apply tenant scoping to query"""
    query = session.query(model)
    if access.tenant_id:
        query = query.filter_by(tenant_id=access.tenant_id)
    return query

def get_or_404_scoped(session: Session, access: UnifiedAccess, model, record_id):
    """Get record with tenant scoping or raise 404"""
    record = session.get(model, record_id)
    if not record:
        return None
    if access.tenant_id and hasattr(record, 'tenant_id') and record.tenant_id != access.tenant_id:
        return None
    return record

# --- Request Schemas ---

class SupplierCreate(BaseModel):
    company_name: str = Field(..., alias="companyName")
    company_code: Optional[str] = Field(None, alias="companyCode")
    tax_number: Optional[str] = Field(None, alias="taxNumber")
    tax_office: Optional[str] = Field(None, alias="taxOffice")
    institution_number: Optional[str] = Field(None, alias="institutionNumber")
    contact_person: Optional[str] = Field(None, alias="contactPerson")
    email: Optional[str] = None
    phone: Optional[str] = None
    mobile: Optional[str] = None
    fax: Optional[str] = None
    website: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    country: str = "Türkiye"
    postal_code: Optional[str] = Field(None, alias="postalCode")
    payment_terms: Optional[str] = Field(None, alias="paymentTerms")
    currency: str = "TRY"
    rating: Optional[float] = None
    notes: Optional[str] = None
    is_active: bool = Field(True, alias="isActive")
    tenant_id: Optional[str] = Field(None, alias="tenantId")

class SupplierUpdate(BaseModel):
    company_name: Optional[str] = Field(None, alias="companyName")
    company_code: Optional[str] = Field(None, alias="companyCode")
    tax_number: Optional[str] = Field(None, alias="taxNumber")
    tax_office: Optional[str] = Field(None, alias="taxOffice")
    institution_number: Optional[str] = Field(None, alias="institutionNumber")
    contact_person: Optional[str] = Field(None, alias="contactPerson")
    email: Optional[str] = None
    phone: Optional[str] = None
    mobile: Optional[str] = None
    fax: Optional[str] = None
    website: Optional[str] = None
    address: Optional[str] = None
    city: Optional[str] = None
    country: Optional[str] = None
    postal_code: Optional[str] = Field(None, alias="postalCode")
    payment_terms: Optional[str] = Field(None, alias="paymentTerms")
    currency: Optional[str] = None
    rating: Optional[float] = None
    notes: Optional[str] = None
    is_active: Optional[bool] = Field(None, alias="isActive")

# --- Routes ---

@router.get("/suppliers", operation_id="listSuppliers", response_model=ResponseEnvelope[List[SupplierRead]])
def get_suppliers(
    page: int = Query(1, ge=1, le=1000000),
    per_page: int = Query(50, ge=1, le=200),
    search: Optional[str] = None,
    is_active: Optional[bool] = None,
    city: Optional[str] = None,
    sort_by: str = "company_name",
    sort_order: str = "asc",
    access: UnifiedAccess = Depends(require_access()),
    db_session: Session = Depends(get_db)
):
    """Get all suppliers with filtering and pagination"""
    try:
        _ensure_supplier_schema()
        query = tenant_scoped_query(access, Supplier, db_session)
        
        if is_active is not None:
            query = query.filter(Supplier.is_active == is_active)
        
        if city:
            query = query.filter(Supplier.city.ilike(f'%{city}%'))
        
        if search:
            search_term = f'%{search}%'
            query = query.filter(
                or_(
                    Supplier.company_name.ilike(search_term),
                    Supplier.company_code.ilike(search_term),
                    Supplier.tax_number.ilike(search_term),
                    Supplier.contact_person.ilike(search_term),
                    Supplier.email.ilike(search_term),
                    Supplier.phone.ilike(search_term)
                )
            )
            query = query.order_by(Supplier.company_name.asc())
        else:
            if hasattr(Supplier, sort_by):
                order_column = getattr(Supplier, sort_by)
                if sort_order == 'desc':
                    query = query.order_by(order_column.desc())
                else:
                    query = query.order_by(order_column.asc())
        
        total = query.count()
        suppliers = query.offset((page - 1) * per_page).limit(per_page).all()
        
        # Use Pydantic schema for type-safe serialization (NO to_dict())
        return ResponseEnvelope(
            data=[SupplierRead.model_validate(s) for s in suppliers],
            meta={
                "page": page,
                "perPage": per_page,
                "total": total,
                "totalPages": (total + per_page - 1) // per_page
            }
        )
    except Exception as e:
        logger.error(f"Get suppliers error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/suppliers/search", operation_id="listSupplierSearch", response_model=ResponseEnvelope[SupplierSearchResponse])
def search_suppliers(
    q: str = Query("", min_length=0),
    limit: int = Query(10, ge=1, le=50),
    access: UnifiedAccess = Depends(require_access()),
    db_session: Session = Depends(get_db)
):
    """Fast supplier search for autocomplete"""
    try:
        _ensure_supplier_schema()
        if not q or len(q) < 2:
            return ResponseEnvelope(data={"suppliers": []})
        
        search_term = f'%{q}%'
        query = tenant_scoped_query(access, Supplier, db_session).filter(
            Supplier.is_active == True,
            or_(
                Supplier.company_name.ilike(search_term),
                Supplier.company_code.ilike(search_term),
                Supplier.tax_number.ilike(search_term),
                Supplier.contact_person.ilike(search_term)
            )
        ).order_by(Supplier.company_name.asc()).limit(limit)
        
        suppliers = query.all()
        
        # Use Pydantic schema for type-safe serialization (NO to_dict())
        return ResponseEnvelope(data=SupplierSearchResponse(
            suppliers=[SupplierRead.model_validate(s) for s in suppliers]
        ))
    except Exception as e:
        logger.error(f"Search suppliers error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/suppliers/stats", operation_id="listSupplierStats", response_model=ResponseEnvelope[SupplierStats])
def get_supplier_stats(
    access: UnifiedAccess = Depends(require_access()),
    db_session: Session = Depends(get_db)
):
    """Get supplier statistics"""
    try:
        _ensure_supplier_schema()
        base_query = tenant_scoped_query(access, Supplier, db_session)
        
        total_suppliers = base_query.count()
        active_suppliers = base_query.filter_by(is_active=True).count()
        inactive_suppliers = total_suppliers - active_suppliers
        
        # Total product relationships
        if access.tenant_id:
            supplier_ids = [s.id for s in base_query.all()]
            total_relationships = db_session.query(ProductSupplier).filter(
                ProductSupplier.supplier_id.in_(supplier_ids),
                ProductSupplier.is_active == True
            ).count() if supplier_ids else 0
        else:
            total_relationships = db_session.query(ProductSupplier).filter_by(is_active=True).count()
        
        # Average rating
        avg_rating_result = base_query.filter(
            Supplier.is_active == True,
            Supplier.rating.isnot(None)
        ).with_entities(func.avg(Supplier.rating)).scalar()
        
        avg_rating = float(avg_rating_result) if avg_rating_result else 0.0
        
        return ResponseEnvelope(
            data=SupplierStats(
                total_suppliers=total_suppliers,
                active_suppliers=active_suppliers,
                inactive_suppliers=inactive_suppliers,
                total_product_relationships=total_relationships,
                average_rating=round(avg_rating, 1)
            )
        )
    except Exception as e:
        logger.error(f"Get supplier stats error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/suppliers/{supplier_id}", operation_id="getSupplier", response_model=ResponseEnvelope[SupplierRead])
def get_supplier(
    supplier_id: int,
    access: UnifiedAccess = Depends(require_access()),
    db_session: Session = Depends(get_db)
):
    """Get a single supplier by ID"""
    try:
        _ensure_supplier_schema()
        supplier = get_or_404_scoped(db_session, access, Supplier, supplier_id)
        if not supplier:
            raise HTTPException(
                status_code=404,
                detail=ApiError(message="Supplier not found", code="SUPPLIER_NOT_FOUND").model_dump(mode="json")
            )
        
        # Use Pydantic schema for type-safe serialization (NO to_dict())
        # Use Pydantic schema for type-safe serialization (NO to_dict())
        supplier_data = SupplierRead.model_validate(supplier).model_dump(by_alias=True)
        
        products_data = []
        for ps in supplier.products:
            if not ps.is_active:
                continue
            ps_data = ProductSupplierRead.model_validate(ps).model_dump(by_alias=True)
            if ps.product:
                 ps_data['product'] = {
                    'id': ps.product.id,
                    'name': ps.product.name,
                    'sku': ps.product.sku
                }
            products_data.append(ps_data)
            
        supplier_data['products'] = products_data
        
        return ResponseEnvelope(data=supplier_data)
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Get supplier error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/suppliers", operation_id="createSuppliers", status_code=201, response_model=ResponseEnvelope[SupplierRead])
def create_supplier(
    supplier_in: SupplierCreateSchema,
    access: UnifiedAccess = Depends(require_access()),
    db_session: Session = Depends(get_db)
):
    """Create a new supplier"""
    try:
        _ensure_supplier_schema()
        access.tenant_id = access.tenant_id or supplier_in.tenant_id
        
        if not access.tenant_id:
            tenant = db_session.query(Tenant).first()
            if tenant:
                access.tenant_id = tenant.id
        
        if not access.tenant_id:
            raise HTTPException(
                status_code=400,
                detail=ApiError(message="access.tenant_id is required", code="TENANT_REQUIRED").model_dump(mode="json")
            )
        
        # Check for duplicate
        existing = db_session.query(Supplier).filter_by(
            tenant_id=access.tenant_id,
            company_name=supplier_in.company_name
        ).first()
        if existing:
            raise HTTPException(
                status_code=409,
                detail=ApiError(message="Supplier with this company name already exists", code="DUPLICATE").model_dump(mode="json")
            )
        
        data = supplier_in.model_dump(by_alias=False, exclude={'name', 'code', 'contact_name', 'tenant_id'})
        
        # Convert empty strings to None for optional unique fields to avoid UNIQUE constraint violations
        if data.get('company_code') == '':
            data['company_code'] = None
        if data.get('tax_number') == '':
            data['tax_number'] = None
            
        supplier = Supplier(tenant_id=access.tenant_id, **data)
        
        db_session.add(supplier)
        db_session.commit()
        
        logger.info(f"Supplier created: {supplier.id}")
        return ResponseEnvelope(data=supplier)
    except HTTPException:
        raise
    except IntegrityError as e:
        db_session.rollback()
        logger.warning(f"Supplier creation failed - duplicate: {e}")
        # Check if it's a company_name duplicate
        if 'company_name' in str(e):
            raise HTTPException(
                status_code=409,
                detail=ApiError(message="Supplier with this company name already exists", code="DUPLICATE_COMPANY_NAME").model_dump(mode="json")
            )
        # Other integrity errors
        raise HTTPException(
            status_code=409,
            detail=ApiError(message="Duplicate entry", code="DUPLICATE").model_dump(mode="json")
        )
    except Exception as e:
        db_session.rollback()
        logger.error(f"Create supplier error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.put("/suppliers/{supplier_id}", operation_id="updateSupplier", response_model=ResponseEnvelope[SupplierRead])
def update_supplier(
    supplier_id: int,
    supplier_in: SupplierUpdateSchema,
    access: UnifiedAccess = Depends(require_access()),
    db_session: Session = Depends(get_db)
):
    """Update an existing supplier"""
    try:
        _ensure_supplier_schema()
        supplier = get_or_404_scoped(db_session, access, Supplier, supplier_id)
        if not supplier:
            raise HTTPException(
                status_code=404,
                detail=ApiError(message="Supplier not found", code="SUPPLIER_NOT_FOUND").model_dump(mode="json")
            )
        
        data = supplier_in.model_dump(exclude_unset=True, by_alias=False)
        
        # Exclude read-only aliases and other non-model fields
        exclude_fields = {'name', 'code', 'contact_name', 'tenant_id'}
        
        # Convert empty strings to None for optional unique fields to avoid UNIQUE constraint violations
        if data.get('company_code') == '':
            data['company_code'] = None
        if data.get('tax_number') == '':
            data['tax_number'] = None
        
        # Check for duplicate name
        if 'company_name' in data and data['company_name'] != supplier.company_name:
            existing = db_session.query(Supplier).filter_by(
                tenant_id=supplier.tenant_id,
                company_name=data['company_name']
            ).first()
            if existing:
                raise HTTPException(
                    status_code=409,
                    detail=ApiError(message="Supplier with this company name already exists", code="DUPLICATE").model_dump(mode="json")
                )
        
        for key, value in data.items():
            if key not in exclude_fields and hasattr(supplier, key):
                setattr(supplier, key, value)
        
        db_session.commit()
        
        logger.info(f"Supplier updated: {supplier.id}")
        return ResponseEnvelope(data=supplier)
    except HTTPException:
        raise
    except Exception as e:
        db_session.rollback()
        logger.error(f"Update supplier error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.delete("/suppliers/{supplier_id}", operation_id="deleteSupplier", response_model=ResponseEnvelope[None])
def delete_supplier(
    supplier_id: int,
    access: UnifiedAccess = Depends(require_access()),
    db_session: Session = Depends(get_db)
):
    """Delete a supplier (soft delete)"""
    try:
        supplier = get_or_404_scoped(db_session, access, Supplier, supplier_id)
        if not supplier:
            raise HTTPException(
                status_code=404,
                detail=ApiError(message="Supplier not found", code="SUPPLIER_NOT_FOUND").model_dump(mode="json")
            )
        
        # Soft delete
        supplier.is_active = False
        
        # Deactivate product relationships
        for ps in supplier.products:
            ps.is_active = False
        
        db_session.commit()
        
        logger.info(f"Supplier deleted (soft): {supplier.id}")
        return ResponseEnvelope(message="Supplier deactivated successfully")
    except HTTPException:
        raise
    except Exception as e:
        db_session.rollback()
        logger.error(f"Delete supplier error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.post("/suppliers/bulk-upload", operation_id="createSupplierBulkUpload", response_model=ResponseEnvelope[Dict[str, Any]])
async def bulk_upload_suppliers(
    file: UploadFile = File(...),
    access: UnifiedAccess = Depends(require_access("suppliers.manage")),
    db: Session = Depends(get_db)
):
    """Bulk upload suppliers from CSV/XLSX"""
    try:
        effective_tenant = access.effective_tenant_id or access.tenant_id

        if not effective_tenant or effective_tenant == 'system':
            raise HTTPException(status_code=400, detail="Tenant context required")

        filename = (file.filename or '').lower()
        content = await file.read()

        rows = []

        def _sanitize_cell(v):
            if v is None: return None
            if not isinstance(v, str): return v
            v = v.strip()
            if v.startswith(('=', '+', '-', '@')): return "'" + v
            return v

        # Parse File
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if load_workbook is None:
                raise HTTPException(status_code=500, detail="Server missing openpyxl dependency")
            try:
                wb = load_workbook(filename=io.BytesIO(content), read_only=True, data_only=True)
                sheet = wb[wb.sheetnames[0]] if wb.sheetnames else None
                if sheet is None: raise HTTPException(status_code=400, detail="XLSX contains no sheets")
                it = sheet.iter_rows(values_only=True)
                headers_row = next(it, None)
                if not headers_row: raise HTTPException(status_code=400, detail="Empty sheet")
                headers = [str(h).strip() if h is not None else '' for h in headers_row]
                for r in it:
                    obj = {}
                    for idx, h in enumerate(headers):
                        val = r[idx] if idx < len(r) else None
                        obj[h] = _sanitize_cell(val)
                    if any(v not in (None, '') for v in obj.values()):
                        rows.append(obj)
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"XLSX parse error: {e}")
        else:
            try:
                try: text_content = content.decode('utf-8-sig')
                except: text_content = content.decode('utf-8', errors='replace')
                try: dialect = csv.Sniffer().sniff(text_content[:4096]); delimiter = dialect.delimiter
                except: delimiter = ','
                rows = [r for r in csv.DictReader(io.StringIO(text_content), delimiter=delimiter)]
                rows = [{k: _sanitize_cell(v) for k, v in r.items()} for r in rows]
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"CSV parse error: {e}")

        created = 0
        updated = 0
        errors = []
        row_num = 0

        for row in rows:
            row_num += 1
            try:
                def get_val(keys):
                    for k in keys:
                        if k in row and row[k]: return row[k]
                    # Case-insensitive fallback
                    normalized_keys = {k.lower(): k for k in row.keys()}
                    for k in keys:
                        lower_k = k.lower()
                        if lower_k in normalized_keys and row[normalized_keys[lower_k]]:
                            return row[normalized_keys[lower_k]]
                    return None

                company_name = get_val(['companyName', 'company_name', 'sirket', 'Şirket Adı', 'firma'])
                company_code = get_val(['companyCode', 'company_code', 'sirket_kodu', 'Şirket Kodu'])
                phone = get_val(['phone', 'telefon', 'Telefon', 'tel'])

                if not (company_name or company_code or phone):
                    errors.append({'row': row_num, 'error': 'En az bir tanımlayıcı gerekli (şirket adı, kodu veya telefon)'})
                    continue

                # Cascading match: company_code > company_name > phone
                existing = None
                if company_code:
                    existing = db.query(Supplier).filter(
                        Supplier.company_code == company_code,
                        Supplier.tenant_id == effective_tenant
                    ).first()
                if not existing and company_name:
                    existing = db.query(Supplier).filter(
                        Supplier.company_name == company_name,
                        Supplier.tenant_id == effective_tenant
                    ).first()
                if not existing and phone:
                    existing = db.query(Supplier).filter(
                        Supplier.phone == phone,
                        Supplier.tenant_id == effective_tenant
                    ).first()

                payload = {
                    'company_name': company_name,
                    'company_code': company_code,
                    'contact_person': get_val(['contactPerson', 'contact_person', 'yetkili', 'Yetkili Kişi']),
                    'phone': phone,
                    'mobile': get_val(['mobile', 'cep', 'Cep Telefon']),
                    'email': get_val(['email', 'e-mail', 'eposta', 'E-posta']),
                    'tax_number': get_val(['taxNumber', 'tax_number', 'vergi_no', 'Vergi No']),
                    'tax_office': get_val(['taxOffice', 'tax_office', 'vergi_dairesi', 'Vergi Dairesi']),
                    'address': get_val(['address', 'adres', 'Adres']),
                    'city': get_val(['city', 'sehir', 'Şehir', 'il']),
                    'district': get_val(['district', 'ilce', 'İlçe']),
                    'country': get_val(['country', 'ulke', 'Ülke']),
                    'payment_terms': get_val(['paymentTerms', 'payment_terms', 'odeme_vade', 'Ödeme Vadesi']),
                    'currency': get_val(['currency', 'doviz', 'Para Birimi']),
                    'notes': get_val(['notes', 'notlar', 'Notlar']),
                    'website': get_val(['website', 'web', 'Web Sitesi']),
                }

                # Clean None values
                payload = {k: v for k, v in payload.items() if v is not None}

                # Handle is_active
                is_active_val = get_val(['isActive', 'is_active', 'aktif', 'Aktif'])
                if is_active_val is not None:
                    if str(is_active_val).lower() in ['true', '1', 'evet', 'aktif', 'yes']:
                        payload['is_active'] = True
                    elif str(is_active_val).lower() in ['false', '0', 'hayır', 'pasif', 'no']:
                        payload['is_active'] = False

                if existing:
                    # Additive update — only fill empty fields
                    for k, v in payload.items():
                        if hasattr(existing, k) and v is not None:
                            current_val = getattr(existing, k, None)
                            if current_val is None or current_val == '':
                                setattr(existing, k, v)
                    updated += 1
                else:
                    if not company_name:
                        errors.append({'row': row_num, 'error': 'Yeni tedarikçi için şirket adı gerekli'})
                        continue
                    supplier = Supplier(tenant_id=effective_tenant, **payload)
                    db.add(supplier)
                    created += 1

                db.begin_nested()
                try:
                    db.flush()
                    db.commit()
                except Exception as e:
                    db.rollback()
                    errors.append({'row': row_num, 'error': str(e)})

            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})

        db.commit()
        return ResponseEnvelope(data={
            'success': True,
            'created': created,
            'updated': updated,
            'errors': errors
        })

    except HTTPException: raise
    except Exception as e:
        db.rollback()
        logger.error(f"Bulk supplier upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))
